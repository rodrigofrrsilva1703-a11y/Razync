import streamlit as st
# Deploy sync: pesquisa de empresas aprovada em 2026-09-01.
import pandas as pd
import re
import struct
import calendar
import io
import os
import tempfile
import unicodedata
import json
import hashlib
import hmac
import zipfile
import difflib
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime
from pypdf import PdfReader

from razync.companies import CONFIGURACOES_AUTOKRAFT, CONFIGURACOES_ACCEDE
from razync.company_catalog import EMPRESAS_POR_REGIME, EMPRESAS_POR_CHAVE
from razync.nibo import processar_extrato_nibo_pdf
from razync.security import proteger_acesso
from razync.bank_validation import diagnostico_pdf_sem_lancamentos, validar_fechamento_saldo

# Configuração da página Web
st.set_page_config(
    page_title="Razync", 
    page_icon="assets/razync-icon.png", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==============================================================================
# ESTILIZAÇÃO CSS DARK MODE
# ==============================================================================
st.markdown("""
    <style>
        :root {
            --hc-bg: #0b0f13;
            --hc-surface: #111820;
            --hc-surface-hover: #16212b;
            --hc-border: #27333e;
            --hc-border-strong: #3b4b59;
            --hc-text: #f4f7fa;
            --hc-muted: #94a4b3;
            --hc-accent: #13b9e8;
            --hc-accent-soft: rgba(19, 185, 232, 0.12);
        }

        .stApp { background-color: var(--hc-bg); color: var(--hc-text); }
        .block-container { padding-top: 3.25rem; padding-bottom: 3rem; max-width: 100%; }
        h1, h2, h3, h4 { color: var(--hc-text) !important; letter-spacing: -0.02em; }
        hr { border-color: var(--hc-border) !important; }

        .stButton > button {
            width: 100% !important;
            border-radius: 7px !important;
            font-weight: 500 !important;
            padding: 0.48rem 1rem !important;
            border: 1px solid var(--hc-border) !important;
            background-color: var(--hc-surface) !important;
            color: #d9e2ea !important;
            transition: background-color 0.15s ease, border-color 0.15s ease, color 0.15s ease !important;
            box-shadow: none !important;
        }
        .stButton > button:hover {
            background-color: var(--hc-surface-hover) !important;
            border-color: var(--hc-accent) !important;
            color: #ffffff !important;
        }
        .stButton > button:focus-visible {
            outline: 2px solid var(--hc-accent) !important;
            outline-offset: 2px !important;
        }

        .metric-card {
            background-color: var(--hc-surface);
            border: 1px solid var(--hc-border);
            padding: 15px;
            border-radius: 8px;
            text-align: center;
        }
        .metric-title {
            font-size: 11px;
            color: var(--hc-muted);
            text-transform: uppercase;
            font-weight: 600;
            margin-bottom: 5px;
            letter-spacing: 0.5px;
        }
        .metric-value { font-size: 18px; color: var(--hc-text); font-weight: 700; }

        section[data-testid="stSidebar"] {
            background-color: #0e141a;
            border-right: 1px solid var(--hc-border);
        }
        section[data-testid="stSidebar"] .stButton > button {
            text-align: left !important;
            justify-content: flex-start !important;
            background-color: transparent !important;
            border-color: transparent !important;
        }
        section[data-testid="stSidebar"] .stButton > button:hover {
            background-color: var(--hc-accent-soft) !important;
            border-color: rgba(19, 185, 232, 0.35) !important;
        }

        .hc-brand-title {
            color: var(--hc-text);
            font-size: 30px;
            line-height: 1.3;
            font-weight: 750;
            letter-spacing: -0.035em;
            margin-top: 7px;
            padding: 2px 0 3px;
        }
        .hc-brand-subtitle {
            color: var(--hc-muted);
            font-size: 13px;
            line-height: 1.5;
            margin-top: 6px;
        }

        .tool-card {
            background-color: var(--hc-surface);
            border: 1px solid var(--hc-border);
            padding: 24px 20px;
            border-radius: 8px;
            text-align: center;
            height: 160px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            transition: border-color 0.15s ease, background-color 0.15s ease;
        }
        .tool-card:hover {
            border-color: var(--hc-accent);
            background-color: var(--hc-surface-hover);
        }

        .st-key-home_card_extratos button,
        .st-key-home_card_razao button,
        .st-key-home_card_organizador button {
            width: 100% !important;
            height: 184px !important;
            min-height: 184px !important;
            max-height: 184px !important;
            box-sizing: border-box !important;
            overflow: hidden !important;
            padding: 24px 20px !important;
            background: linear-gradient(180deg, rgba(19, 185, 232, 0.035) 0%, var(--hc-surface) 34%) !important;
            border: 1px solid var(--hc-border) !important;
            border-top: 2px solid rgba(19, 185, 232, 0.72) !important;
            border-radius: 9px !important;
            color: var(--hc-muted) !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            text-align: center !important;
            white-space: pre-line !important;
            line-height: 1.55 !important;
            font-size: 12px !important;
            font-weight: 400 !important;
            transition: background-color 0.15s ease, border-color 0.15s ease, transform 0.15s ease, box-shadow 0.15s ease !important;
            box-shadow: 0 8px 24px rgba(0, 0, 0, 0.14) !important;
        }
        .st-key-home_card_extratos button p,
        .st-key-home_card_razao button p,
        .st-key-home_card_organizador button p {
            white-space: pre-line !important;
            margin: 0 !important;
        }
        .st-key-home_card_extratos button strong,
        .st-key-home_card_razao button strong,
        .st-key-home_card_organizador button strong {
            font-size: 18px !important;
            line-height: 1.3 !important;
            font-weight: 700 !important;
            color: var(--hc-text) !important;
        }
        .st-key-home_card_extratos button:hover,
        .st-key-home_card_razao button:hover,
        .st-key-home_card_organizador button:hover {
            background: linear-gradient(180deg, rgba(19, 185, 232, 0.075) 0%, var(--hc-surface-hover) 38%) !important;
            border-color: var(--hc-accent) !important;
            color: #c4d0da !important;
            transform: translateY(-2px);
            box-shadow: 0 12px 30px rgba(0, 0, 0, 0.22) !important;
        }

        /* Cards pequenos para selecionar as empresas dentro do Grupo Autokraft. */
        .st-key-org_autokraft_card_0 button,
        .st-key-org_autokraft_card_1 button,
        .st-key-org_autokraft_card_2 button {
            min-height: 88px !important;
            height: 88px !important;
            padding: 8px 10px !important;
            border-radius: 8px !important;
            white-space: pre-line !important;
            line-height: 1.18 !important;
            font-size: 11px !important;
            text-align: center !important;
            overflow: hidden !important;
            background: #050b12 !important;
            border: 1px solid #12324a !important;
            box-shadow: none !important;
            transform: none !important;
        }
        .st-key-org_autokraft_card_0 button:hover,
        .st-key-org_autokraft_card_1 button:hover,
        .st-key-org_autokraft_card_2 button:hover {
            background: #081725 !important;
            border-color: #1d6f9b !important;
            box-shadow: none !important;
            transform: none !important;
        }
        .st-key-org_autokraft_card_0 button p,
        .st-key-org_autokraft_card_1 button p,
        .st-key-org_autokraft_card_2 button p {
            white-space: pre-line !important;
            margin: 0 !important;
            line-height: 1.18 !important;
        }

        /* Cards de empresas: visual premium, compacto e responsivo. */
        .st-key-org_empresa_card_nova,
        .st-key-org_empresa_card_autokraft_industrial,
        .st-key-org_empresa_card_autokraft_projetos,
        .st-key-org_empresa_card_isa,
        .st-key-org_empresa_card_accede_automacao,
        .st-key-org_empresa_card_accede_equipamentos {
            width: 100% !important;
            min-width: 0 !important;
            max-width: 260px !important;
            display: flex !important;
            justify-content: flex-start !important;
            margin: 0 !important;
        }

        .st-key-org_empresa_card_nova button,
        .st-key-org_empresa_card_autokraft_industrial button,
        .st-key-org_empresa_card_autokraft_projetos button,
        .st-key-org_empresa_card_isa button,
        .st-key-org_empresa_card_accede_automacao button,
        .st-key-org_empresa_card_accede_equipamentos button {
            position: relative !important;
            width: 100% !important;
            min-width: 0 !important;
            max-width: 260px !important;
            height: 154px !important;
            min-height: 154px !important;
            max-height: 154px !important;
            box-sizing: border-box !important;
            overflow: hidden !important;
            padding: 22px 20px 20px !important;
            cursor: pointer !important;
            background:
                radial-gradient(circle at top left, rgba(19, 185, 232, 0.10), transparent 42%),
                linear-gradient(145deg, #07101a 0%, #050a10 72%) !important;
            border: 1px solid rgba(40, 104, 145, 0.50) !important;
            border-top: 2px solid rgba(19, 185, 232, 0.78) !important;
            border-radius: 14px !important;
            color: #f5f9fc !important;
            display: flex !important;
            align-items: flex-end !important;
            justify-content: flex-start !important;
            text-align: left !important;
            white-space: normal !important;
            overflow-wrap: anywhere !important;
            box-shadow: none !important;
            transition:
                background 160ms ease,
                border-color 160ms ease,
                transform 160ms ease !important;
        }

        .st-key-org_empresa_card_nova button::before,
        .st-key-org_empresa_card_autokraft_industrial button::before,
        .st-key-org_empresa_card_autokraft_projetos button::before,
        .st-key-org_empresa_card_isa button::before,
        .st-key-org_empresa_card_accede_automacao button::before,
        .st-key-org_empresa_card_accede_equipamentos button::before {
            content: '' !important;
            position: absolute !important;
            top: 17px !important;
            left: 19px !important;
            width: 28px !important;
            height: 4px !important;
            border-radius: 99px !important;
            background: rgba(19, 185, 232, 0.88) !important;
        }

        .st-key-org_empresa_card_nova button p,
        .st-key-org_empresa_card_autokraft_industrial button p,
        .st-key-org_empresa_card_autokraft_projetos button p,
        .st-key-org_empresa_card_isa button p,
        .st-key-org_empresa_card_accede_automacao button p,
        .st-key-org_empresa_card_accede_equipamentos button p {
            width: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
            white-space: normal !important;
            overflow-wrap: anywhere !important;
            line-height: 1.24 !important;
        }

        .st-key-org_empresa_card_nova button strong,
        .st-key-org_empresa_card_autokraft_industrial button strong,
        .st-key-org_empresa_card_autokraft_projetos button strong,
        .st-key-org_empresa_card_isa button strong,
        .st-key-org_empresa_card_accede_automacao button strong,
        .st-key-org_empresa_card_accede_equipamentos button strong {
            display: block !important;
            width: 100% !important;
            color: #f5f9fc !important;
            font-size: 17px !important;
            line-height: 1.22 !important;
            font-weight: 730 !important;
            letter-spacing: -0.018em !important;
        }

        .st-key-org_empresa_card_nova button:hover,
        .st-key-org_empresa_card_autokraft_industrial button:hover,
        .st-key-org_empresa_card_autokraft_projetos button:hover,
        .st-key-org_empresa_card_isa button:hover,
        .st-key-org_empresa_card_accede_automacao button:hover,
        .st-key-org_empresa_card_accede_equipamentos button:hover {
            background:
                radial-gradient(circle at top left, rgba(19, 185, 232, 0.16), transparent 44%),
                linear-gradient(145deg, #091522 0%, #06101a 72%) !important;
            border-color: rgba(19, 185, 232, 0.90) !important;
            transform: translateY(-2px) !important;
            box-shadow: none !important;
        }

        @media (max-width: 1280px) {
            .st-key-org_empresa_card_nova,
            .st-key-org_empresa_card_autokraft_industrial,
            .st-key-org_empresa_card_autokraft_projetos,
            .st-key-org_empresa_card_isa,
            .st-key-org_empresa_card_nova button,
            .st-key-org_empresa_card_autokraft_industrial button,
            .st-key-org_empresa_card_autokraft_projetos button,
            .st-key-org_empresa_card_isa button,
        .st-key-org_empresa_card_accede_automacao button,
        .st-key-org_empresa_card_accede_equipamentos button {
                max-width: 230px !important;
            }
        }

        @media (max-width: 1050px) {
            .st-key-org_empresa_card_nova,
            .st-key-org_empresa_card_autokraft_industrial,
            .st-key-org_empresa_card_autokraft_projetos,
            .st-key-org_empresa_card_isa,
            .st-key-org_empresa_card_nova button,
            .st-key-org_empresa_card_autokraft_industrial button,
            .st-key-org_empresa_card_autokraft_projetos button,
            .st-key-org_empresa_card_isa button,
        .st-key-org_empresa_card_accede_automacao button,
        .st-key-org_empresa_card_accede_equipamentos button {
                max-width: 100% !important;
            }
        }

        [data-testid="stFileUploaderDropzone"] {
            background-color: var(--hc-surface) !important;
            border: 1px dashed var(--hc-border-strong) !important;
            border-radius: 8px !important;
        }
        [data-testid="stFileUploaderDropzone"]:hover {
            border-color: var(--hc-accent) !important;
            background-color: var(--hc-surface-hover) !important;
        }
        [data-testid="stDataFrame"] {
            border: 1px solid var(--hc-border);
            border-radius: 8px;
            overflow: hidden;
        }

        .alerta-dominio {
            background-color: #2b1719;
            border-left: 4px solid #f05d66;
            padding: 16px;
            border-radius: 6px;
            margin-bottom: 20px;
        }
        .alerta-dominio h4 { margin-top: 0; color: #ff7b83; font-size: 16px; }
        .alerta-dominio p { margin-bottom: 0; color: #d7dde3; font-size: 14px; }
        .aviso-banner {
            background-color: var(--hc-surface);
            border: 1px solid var(--hc-border);
            padding: 12px 16px;
            border-radius: 7px;
            margin-bottom: 20px;
        }
        .aviso-banner p { margin: 0; color: #c5d0da; font-size: 14px; }

        /* Textos explicativos das ferramentas: mais visíveis e fáceis de localizar. */
        [data-testid="stCaptionContainer"] {
            margin: 8px 0 14px !important;
            padding: 10px 13px !important;
            border-left: 3px solid rgba(19, 185, 232, 0.78) !important;
            border-radius: 6px !important;
            background: rgba(19, 185, 232, 0.065) !important;
        }
        [data-testid="stCaptionContainer"] p {
            color: #d6e4ee !important;
            font-size: 16px !important;
            line-height: 1.6 !important;
            font-weight: 550 !important;
            margin: 0 !important;
        }

        /* Hierarquia visual das ferramentas: mais clara sem deixar o app pesado. */
        [data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 8px !important;
            border-bottom: 1px solid var(--hc-border) !important;
            margin-bottom: 12px !important;
        }
        [data-testid="stTabs"] button[role="tab"] {
            min-height: 42px !important;
            padding: 8px 14px !important;
            border-radius: 8px 8px 0 0 !important;
            color: #aab9c6 !important;
            font-weight: 600 !important;
        }
        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #eef8ff !important;
            background: rgba(19, 185, 232, 0.08) !important;
            border-bottom: 2px solid var(--hc-accent) !important;
        }
        [data-testid="stFileUploader"] {
            margin: 10px 0 14px !important;
        }
        [data-testid="stFileUploaderDropzone"] {
            min-height: 92px !important;
            padding: 14px !important;
        }
        [data-testid="stMetric"] {
            background: rgba(17, 24, 32, 0.72) !important;
            border: 1px solid var(--hc-border) !important;
            border-radius: 8px !important;
            padding: 12px 14px !important;
        }
        .hc-review-box {
            margin: 14px 0 12px;
            padding: 14px 16px;
            border: 1px solid rgba(19, 185, 232, 0.28);
            border-left: 4px solid var(--hc-accent);
            border-radius: 8px;
            background: rgba(19, 185, 232, 0.055);
        }
        .hc-review-title {
            color: #eef8ff;
            font-size: 17px;
            font-weight: 700;
            margin-bottom: 4px;
        }
        .hc-review-text {
            color: #b9cad7;
            font-size: 13.5px;
            line-height: 1.5;
        }
        .hc-step-badge {
            display: inline-block;
            padding: 4px 9px;
            margin: 0 6px 6px 0;
            border-radius: 999px;
            background: #0a1722;
            border: 1px solid #1c4057;
            color: #cbe9f6;
            font-size: 11px;
            font-weight: 650;
        }

        .stTextInput { margin-top: -2px; }

        @keyframes hc-loader-enter {
            from { opacity: 0; transform: translateY(5px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @keyframes hc-loader-line {
            0% { transform: translateX(-140%); }
            100% { transform: translateX(430%); }
        }
        [data-testid="stSpinner"] {
            position: relative;
            overflow: hidden;
            width: 100%;
            box-sizing: border-box;
            padding: 14px 16px !important;
            margin: 8px 0 12px;
            border: 1px solid var(--hc-border);
            border-radius: 8px;
            background: linear-gradient(90deg, rgba(19, 185, 232, 0.055), var(--hc-surface) 45%);
            box-shadow: 0 8px 22px rgba(0, 0, 0, 0.12);
            animation: hc-loader-enter 180ms ease-out both;
        }
        [data-testid="stSpinner"]::after {
            content: "";
            position: absolute;
            left: 0;
            bottom: 0;
            width: 26%;
            height: 2px;
            border-radius: 99px;
            background: linear-gradient(90deg, transparent, var(--hc-accent), transparent);
            box-shadow: 0 0 12px rgba(19, 185, 232, 0.45);
            animation: hc-loader-line 1.25s cubic-bezier(0.4, 0, 0.2, 1) infinite;
        }
        [data-testid="stSpinner"] p {
            color: #d9e4ec !important;
            font-size: 13px !important;
            font-weight: 500 !important;
        }
        [data-testid="stSpinner"] svg {
            color: var(--hc-accent) !important;
        }

        /* A animação ocorre somente no primeiro render após trocar de tela. */
        @keyframes hc-page-enter {
            0% {
                opacity: 0;
                transform: translate3d(0, 16px, 0) scale(0.992);
            }
            62% {
                opacity: 1;
            }
            100% {
                opacity: 1;
                transform: translate3d(0, 0, 0) scale(1);
            }
        }
        .hc-page-transition-marker {
            display: none !important;
        }
        .block-container:has(.hc-page-transition-marker) {
            animation: hc-page-enter 360ms cubic-bezier(0.16, 1, 0.3, 1) both;
            transform-origin: 50% 18%;
            backface-visibility: hidden;
            will-change: opacity, transform;
        }

        /* Microinterações discretas deixam os controles mais responsivos. */
        [data-baseweb="tab"],
        [data-testid="stExpander"],
        [data-testid="stFileUploaderDropzone"],
        [data-testid="stDataFrame"] {
            transition: border-color 180ms ease, background-color 180ms ease,
                        color 180ms ease, box-shadow 220ms ease !important;
        }
        [data-baseweb="tab"]:hover {
            color: var(--hc-text) !important;
        }
        .stButton > button:active {
            transform: translateY(1px) scale(0.986) !important;
            transition-duration: 80ms !important;
        }
        @media (prefers-reduced-motion: reduce) {
            .block-container:has(.hc-page-transition-marker) {
                animation: none !important;
                filter: none !important;
            }
            [data-testid="stSpinner"]::after {
                animation: none !important;
            }
            .stButton > button,
            [data-baseweb="tab"],
            [data-testid="stExpander"],
            [data-testid="stFileUploaderDropzone"] {
                transition-duration: 0.01ms !important;
            }
        }
    </style>
""", unsafe_allow_html=True)


# Camada visual unificada da reforma 2026.
st.markdown("""
<style>
/* ---------- Razync Design System 2026 ---------- */
:root {
    --rz-bg: var(--hc-bg);
    --rz-panel: var(--hc-surface);
    --rz-panel-hover: var(--hc-surface-hover);
    --rz-line: var(--hc-border);
    --rz-line-strong: var(--hc-border-strong);
    --rz-text: var(--hc-text);
    --rz-muted: var(--hc-muted);
    --rz-accent: var(--hc-accent);
    --rz-accent-soft: var(--hc-accent-soft);
    --rz-success: #22a06b;
    --rz-warning: #d89b2b;
    --rz-danger: #d95757;
    --rz-radius-sm: 8px;
    --rz-radius-md: 12px;
    --rz-radius-lg: 16px;
    --rz-shadow: 0 12px 34px rgba(2, 8, 23, 0.10);
}

.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background:
        radial-gradient(circle at 76% -12%, rgba(19, 185, 232, 0.055), transparent 28rem),
        var(--rz-bg) !important;
}

.block-container {
    width: min(100%, 1500px) !important;
    max-width: 1500px !important;
    padding: 2.15rem clamp(1.15rem, 2.6vw, 3rem) 4rem !important;
}

h1 {
    font-size: clamp(1.75rem, 3vw, 2.45rem) !important;
    line-height: 1.12 !important;
    font-weight: 720 !important;
    letter-spacing: -0.038em !important;
}
h2 { font-size: clamp(1.35rem, 2vw, 1.75rem) !important; }
h3 { font-size: 1.16rem !important; }
h4, h5 { letter-spacing: -0.012em !important; }

[data-testid="stCaptionContainer"] {
    padding: 0 !important;
    background: transparent !important;
    border-left: 0 !important;
}
[data-testid="stCaptionContainer"] p {
    color: var(--rz-muted) !important;
    font-size: 0.87rem !important;
    line-height: 1.55 !important;
}

/* ---------- Sidebar ---------- */
section[data-testid="stSidebar"] {
    width: 272px !important;
    background: color-mix(in srgb, var(--rz-panel) 94%, var(--rz-bg)) !important;
    border-right: 1px solid var(--rz-line) !important;
}
section[data-testid="stSidebar"] > div:first-child {
    padding: 1.15rem 0.9rem !important;
}
.rz-sidebar-brand {
    display: flex;
    align-items: center;
    gap: 0.72rem;
    margin: 0.1rem 0 1.15rem;
    padding: 0.25rem 0.35rem;
}
.rz-sidebar-brand-copy strong {
    display: block;
    color: var(--rz-text);
    font-size: 0.98rem;
    letter-spacing: -0.02em;
}
.rz-sidebar-brand-copy span {
    color: var(--rz-muted);
    font-size: 0.69rem;
    letter-spacing: 0.055em;
    text-transform: uppercase;
}
.rz-nav-label {
    color: var(--rz-muted);
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 0.105em;
    text-transform: uppercase;
    padding: 0 0.65rem;
    margin: 0.3rem 0 0.4rem;
}
section[data-testid="stSidebar"] .stButton > button {
    min-height: 2.55rem !important;
    padding: 0.48rem 0.72rem !important;
    margin-bottom: 0.18rem !important;
    border: 1px solid transparent !important;
    border-radius: 9px !important;
    background: transparent !important;
    color: var(--rz-muted) !important;
    font-size: 0.84rem !important;
    font-weight: 540 !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: var(--rz-accent-soft) !important;
    border-color: color-mix(in srgb, var(--rz-accent) 24%, transparent) !important;
    color: var(--rz-text) !important;
    transform: none !important;
}
section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: var(--rz-accent-soft) !important;
    border-color: color-mix(in srgb, var(--rz-accent) 34%, transparent) !important;
    color: var(--rz-accent) !important;
}
section[data-testid="stSidebar"] [data-testid="stRadio"] {
    padding: 0.65rem 0.55rem 0.2rem;
    margin-top: 0.65rem;
    border-top: 1px solid var(--rz-line);
}
section[data-testid="stSidebar"] [data-testid="stRadio"] label p {
    font-size: 0.74rem !important;
}

/* ---------- Cabeçalhos ---------- */
.rz-page-header {
    margin: 0 0 1.45rem;
    padding: 0 0 1.15rem;
    border-bottom: 1px solid var(--rz-line);
}
.rz-page-kicker {
    color: var(--rz-accent);
    font-size: 0.67rem;
    font-weight: 760;
    letter-spacing: 0.115em;
    text-transform: uppercase;
    margin-bottom: 0.38rem;
}
.rz-page-title {
    color: var(--rz-text);
    font-size: clamp(1.65rem, 3vw, 2.35rem);
    line-height: 1.15;
    font-weight: 730;
    letter-spacing: -0.038em;
    margin: 0;
}
.rz-page-description {
    max-width: 780px;
    color: var(--rz-muted);
    font-size: 0.9rem;
    line-height: 1.6;
    margin: 0.42rem 0 0;
}

/* ---------- Home ---------- */
.rz-home-hero {
    display: grid;
    grid-template-columns: minmax(0, 1fr) auto;
    align-items: end;
    gap: 1.5rem;
    padding: 0.45rem 0 1.55rem;
    margin-bottom: 1rem;
    border-bottom: 1px solid var(--rz-line);
}
.rz-home-eyebrow {
    color: var(--rz-accent);
    font-size: 0.68rem;
    font-weight: 760;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
}
.rz-home-title {
    color: var(--rz-text);
    font-size: clamp(2rem, 4vw, 3.15rem);
    font-weight: 735;
    line-height: 1.02;
    letter-spacing: -0.052em;
}
.rz-home-copy {
    max-width: 680px;
    color: var(--rz-muted);
    font-size: 0.94rem;
    line-height: 1.65;
    margin-top: 0.65rem;
}
.rz-home-status {
    color: var(--rz-muted);
    font-size: 0.72rem;
    white-space: nowrap;
    padding-bottom: 0.25rem;
}
.rz-home-status::before {
    content: "";
    display: inline-block;
    width: 7px;
    height: 7px;
    margin-right: 0.45rem;
    border-radius: 50%;
    background: #22a06b;
    box-shadow: 0 0 0 4px rgba(34,160,107,.10);
}
.rz-section-label {
    color: var(--rz-muted);
    font-size: 0.68rem;
    font-weight: 720;
    letter-spacing: 0.105em;
    text-transform: uppercase;
    margin: 1.2rem 0 0.65rem;
}

.st-key-home_card_extratos button,
.st-key-home_card_razao button,
.st-key-home_card_organizador button {
    height: 126px !important;
    min-height: 126px !important;
    max-height: 126px !important;
    padding: 1.05rem 1.1rem !important;
    align-items: flex-start !important;
    justify-content: flex-end !important;
    text-align: left !important;
    background: var(--rz-panel) !important;
    border: 1px solid var(--rz-line) !important;
    border-top: 1px solid var(--rz-line) !important;
    border-radius: var(--rz-radius-md) !important;
    box-shadow: none !important;
    color: var(--rz-muted) !important;
    font-size: 0.78rem !important;
    line-height: 1.45 !important;
}
.st-key-home_card_extratos button:hover,
.st-key-home_card_razao button:hover,
.st-key-home_card_organizador button:hover {
    background: var(--rz-panel-hover) !important;
    border-color: color-mix(in srgb, var(--rz-accent) 48%, var(--rz-line)) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 10px 24px rgba(2,8,23,.08) !important;
}
.st-key-home_card_extratos button strong,
.st-key-home_card_razao button strong,
.st-key-home_card_organizador button strong {
    color: var(--rz-text) !important;
    font-size: 1rem !important;
    font-weight: 650 !important;
}

/* ---------- Componentes de trabalho ---------- */
[data-testid="stFileUploaderDropzone"] {
    min-height: 116px !important;
    padding: 1rem !important;
    background: color-mix(in srgb, var(--rz-panel) 72%, transparent) !important;
    border: 1px dashed var(--rz-line-strong) !important;
    border-radius: var(--rz-radius-md) !important;
    box-shadow: none !important;
}
[data-testid="stFileUploaderDropzone"]:hover {
    background: var(--rz-accent-soft) !important;
    border-color: var(--rz-accent) !important;
}
[data-testid="stFileUploaderDropzone"] button {
    min-height: 2.25rem !important;
    border-radius: var(--rz-radius-sm) !important;
}

.stTextInput input,
.stNumberInput input,
.stDateInput input,
[data-baseweb="select"] > div {
    min-height: 2.7rem !important;
    border-radius: var(--rz-radius-sm) !important;
    border-color: var(--rz-line) !important;
    background: var(--rz-panel) !important;
    box-shadow: none !important;
}
.stTextInput input:focus,
.stNumberInput input:focus,
.stDateInput input:focus {
    border-color: var(--rz-accent) !important;
    box-shadow: 0 0 0 3px var(--rz-accent-soft) !important;
}

[data-baseweb="tab-list"] {
    gap: 0.3rem !important;
    padding: 0.22rem !important;
    border: 1px solid var(--rz-line);
    border-radius: 10px;
    background: color-mix(in srgb, var(--rz-panel) 72%, transparent);
}
[data-baseweb="tab"] {
    min-height: 2.5rem !important;
    padding: 0.45rem 0.75rem !important;
    border-radius: 7px !important;
    font-size: 0.8rem !important;
}
[data-baseweb="tab"][aria-selected="true"] {
    background: var(--rz-panel) !important;
    color: var(--rz-text) !important;
}

[data-testid="stExpander"] {
    border: 1px solid var(--rz-line) !important;
    border-radius: var(--rz-radius-md) !important;
    background: color-mix(in srgb, var(--rz-panel) 70%, transparent) !important;
    box-shadow: none !important;
}
[data-testid="stExpander"] summary {
    min-height: 2.9rem;
}

[data-testid="stDataFrame"] {
    border: 1px solid var(--rz-line) !important;
    border-radius: var(--rz-radius-md) !important;
    overflow: hidden;
    box-shadow: none !important;
}

[data-testid="stMetric"] {
    min-height: 96px;
    padding: 0.85rem 0.95rem;
    border: 1px solid var(--rz-line);
    border-radius: var(--rz-radius-md);
    background: var(--rz-panel);
}
[data-testid="stMetricLabel"] p {
    color: var(--rz-muted) !important;
    font-size: 0.71rem !important;
    font-weight: 680 !important;
    letter-spacing: 0.045em;
    text-transform: uppercase;
}
[data-testid="stMetricValue"] {
    color: var(--rz-text) !important;
    font-size: 1.35rem !important;
}

.metric-card {
    min-height: 96px;
    padding: 0.85rem 0.95rem !important;
    display: flex;
    flex-direction: column;
    justify-content: center;
    border-radius: var(--rz-radius-md) !important;
    background: var(--rz-panel) !important;
    border: 1px solid var(--rz-line) !important;
    box-shadow: none !important;
}
.metric-title { font-size: 0.66rem !important; letter-spacing: .055em !important; }
.metric-value { font-size: 1.22rem !important; }

.stAlert {
    border: 1px solid var(--rz-line) !important;
    border-radius: var(--rz-radius-md) !important;
    box-shadow: none !important;
}
.aviso-banner {
    margin: 0.8rem 0 1.1rem !important;
    padding: 0.75rem 0.9rem !important;
    border: 1px solid var(--rz-line) !important;
    border-left: 3px solid var(--rz-accent) !important;
    border-radius: 0 var(--rz-radius-sm) var(--rz-radius-sm) 0 !important;
    background: color-mix(in srgb, var(--rz-panel) 75%, transparent) !important;
}
.aviso-banner p { margin: 0 !important; color: var(--rz-muted) !important; }

.stDownloadButton > button {
    min-height: 2.7rem !important;
    border-radius: var(--rz-radius-sm) !important;
    background: var(--rz-accent) !important;
    border-color: var(--rz-accent) !important;
    color: #03131a !important;
    font-weight: 650 !important;
}
.stDownloadButton > button:hover {
    filter: brightness(1.05);
    transform: translateY(-1px) !important;
}

.stButton > button {
    min-height: 2.55rem;
    border-radius: var(--rz-radius-sm) !important;
    font-size: 0.84rem !important;
}
.stButton > button:focus-visible,
.stDownloadButton > button:focus-visible,
input:focus-visible {
    outline: 2px solid var(--rz-accent) !important;
    outline-offset: 2px !important;
}

hr {
    margin: 1.35rem 0 !important;
    border-color: var(--rz-line) !important;
}

/* ---------- Responsividade ---------- */
@media (max-width: 900px) {
    .block-container { padding: 1.45rem 1rem 3rem !important; }
    .rz-home-hero { grid-template-columns: 1fr; gap: 0.5rem; }
    .rz-home-status { white-space: normal; }
    [data-testid="stHorizontalBlock"] {
        gap: 0.75rem !important;
    }
}
@media (max-width: 640px) {
    .block-container { padding-top: 1rem !important; }
    .rz-page-title { font-size: 1.65rem; }
    .rz-home-title { font-size: 2rem; }
    .st-key-home_card_extratos button,
    .st-key-home_card_razao button,
    .st-key-home_card_organizador button {
        height: 105px !important;
        min-height: 105px !important;
        max-height: 105px !important;
    }
    [data-baseweb="tab-list"] {
        overflow-x: auto !important;
        flex-wrap: nowrap !important;
    }
    [data-baseweb="tab"] { white-space: nowrap !important; }
    [data-testid="stDataFrame"] { max-width: calc(100vw - 2rem) !important; }
}
@media (prefers-reduced-motion: reduce) {
    *, *::before, *::after {
        scroll-behavior: auto !important;
        animation-duration: 0.01ms !important;
        transition-duration: 0.01ms !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# CONTROLE DE ACESSO
# ==============================================================================
# Implementação isolada em razync/security.py para facilitar manutenção e testes.
SEGURANCA_POR_SENHA_ATIVA = proteger_acesso()


if "tema_razync" not in st.session_state:
    st.session_state["tema_razync"] = "Escuro"
with st.sidebar:
    col_logo_sidebar, col_marca_sidebar = st.columns([0.24, 0.76], gap="small")
    with col_logo_sidebar:
        st.image("assets/razync-icon.png", width=46)
    with col_marca_sidebar:
        st.markdown(
            """
            <div class="rz-sidebar-brand-copy">
                <strong>Razync</strong>
                <span>Operações contábeis</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
    tema = st.radio(
        "Aparência",
        ["Escuro", "Claro"],
        horizontal=True,
        key="tema_razync_radio",
    )
if tema != st.session_state["tema_razync"]:
    st.session_state["tema_razync"] = tema
if st.session_state["tema_razync"] == "Claro":
    st.markdown("""<style>
    :root {
        --hc-bg: #f4f7fb;
        --hc-surface: #ffffff;
        --hc-surface-hover: #f0f5f9;
        --hc-border: #d7e1e9;
        --hc-border-strong: #b9c8d4;
        --hc-text: #17212b;
        --hc-muted: #607181;
        --hc-accent: #0784b8;
        --hc-accent-soft: rgba(7,132,184,.09);
    }

    .stApp,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"] {
        background: #f4f7fb !important;
        color: #17212b !important;
    }
    .block-container { background: transparent !important; }

    section[data-testid="stSidebar"] {
        background: #eef3f7 !important;
        border-right: 1px solid #d4dfe8 !important;
    }
    section[data-testid="stSidebar"] * { color: #23313d; }
    section[data-testid="stSidebar"] .stButton > button {
        color: #314250 !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: #e2edf4 !important;
        border-color: #a8c7d9 !important;
    }

    h1,h2,h3,h4,h5,h6,
    .hc-brand-title,
    .metric-value,
    .hc-review-title {
        color: #17212b !important;
    }
    p, label, .stMarkdown, .stText, [data-testid="stWidgetLabel"] {
        color: #344552;
    }
    .hc-brand-subtitle,
    .metric-title,
    .hc-review-text { color: #657685 !important; }

    .stButton > button {
        background: #ffffff !important;
        color: #253746 !important;
        border-color: #cfdbe4 !important;
        box-shadow: 0 1px 2px rgba(31, 49, 64, .04) !important;
    }
    .stButton > button:hover {
        background: #f3f8fb !important;
        color: #102532 !important;
        border-color: #63a9c8 !important;
        box-shadow: 0 3px 10px rgba(37, 77, 99, .08) !important;
    }

    [data-testid="stCaptionContainer"] {
        background: #edf7fb !important;
        border-left-color: #0784b8 !important;
    }
    [data-testid="stCaptionContainer"] p {
        color: #3c5666 !important;
    }

    [data-testid="stFileUploaderDropzone"],
    [data-testid="stMetric"],
    .metric-card,
    .aviso-banner,
    .hc-review-box,
    [data-testid="stExpander"] {
        background: #ffffff !important;
        border-color: #d5e0e8 !important;
        color: #17212b !important;
        box-shadow: 0 1px 3px rgba(31, 49, 64, .035) !important;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        background: #f5fafc !important;
        border-color: #70b2cf !important;
    }

    input, textarea,
    [data-baseweb="select"] > div,
    [data-baseweb="input"] > div {
        background: #ffffff !important;
        color: #17212b !important;
        border-color: #cad7e1 !important;
    }
    input::placeholder, textarea::placeholder { color: #8797a4 !important; }

    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        border-bottom-color: #d5e0e8 !important;
    }
    [data-testid="stTabs"] button[role="tab"] {
        color: #617382 !important;
        background: transparent !important;
    }
    [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
        color: #075f84 !important;
        background: #eaf5fa !important;
        border-bottom-color: #0784b8 !important;
    }

    [data-testid="stDataFrame"],
    [data-testid="stTable"] {
        background: #ffffff !important;
        border-color: #d5e0e8 !important;
    }

    .hc-step-badge {
        background: #edf6fa !important;
        border-color: #b8d6e4 !important;
        color: #28566b !important;
    }

    .st-key-org_empresa_card_nova button,
    .st-key-org_empresa_card_autokraft_industrial button,
    .st-key-org_empresa_card_autokraft_projetos button,
    .st-key-org_empresa_card_isa button,
        .st-key-org_empresa_card_accede_automacao button,
        .st-key-org_empresa_card_accede_equipamentos button {
        background: linear-gradient(145deg, #ffffff 0%, #f3f8fb 100%) !important;
        border-color: #bfd5e1 !important;
        border-top-color: #0784b8 !important;
        color: #17212b !important;
        box-shadow: 0 4px 14px rgba(38, 74, 95, .06) !important;
    }
    .st-key-org_empresa_card_nova button strong,
    .st-key-org_empresa_card_autokraft_industrial button strong,
    .st-key-org_empresa_card_autokraft_projetos button strong,
    .st-key-org_empresa_card_isa button strong,
        .st-key-org_empresa_card_accede_automacao button strong,
        .st-key-org_empresa_card_accede_equipamentos button strong {
        color: #17212b !important;
    }
    .st-key-org_empresa_card_nova button:hover,
    .st-key-org_empresa_card_autokraft_industrial button:hover,
    .st-key-org_empresa_card_autokraft_projetos button:hover,
    .st-key-org_empresa_card_isa button:hover,
        .st-key-org_empresa_card_accede_automacao button:hover,
        .st-key-org_empresa_card_accede_equipamentos button:hover {
        background: linear-gradient(145deg, #ffffff 0%, #eaf5fa 100%) !important;
        border-color: #5ea7c7 !important;
        box-shadow: 0 7px 18px rgba(38, 74, 95, .09) !important;
    }

    .st-key-ng_card_matriz button,
    .st-key-ng_card_filial button,
    .st-key-ng_card_matriz_ativo button,
    .st-key-ng_card_filial_ativo button {
        background: #ffffff !important;
        color: #263b49 !important;
        border-color: #c6d7e2 !important;
    }
    .st-key-ng_card_matriz_ativo button,
    .st-key-ng_card_filial_ativo button {
        background: #e9f5fb !important;
        border-color: #208bb7 !important;
        color: #164f68 !important;
    }

    [data-testid="stSpinner"] {
        background: linear-gradient(90deg, #edf7fb, #ffffff 45%) !important;
        border-color: #d2e0e8 !important;
        box-shadow: none !important;
    }
    [data-testid="stSpinner"] p { color: #405766 !important; }

    .alerta-dominio {
        background: #fff2f3 !important;
        border-left-color: #d94755 !important;
    }
    .alerta-dominio h4 { color: #a92c38 !important; }
    .alerta-dominio p { color: #62474b !important; }
    </style>""", unsafe_allow_html=True)

# ==============================================================================
# FUNÇÕES DE LIMPEZA E FORMATAÇÃO (MECÂNICAS)
# ==============================================================================
def limpar_caracteres_ilegais(val):
    if isinstance(val, str): return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', val)
    return val

def normalizar_texto(texto):
    if not isinstance(texto, str): return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').lower()

def formatar_moeda(valor):
    try: return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except: return "R$ 0,00"

def formatar_dataframe_moeda_br(df, colunas):
    """Formata apenas a cópia exibida; os dados originais continuam numéricos."""
    exibicao = df.copy()
    for coluna in colunas:
        if coluna in exibicao.columns:
            exibicao[coluna] = exibicao[coluna].apply(
                lambda valor: formatar_moeda(valor) if pd.notna(valor) and valor != '' else ''
            )
    return exibicao

def executar_com_loading(mensagem, funcao, *args, **kwargs):
    """Exibe o carregamento visual apenas enquanto a operação estiver executando."""
    with st.spinner(mensagem):
        return funcao(*args, **kwargs)

def sanitizar_dataframe(df):
    for col in df.select_dtypes(include=['object', 'string']).columns: df[col] = df[col].apply(limpar_caracteres_ilegais)
    return df

def interpretar_sinal_inteligente(historico_str, valor_num, explicit_nature=""):
    """
    Motor Inteligente Universal:
    Avalia se o lançamento é Entrada (+) ou Saída (-) com base em indicadores
    explícitos e análise semântica avançada do histórico.
    """
    val = abs(float(valor_num))
    ind = normalizar_texto(str(explicit_nature)).strip()
    ind_tokens = set(re.findall(r'[a-z]+', ind))
    h_norm = normalizar_texto(historico_str)
    
    # 1. Indicador explícito de natureza (C/D, Crédito/Débito, Entrada/Saída)
    natureza_debito = (
        ind in {'d', 'deb', 'db', 'debito', 'saida'} or
        bool(ind_tokens.intersection({'debito', 'saida', 'pagamento', 'pagto', 'emitido'}))
    )
    natureza_credito = (
        ind in {'c', 'cred', 'cr', 'credito', 'entrada'} or
        bool(ind_tokens.intersection({'credito', 'entrada', 'recebimento', 'recebido'}))
    )
    if natureza_debito:
        return -val
    if natureza_credito:
        return val
        
    # 2. Se o valor já veio negativo do arquivo original
    if valor_num < 0:
        return -val

    # 3. Indicadores inequívocos de entrada têm prioridade sobre termos como
    # "aplic" que podem aparecer no complemento de um rendimento.
    termos_entrada = [
        'pix recebido', 'ted recebida', 'ted recebido', 'recebimento',
        'recebimentos', 'rendimento', 'rendimentos', 'deposito',
        'boleto recebido', 'boletos recebidos', 'estorno cred', 'credito'
    ]
    if any(termo in h_norm for termo in termos_entrada):
        return val
        
    # 4. Análise semântica inteligente por palavras-chave de saída no histórico
    termos_saida = [
        'ted emitido', 'ted emi do', 'pix env', 'pix enviado', 'ted env', 'doc env', 'pagto', 'pagamento', 
        'tarifa', 'manut', 'cobranca', 'debito', 'saque', 'compra', 'cartao', 
        'transferencia env', 'transf env', 'cpfl', 'darf', 'gps', 'iss', 'imposto',
        'aplicacao', 'aplic', 'investimento', 'estorno deb', 'saida', 'db', 'sispag',
        'concessionaria', 'tributo', 'boleto pago', 'tarifa emissao', 'tarifa emissao de ted', 'emitido'
    ]
    
    if any(termo in h_norm for termo in termos_saida):
        return -val
        
    # Padrão para recebimentos, pix recebido, ted recebida, rendimentos, etc.
    return val

def limpar_valor_monetario(v_val):
    if pd.isna(v_val) or v_val == '': return 0.0
    if isinstance(v_val, (int, float)): return float(v_val)
    
    s = str(v_val).strip().upper()
    is_negative = False
    if '-' in s or s.endswith('D') or s.endswith('SAÍDA') or s.endswith('SAIDA') or re.search(r'\(\s*[\d\.,]+\s*\)', s):
        is_negative = True
        
    s = re.sub(r'[^\d,\.]', '', s)
    if not s: return 0.0
    
    if ',' in s and '.' in s:
        last_dot, last_comma = s.rfind('.'), s.rfind(',')
        s = s.replace('.', '').replace(',', '.') if last_comma > last_dot else s.replace(',', '')
    elif ',' in s: 
        s = s.replace(',', '.')
        
    try: 
        val = float(s)
        return -abs(val) if is_negative else abs(val)
    except: 
        return 0.0

def identificar_banco_inteligente(texto_conteudo, filename_str=""):
    # O nome do arquivo tem prioridade para impedir que fornecedores citados no
    # histórico sejam confundidos com o banco emissor do extrato.
    nome = normalizar_texto(str(filename_str)).upper()
    cabecalho = normalizar_texto(str(texto_conteudo)[:6000]).upper()
    digitos_nome = re.sub(r'\D', '', str(filename_str))
    digitos_cabecalho = re.sub(r'\D', '', str(texto_conteudo)[:6000])

    # As empresas também nomeiam os extratos apenas com agência/conta. Essa
    # identificação não depende do mês ou do ano presentes no nome do arquivo.
    contas_nova_geracao = [
        ('995495', 'BANCO ITAU'),
        ('4519906', 'BANCO BRADESCO'),
        ('6739471', 'BANCO FIBRA'),
    ]
    for conta, banco in contas_nova_geracao:
        if conta in digitos_nome:
            return banco

    bancos = [
        (['ITAU'], 'BANCO ITAU'),
        (['BRADESCO'], 'BANCO BRADESCO'),
        (['FIBRA'], 'BANCO FIBRA'),
        (['DAYCOVAL', 'DAYCONNECT'], 'BANCO DAYCOVAL'),
        (['SANTANDER'], 'BANCO SANTANDER'),
        (['SICOOB'], 'SICOOB'),
        (['SICREDI'], 'SICREDI'),
        (['NUBANK', 'NU PAGAMENTO'], 'NUBANK'),
        (['CAIXA ECONOMICA'], 'CAIXA ECONOMICA'),
        (['BANCO DO BRASIL'], 'BANCO DO BRASIL'),
        (['BANCO INTER'], 'BANCO INTER'),
    ]
    for termos, banco in bancos:
        if any(termo in nome for termo in termos):
            return banco
    for termos, banco in bancos:
        if any(termo in cabecalho for termo in termos):
            return banco
    for conta, banco in contas_nova_geracao:
        if conta in digitos_cabecalho:
            return banco

    # Itaú Empresas: em alguns PDFs o logotipo é imagem e a palavra "Itaú"
    # não existe na camada de texto. Identificamos então pela assinatura estrutural
    # exclusiva do extrato detalhado, sem depender do nome do arquivo.
    assinatura_itau = (
        (
            'LANCAMENTOS DO PERIODO' in cabecalho
            and 'RAZAO SOCIAL' in cabecalho
            and 'CNPJ/CPF' in cabecalho
            and 'VALOR (R$)' in cabecalho
            and 'SALDO (R$)' in cabecalho
            and ('LIMITE DA CONTA' in cabecalho or 'SALDO TOTAL' in cabecalho)
        )
        or (
            'LANCAMENTOS PERIODO' in cabecalho
            and 'CONTA CORRENTE' in cabecalho
            and 'AG/ORIGEM' in cabecalho
            and 'VALOR (R$)' in cabecalho
            and 'SALDO (R$)' in cabecalho
            and ('SISPAG' in cabecalho or 'APLIC AUT' in cabecalho)
        )
    )
    if assinatura_itau:
        return 'BANCO ITAU'

    if '58.616.418' in str(texto_conteudo)[:6000]: return 'BANCO FIBRA'
    if re.search(r'\b0?341\b', cabecalho): return 'BANCO ITAU'
    return "BANCO CONTA CORRENTE"

# ==============================================================================
# MOTORES DE EXTRAÇÃO UNIVERSAL
# ==============================================================================
def processar_ofx(file_bytes, filename):
    lancamentos = []
    texto = ""
    for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
        try: texto = file_bytes.decode(enc); break
        except: pass
    if not texto: texto = file_bytes.decode('latin1', errors='ignore')
    
    banco_detectado = identificar_banco_inteligente(texto, filename)
    raw_blocks = re.split(r'<STMTTRN>', texto, flags=re.IGNORECASE)
    
    for block in raw_blocks[1:]:
        block_clean = re.split(r'</STMTTRN>|</BANKTRANLIST>', block, flags=re.IGNORECASE)[0]
        match_date = re.search(r'<DTPOSTED>\s*(\d{4}[-/\.]?\d{2}[-/\.]?\d{2}|\d{8})', block_clean, re.IGNORECASE)
        match_amt = re.search(r'<TRNAMT>\s*([\+\-]?[\d\.\,]+)', block_clean, re.IGNORECASE)
        match_memo = re.search(r'<(?:MEMO|NAME|PAYEE)>\s*(.*?)(?:\r|\n|<|$)', block_clean, re.IGNORECASE)
        match_type = re.search(r'<TRNTYPE>\s*([A-Z]+)', block_clean, re.IGNORECASE)
        
        if match_date and match_amt:
            dt_s = match_date.group(1).replace('-', '').replace('/', '').replace('.', '')
            if len(dt_s) >= 8: data_fmt = f"{dt_s[6:8]}/{dt_s[4:6]}/{dt_s[:4]}"
            else: continue
            
            valor_bruto = limpar_valor_monetario(match_amt.group(1).strip())
            trntype = match_type.group(1).upper() if match_type else ""
            historico = limpar_caracteres_ilegais(match_memo.group(1).strip().replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')) if match_memo else "TRANSACAO OFX"
            
            if 'SALDO' in historico.upper(): continue
            
            if trntype in ['DEBIT', 'PAYMENT', 'FEE']:
                valor_float = -abs(valor_bruto)
            elif trntype in ['CREDIT', 'DEP', 'DIRECTDEP']:
                valor_float = abs(valor_bruto)
            else:
                valor_float = interpretar_sinal_inteligente(historico, valor_bruto)
                
            if valor_float != 0: 
                lancamentos.append({'DESCRIÇÃO': banco_detectado, 'DATA': data_fmt, 'VALOR': valor_float, 'DÉBITO': '', 'CRÉDITO': '', 'HISTÓRICO': historico})
    return lancamentos

def processar_planilha_universal(file_bytes, filename):
    lancamentos, df = [], None
    ext = os.path.splitext(filename)[1].lower()
    if ext in ['.xlsx', '.xls']:
        try:
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            for sheet in xls.sheet_names:
                df_temp = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                if df_temp is not None and not df_temp.empty and df_temp.shape[1] > 1:
                    df = df_temp
                    break
        except Exception:
            try: dfs = pd.read_html(io.BytesIO(file_bytes))
            except Exception: pass
            if dfs: df = dfs[0]
    if df is None:
        for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1', 'utf-16']:
            for sep in [';', ',', '\t', '|']:
                try: df_temp = pd.read_csv(io.BytesIO(file_bytes), sep=sep, encoding=enc, dtype=str)
                except Exception: pass
                if df_temp is not None and df_temp.shape[1] > 1: df = df_temp; break
            if df is not None: break
    if df is None or df.empty: return []
    
    texto_amostra = " ".join([str(v) for row in df.head(10).values for v in row if pd.notna(v)])
    banco_detectado = identificar_banco_inteligente(texto_amostra, filename)
    
    header_idx = None
    for idx, row in df.iterrows():
        row_str = normalizar_texto(" ".join([str(v) for v in row.values if pd.notna(v)]))
        if ('data' in row_str or 'dt' in row_str) and ('valor' in row_str or 'credito' in row_str or 'debito' in row_str or 'lancamento' in row_str or 'historico' in row_str):
            header_idx = idx; break
            
    if header_idx is not None:
        df.columns = [str(v).strip() for v in df.iloc[header_idx].values]
        df = df.iloc[header_idx+1:].copy()
    
    cols_map = {c: normalizar_texto(c) for c in df.columns}
    col_data = next((c for c, nc in cols_map.items() if any(p in nc for p in ['data', 'dt', 'date', 'dia'])), None)
    col_hist = next((c for c, nc in cols_map.items() if any(p in nc for p in ['lancamento', 'historico', 'hist', 'razao social', 'descric', 'detalhe', 'memo'])), None)
    col_val = next((c for c, nc in cols_map.items() if any(p in nc for p in ['valor', 'val', 'monto', 'amount'])), None)
    col_cred = next((c for c, nc in cols_map.items() if any(p in nc for p in ['credito', 'credit', 'entrada', 'vlr_cred', 'crd'])), None)
    col_deb = next((c for c, nc in cols_map.items() if any(p in nc for p in ['debito', 'debit', 'saida', 'vlr_deb', 'deb'])), None)
    col_tipo = next((c for c, nc in cols_map.items() if any(p in nc for p in ['tipo', 'natureza', 'operacao', 'c/d'])), None)

    if not col_data: return []
    
    for _, row in df.iterrows():
        dt_raw = str(row[col_data]).strip() if pd.notna(row[col_data]) else ''
        if dt_raw.upper() in ['TOTAL', 'ÚLTIMOS LANÇAMENTOS', 'ULTIMOS LANCAMENTOS', 'SALDOS INVEST FÁCIL / PLUS', 'NAN', 'SALDO ANTERIOR']:
            if dt_raw.upper() == 'TOTAL': break
            continue
            
        match_dt = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', dt_raw)
        if not match_dt: continue
        dt_fmt = match_dt.group(1).replace('-', '/')
        
        hist_raw = limpar_caracteres_ilegais(str(row[col_hist]).strip()) if col_hist and pd.notna(row[col_hist]) else 'MOVIMENTO BANCARIO'
        hist_fmt = hist_raw if hist_raw.lower() != 'nan' else 'MOVIMENTO BANCARIO'
        if any(term in hist_fmt.upper() for term in ['SALDO', 'SUBTOTAL', 'TOTAL', 'TRANSPORTAR']): continue
        
        valor_float = 0.0
        
        if col_cred or col_deb:
            v_cred = limpar_valor_monetario(row[col_cred]) if col_cred and pd.notna(row[col_cred]) else 0.0
            v_deb = limpar_valor_monetario(row[col_deb]) if col_deb and pd.notna(row[col_deb]) else 0.0
            if v_cred != 0: valor_float = abs(v_cred)
            elif v_deb != 0: valor_float = -abs(v_deb)
        elif col_val and pd.notna(row[col_val]):
            val_cru = limpar_valor_monetario(row[col_val])
            tipo_str = str(row[col_tipo]).strip() if col_tipo and pd.notna(row[col_tipo]) else ""
            valor_float = interpretar_sinal_inteligente(hist_fmt, val_cru, tipo_str)

        if valor_float != 0:
            lancamentos.append({'DESCRIÇÃO': banco_detectado, 'DATA': dt_fmt, 'VALOR': valor_float, 'DÉBITO': '', 'CRÉDITO': '', 'HISTÓRICO': hist_fmt})
    return lancamentos

def extrair_periodo_extrato(caminho_pdf):
    try:
        reader = PdfReader(caminho_pdf, strict=False)
        texto = "".join([p.extract_text() for p in reader.pages[:3]])
        datas = re.findall(r'(\d{2}/\d{2}/\d{4})', texto)
        if len(datas) >= 2: return datetime.strptime(datas[0], '%d/%m/%Y'), datetime.strptime(datas[1], '%d/%m/%Y')
    except: pass
    return None, None

def processar_pdf_layout_universal(reader, banco_identificado):
    """
    Analisa tabelas bancárias pela estrutura do próprio PDF, sem regras por banco.

    O motor combina: cabeçalhos detectados, posição das colunas, última data
    válida, sinais C/D, semântica do histórico e variação matemática do saldo.
    """
    lancamentos = []
    tabela_ativa = False
    encontrou_cabecalho = False
    data_atual = None
    saldo_anterior = None
    pos_credito = None
    pos_debito = None
    pos_valor = None
    pos_saldo = None
    historico_pendente = ""

    date_regex = re.compile(r'(?<!\d)(\d{2}/\d{2}/\d{4})(?!\d)')
    valor_regex = re.compile(
        r'(?<!\d)(?:R\$\s*)?(\(?\s*[+-]?\s*\d{1,3}(?:\.\d{3})*,\d{2}\s*\)?\s*[CD]?)(?!\d)',
        re.IGNORECASE
    )

    def limpar_historico_linha(linha, data_linha=None):
        hist = linha
        if data_linha:
            hist = hist.replace(data_linha, ' ', 1)
        hist = valor_regex.sub(' ', hist).replace('R$', ' ')
        hist = hist.replace('Emi\x00do', 'Emitido').replace('\x00', '')
        return re.sub(r'\s+', ' ', hist).strip(' -|')

    def linha_auxiliar_valida(linha):
        norm = normalizar_texto(linha)
        if re.match(r'^\s*\d+\s*/\s*\d+\s*$', linha):
            return False
        bloqueios = [
            'extrato mensal', 'nome do usuario', 'data da operacao', 'folha ',
            'pagina ', 'sujeito a alteracoes', 'fim de relatorio', 'cnpj:',
            'agencia | conta', 'total disponivel', 'extrato de:', 'lancamento dcto',
            'lembramos que', 'movimentacao de saldo', 'sua validade restrita'
        ]
        return bool(linha.strip()) and not any(item in norm for item in bloqueios)

    parar_processamento = False
    for pagina in reader.pages:
        try:
            texto_layout = pagina.extract_text(extraction_mode="layout") or ""
        except (TypeError, ValueError):
            texto_layout = pagina.extract_text() or ""

        linhas = texto_layout.splitlines()
        for indice_linha, linha in enumerate(linhas):
            norm = normalizar_texto(linha)

            # Encerra o extrato principal antes de avisos, projeções ou uma nova
            # tabela de lançamentos futuros existente no mesmo PDF.
            if lancamentos and (
                'lancamentos futuros do periodo' in norm or
                norm.startswith('aviso: os saldos acima') or
                (norm.startswith('saldo de ') and not date_regex.search(linha)) or
                norm.startswith('posicao em:')
            ):
                parar_processamento = True
                break

            # Detecta dinamicamente as colunas da tabela.
            tem_data = re.search(r'\bdata\b', norm) is not None
            tem_coluna_monetaria = any(k in norm for k in ['credito', 'debito', 'valor', 'saldo'])
            if tem_data and tem_coluna_monetaria:
                tabela_ativa = True
                encontrou_cabecalho = True
                if 'credito' in norm:
                    pos_credito = normalizar_texto(linha).find('credito')
                if 'debito' in norm:
                    pos_debito = normalizar_texto(linha).find('debito')
                if 'valor' in norm:
                    pos_valor = normalizar_texto(linha).find('valor')
                if 'saldo' in norm:
                    pos_saldo = normalizar_texto(linha).find('saldo')
                continue

            # Finaliza a primeira tabela principal antes de resumos ou outras seções.
            if tabela_ativa and lancamentos and re.match(r'^\s*total\b', norm):
                parar_processamento = True
                break

            ocorrencias = list(valor_regex.finditer(linha))
            match_data = date_regex.search(linha)
            datas_na_linha = date_regex.findall(linha)

            # PDFs sem cabeçalho ainda podem iniciar por uma linha transacional.
            if not tabela_ativa and not encontrou_cabecalho:
                if (match_data and len(datas_na_linha) == 1 and ocorrencias and
                        not any(k in norm for k in ['periodo', 'saldo', 'disponivel', 'limite'])):
                    tabela_ativa = True
                else:
                    continue

            if not tabela_ativa:
                continue

            # Saldo anterior/de abertura serve para validar matematicamente o sinal.
            if 'saldo anterior' in norm or 'saldo inicial' in norm:
                if match_data:
                    data_atual = match_data.group(1)
                if ocorrencias:
                    saldo_anterior = limpar_valor_monetario(ocorrencias[-1].group(1))
                continue

            # Linhas isoladas de saldo não são lançamentos.
            if ('saldo' in norm and not any(k in norm for k in ['rentab', 'rendimento'])):
                if ocorrencias:
                    saldo_anterior = limpar_valor_monetario(ocorrencias[-1].group(1))
                continue

            # Resumos financeiros e limites não representam movimentações.
            if any(k in norm for k in [
                'disponivel', 'limite adicional', 'bloqueado', 'c.p.m.f',
                'provisionado', 'lancamentos futuros', 'tarifas pendentes',
                'previsao encargos', 'posicao em:'
            ]):
                continue

            # Uma linha com movimento + saldo tem ao menos dois valores. Quando há
            # apenas um, o cabeçalho/posição e o histórico definem sua natureza.
            if not ocorrencias:
                # Alguns PDFs quebram um valor alto: a linha da transação termina
                # em "R$" e o número aparece sozinho na linha seguinte.
                if match_data and len(datas_na_linha) == 1 and 'R$' in linha:
                    data_atual = match_data.group(1)
                    historico_pendente = limpar_historico_linha(linha, data_atual)
                    continue
                if lancamentos and linha_auxiliar_valida(linha):
                    complemento = limpar_historico_linha(linha)
                    if complemento and not date_regex.search(complemento):
                        hist_atual = lancamentos[-1]['HISTÓRICO']
                        lancamentos[-1]['HISTÓRICO'] = re.sub(r'\s+', ' ', f"{hist_atual} {complemento}").strip()
                continue

            if match_data:
                data_atual = match_data.group(1)
            if not data_atual:
                continue

            # Valores anteriores ao último são movimento; o último é saldo quando
            # a tabela possui coluna Saldo e há mais de um valor na linha.
            tem_saldo_linha = pos_saldo is not None and len(ocorrencias) >= 2
            saldo_linha = limpar_valor_monetario(ocorrencias[-1].group(1)) if tem_saldo_linha else None
            candidatos = ocorrencias[:-1] if tem_saldo_linha else ocorrencias
            candidatos_validos = [m for m in candidatos if limpar_valor_monetario(m.group(1)) != 0]
            if not candidatos_validos:
                if saldo_linha is not None:
                    saldo_anterior = saldo_linha
                continue

            mov = candidatos_validos[0]
            token_mov = mov.group(1).strip()
            valor_bruto = limpar_valor_monetario(token_mov)
            valor_abs = abs(valor_bruto)
            natureza = ''
            if re.search(r'D\s*$', token_mov, re.IGNORECASE):
                natureza = 'D'
            elif re.search(r'C\s*$', token_mov, re.IGNORECASE):
                natureza = 'C'

            hist_linha = limpar_historico_linha(linha, match_data.group(1) if match_data else None)
            hist = re.sub(r'\s+', ' ', f"{historico_pendente} {hist_linha}").strip()
            historico_pendente = ""
            hist_norm = normalizar_texto(hist)
            if any(k in hist_norm for k in ['saldo invest', 'saldo anterior', 'saldo final', 'saldo do dia']):
                if saldo_linha is not None:
                    saldo_anterior = saldo_linha
                continue

            # Prioridade 1: sinal explícito no próprio valor.
            if valor_bruto < 0 or natureza == 'D':
                valor_final = -valor_abs
            elif natureza == 'C':
                valor_final = valor_abs
            else:
                valor_final = None

            # Prioridade 2: diferença exata entre saldo atual e saldo anterior.
            if valor_final is None and saldo_linha is not None and saldo_anterior is not None:
                diferenca = round(saldo_linha - saldo_anterior, 2)
                if abs(abs(diferenca) - valor_abs) <= 0.05:
                    valor_final = valor_abs if diferenca > 0 else -valor_abs

            # Prioridade 3: posição em colunas Débito/Crédito detectadas.
            if valor_final is None and pos_debito is not None and pos_credito is not None:
                pos_movimento = linha.find('R$', max(0, mov.start() - 4), mov.end())
                if pos_movimento < 0:
                    pos_movimento = mov.start(1)
                dist_debito = abs(pos_movimento - pos_debito)
                dist_credito = abs(pos_movimento - pos_credito)
                valor_final = -valor_abs if dist_debito < dist_credito else valor_abs

            # Prioridade 4: sinal semântico como último recurso.
            if valor_final is None:
                valor_final = interpretar_sinal_inteligente(hist, valor_bruto, natureza)

            if valor_abs != 0:
                lancamentos.append({
                    'DESCRIÇÃO': banco_identificado,
                    'DATA': data_atual,
                    'VALOR': valor_final,
                    'DÉBITO': '',
                    'CRÉDITO': '',
                    'HISTÓRICO': limpar_caracteres_ilegais(hist or 'MOVIMENTO BANCARIO')
                })

            if saldo_linha is not None:
                saldo_anterior = saldo_linha

        if parar_processamento:
            break

    return lancamentos

def extrair_valor_lancamento_pdf(texto_bloco):
    """
    Seleciona o valor da movimentação sem confundi-lo com o saldo.

    Extratos normalmente exibem o valor do lançamento antes do saldo. Esta
    função também descarta números ligados explicitamente a saldo anterior,
    saldo atual, disponível, limite e resumos semelhantes.
    """
    padrao_valor = re.compile(
        r'(?<!\d)(?:R\$\s*)?(\(?\s*[+-]?[\d\.]+,\d{2}\s*\)?\s*[CD]?)(?!\d)',
        re.IGNORECASE
    )
    ocorrencias = list(padrao_valor.finditer(texto_bloco))
    if not ocorrencias:
        return None, ""

    termos_resumo = [
        'saldo', 'disponivel', 'limite', 'bloqueado', 'provisionado',
        'saldo atual', 'saldo anterior', 'saldo final', 'saldo do dia'
    ]
    candidatos = []
    for ocorrencia in ocorrencias:
        contexto_antes = normalizar_texto(texto_bloco[max(0, ocorrencia.start() - 35):ocorrencia.start()])
        if any(termo in contexto_antes for termo in termos_resumo):
            continue
        candidatos.append(ocorrencia)

    # Se todos foram marcados como resumo, não cria um lançamento de saldo.
    if not candidatos:
        return None, ""

    # Nos formatos Valor + Saldo, Débito + Saldo ou Crédito + Saldo, o primeiro
    # valor monetário útil pertence ao lançamento e os seguintes são saldos.
    escolhido = candidatos[0]
    token = escolhido.group(1).strip()
    natureza = ""
    if re.search(r'D\s*$', token, re.IGNORECASE):
        natureza = "D"
    elif re.search(r'C\s*$', token, re.IGNORECASE):
        natureza = "C"

    return token, natureza

def processar_pdf_itau_detalhado(reader, banco_identificado):
    """Lê os formatos detalhados do Itaú Empresas sem importar linhas de saldo."""
    textos = [(pagina.extract_text() or '') for pagina in reader.pages]
    texto_total = '\n'.join(textos)
    texto_norm = normalizar_texto(texto_total)

    assinatura_moderno = (
        'lancamentos do periodo' in texto_norm
        and 'razao social' in texto_norm
        and 'valor (r$)' in texto_norm
        and 'saldo (r$)' in texto_norm
    )
    assinatura_abreviado = (
        'lancamentos periodo' in texto_norm
        and 'conta corrente' in texto_norm
        and 'ag/origem' in texto_norm
        and 'valor (r$)' in texto_norm
        and 'saldo (r$)' in texto_norm
        and ('sispag' in texto_norm or 'aplic aut' in texto_norm)
    )
    if (
        banco_identificado not in {'BANCO ITAU', 'BANCO ITAÚ'}
        and 'itau' not in texto_norm
        and not assinatura_moderno
        and not assinatura_abreviado
    ):
        return []

    termos_saldo = [
        'saldo anterior', 'saldo aplic', 'saldo aplic. aut',
        'saldo total disponivel dia', 'saldo movimentacao conta',
        'sdo aplic aut mais ap', 'saldo em conta corrente',
        'saldo da conta corrente', 'saldo disponivel sem investimentos',
        'saldo em aplicacao automatica', 'valor total em aplicacoes automaticas',
        'saldo total disponivel', 'saldo total', 'limite da conta',
        'total disponivel para uso', 'utilizado', 'disponivel',
    ]
    padrao_valor = re.compile(r'(?<!\d)([-+]?\s*\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)')

    # Formato Itaú Empresas em que o PDF extrai datas como "02 / mar" e,
    # frequentemente, cola a data DEPOIS do valor: "-1.621,0002 / mar".
    # O período declarado no cabeçalho define o ano e também impede que a seção
    # de lançamentos futuros (por exemplo, abril em um extrato de março) seja lida.
    periodo_match = re.search(
        r'lancamentos\s+periodo\s*:\s*(\d{2}/\d{2}/\d{4})\s+ate\s+(\d{2}/\d{2}/\d{4})',
        texto_norm,
        re.IGNORECASE,
    )
    if assinatura_abreviado and periodo_match:
        try:
            periodo_inicio = datetime.strptime(periodo_match.group(1), '%d/%m/%Y').date()
            periodo_fim = datetime.strptime(periodo_match.group(2), '%d/%m/%Y').date()
        except ValueError:
            periodo_inicio = periodo_fim = None

        if periodo_inicio and periodo_fim:
            meses = {
                'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4,
                'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8,
                'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
            }
            moeda = r'[-+]?\d{1,3}(?:\.\d{3})*,\d{2}'
            padrao_data_sufixo = re.compile(
                rf'^(?P<hist>.+?)\s+(?P<valor>{moeda})(?P<dia>\d{{2}})\s*/\s*(?P<mes>[A-Za-zÀ-ÿ#]+)$'
            )
            padrao_data_prefixo = re.compile(
                rf'^(?P<dia>\d{{2}})\s*/\s*(?P<mes>[A-Za-zÀ-ÿ#]+)\s+(?P<hist>.+?)\s+(?P<valor>{moeda})$'
            )

            def resolver_data_curta(dia_raw, mes_raw):
                mes_norm = normalizar_texto(str(mes_raw)).replace('#', '')[:3]
                numero_mes = meses.get(mes_norm)
                if not numero_mes:
                    return None
                for ano in sorted({periodo_inicio.year, periodo_fim.year}):
                    try:
                        candidato = datetime(ano, numero_mes, int(dia_raw)).date()
                    except ValueError:
                        continue
                    if periodo_inicio <= candidato <= periodo_fim:
                        return candidato
                return None

            lancamentos_abreviados = []
            for texto_pagina in textos:
                for linha_bruta in texto_pagina.splitlines():
                    linha = re.sub(r'\s+', ' ', linha_bruta).strip()
                    if not linha:
                        continue
                    correspondencia = padrao_data_sufixo.match(linha)
                    if correspondencia is None:
                        correspondencia = padrao_data_prefixo.match(linha)
                    if correspondencia is None:
                        continue

                    data_lancamento = resolver_data_curta(
                        correspondencia.group('dia'), correspondencia.group('mes')
                    )
                    if data_lancamento is None:
                        continue

                    historico = re.sub(r'\s+', ' ', correspondencia.group('hist')).strip(' -|')
                    historico_norm = normalizar_texto(historico)
                    if not historico or any(termo in historico_norm for termo in termos_saldo):
                        continue

                    valor = limpar_valor_monetario(correspondencia.group('valor'))
                    if abs(valor) < 0.005:
                        continue

                    lancamentos_abreviados.append({
                        'DESCRIÇÃO': banco_identificado if banco_identificado in {'BANCO ITAU', 'BANCO ITAÚ'} else 'BANCO ITAU',
                        'DATA': data_lancamento.strftime('%d/%m/%Y'),
                        'VALOR': valor,
                        'DÉBITO': '',
                        'CRÉDITO': '',
                        'HISTÓRICO': limpar_caracteres_ilegais(historico),
                    })

            if lancamentos_abreviados:
                return lancamentos_abreviados

    # Formato detalhado atual, com datas completas no início da linha.
    linhas_por_pagina = [
        [re.sub(r'\s+', ' ', linha).strip() for linha in texto.splitlines() if linha.strip()]
        for texto in textos
    ]
    padrao_data = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(.*)$')
    blocos = []
    atual = None

    for linhas_pagina in linhas_por_pagina:
        for linha in linhas_pagina:
            m = padrao_data.match(linha)
            if m:
                # Na quebra de página o Itaú pode repetir a mesma data para
                # continuar um lançamento cujo valor ficou na página seguinte.
                atual_sem_valor = bool(atual) and not padrao_valor.search(atual[1])
                if atual_sem_valor and atual[0] == m.group(1):
                    atual[1] += ' ' + m.group(2)
                    continue
                if atual:
                    blocos.append(atual)
                atual = [m.group(1), m.group(2)]
            elif atual:
                atual[1] += ' ' + linha
    if atual:
        blocos.append(atual)

    lancamentos = []
    for data_str, conteudo in blocos:
        conteudo_norm = normalizar_texto(conteudo)
        if any(termo in conteudo_norm for termo in termos_saldo):
            continue
        if conteudo_norm.startswith(('aviso', 'atualizado em')):
            continue

        valores = list(padrao_valor.finditer(conteudo))
        if not valores:
            continue

        valor_match = valores[-1]
        valor = limpar_valor_monetario(valor_match.group(1))
        if abs(valor) < 0.005:
            continue

        historico = (conteudo[:valor_match.start()] + ' ' + conteudo[valor_match.end():]).strip()
        historico = re.sub(r'\s+', ' ', historico).strip(' -|')
        if not historico:
            historico = 'MOVIMENTO BANCARIO'

        lancamentos.append({
            'DESCRIÇÃO': banco_identificado or 'BANCO ITAU',
            'DATA': data_str,
            'VALOR': valor,
            'DÉBITO': '',
            'CRÉDITO': '',
            'HISTÓRICO': limpar_caracteres_ilegais(historico),
        })

    return lancamentos

def processar_pdf_daycoval_detalhado(reader, banco_identificado):
    """
    Lê extratos detalhados Dayconnect/Daycoval antigos e recentes.

    Alguns PDFs preservam o texto normalmente; outros inserem espaços entre
    letras, datas e valores (ex.: ``01/ 05 T A R I F A ... - R $  7 , 3 7``).
    O parser normaliza essa fragmentação, preserva lançamentos repetidos reais,
    respeita o período declarado no extrato e usa o modo layout somente como
    fallback quando a extração textual simples não produz lançamentos.
    """
    textos_simples = [(pagina.extract_text() or '') for pagina in reader.pages]
    texto_identificacao = normalizar_texto('\n'.join(textos_simples[:2]))
    if (
        banco_identificado != 'BANCO DAYCOVAL'
        and 'daycoval' not in texto_identificacao
        and 'dayconnect' not in texto_identificacao
    ):
        return []

    padrao_data_inicio = re.compile(
        r'^\s*(\d)\s*(\d)\s*/\s*(\d)\s*(\d)\s+(.*)$'
    )
    padrao_data_completa = re.compile(
        r'(\d)\s*(\d)\s*/\s*(\d)\s*(\d)\s*/\s*(2)\s*(0)\s*(\d)\s*(\d)'
    )
    padrao_moeda = re.compile(
        r'([+-]?\s*R\s*\$\s*[+-]?\s*\d[\d\s.]*,\s*\d\s*\d)',
        re.IGNORECASE
    )

    def desfragmentar_linha(linha):
        texto = str(linha or '').replace('\xa0', ' ')
        tokens = texto.split()
        alfabeticos = [
            token for token in tokens
            if any(caractere.isalpha() for caractere in token)
        ]
        unitarios = [
            token for token in alfabeticos
            if len(re.sub(r'[^A-Za-zÀ-ÿ]', '', token)) == 1
        ]
        fragmentado = bool(alfabeticos) and (
            len(unitarios) / len(alfabeticos) >= 0.45
        )
        if fragmentado:
            # Remove somente ESPAÇO SIMPLES entre letras. Espaços duplos do PDF
            # continuam separando palavras e são colapsados apenas ao final.
            texto = re.sub(
                r'(?<=[A-Za-zÀ-ÿ]) (?=[A-Za-zÀ-ÿ])', '', texto
            )
        return re.sub(r'\s+', ' ', texto).strip()

    def converter_moeda(valor_raw):
        texto = re.sub(r'\s+', '', str(valor_raw).upper()).replace('R$', '')
        sinal = -1.0 if texto.startswith('-') else 1.0
        texto = texto.lstrip('+-').replace('.', '').replace(',', '.')
        try:
            return sinal * float(texto)
        except (TypeError, ValueError):
            return 0.0

    def extrair_data_match(match):
        dia = match.group(1) + match.group(2)
        mes = match.group(3) + match.group(4)
        ano = '20' + match.group(7) + match.group(8)
        try:
            return datetime.strptime(
                f'{dia}/{mes}/{ano}', '%d/%m/%Y'
            ).date()
        except ValueError:
            return None

    def localizar_periodo(texto_total):
        linhas = texto_total.splitlines()
        for indice, linha in enumerate(linhas):
            normalizada = normalizar_texto(desfragmentar_linha(linha))
            if 'periodo' not in normalizada:
                continue
            janela = ' '.join(linhas[indice:indice + 4])
            datas = list(padrao_data_completa.finditer(janela))
            if len(datas) >= 2:
                data_inicial = extrair_data_match(datas[0])
                data_final = extrair_data_match(datas[1])
                if data_inicial and data_final and data_inicial <= data_final:
                    return data_inicial, data_final
        return None

    def processar_fonte(texto_total):
        periodo = localizar_periodo(texto_total)
        if periodo:
            ano_referencia = periodo[0].year
        else:
            data_completa = padrao_data_completa.search(texto_total)
            data_referencia = extrair_data_match(data_completa) if data_completa else None
            ano_referencia = data_referencia.year if data_referencia else datetime.now().year

        blocos = []
        atual = None
        prefixos_fim = (
            'impressao realizada', 'central de atendimento',
            'horario de atendimento', 'sac daycoval',
            'central para deficientes', 'ouvidoria:', 'os saldos acima',
            'saldo anterior', 'extrato detalhado', 'conta corrente',
            'saldo disponivel', 'titular', 'periodo', 'agencia', 'conta ',
            'saldo atual', 'limite ', 'saldo bloqueado', 'valor bloqueado',
            'provisao de encargos', 'lancamentos futuros'
        )

        for linha_bruta in texto_total.splitlines():
            correspondencia = padrao_data_inicio.match(linha_bruta)
            if correspondencia:
                if atual:
                    blocos.append(atual)
                atual = {
                    'dia': correspondencia.group(1) + correspondencia.group(2),
                    'mes': correspondencia.group(3) + correspondencia.group(4),
                    'linhas': [correspondencia.group(5)]
                }
                continue

            if atual:
                linha_normalizada = normalizar_texto(
                    desfragmentar_linha(linha_bruta)
                )
                if (
                    any(linha_normalizada.startswith(prefixo) for prefixo in prefixos_fim)
                    or ('feira' in linha_normalizada and 'saldo:' in linha_normalizada)
                ):
                    blocos.append(atual)
                    atual = None
                    continue
                atual['linhas'].append(linha_bruta)

        if atual:
            blocos.append(atual)

        lancamentos = []
        for bloco in blocos:
            conteudo = ' '.join(bloco['linhas'])
            moedas = list(padrao_moeda.finditer(conteudo))
            if not moedas:
                continue

            # O primeiro valor monetário pertence ao lançamento. Isso evita que
            # um "Saldo Anterior" posterior contamine o último movimento da página.
            moeda = moedas[0]
            valor = converter_moeda(moeda.group(1))
            if abs(valor) < 0.005:
                continue

            historico = desfragmentar_linha(
                conteudo[:moeda.start()]
            ).strip(' -|')
            if not historico:
                continue
            historico_normalizado = normalizar_texto(historico)
            if historico_normalizado.startswith('saldo ') or 'saldo:' in historico_normalizado:
                continue

            try:
                data_lancamento = datetime(
                    ano_referencia,
                    int(bloco['mes']),
                    int(bloco['dia'])
                ).date()
            except ValueError:
                continue

            # Extratos emitidos no início do mês seguinte podem exibir um movimento
            # posterior ao período solicitado. Para conciliação, prevalece o período
            # declarado pelo próprio banco.
            if periodo and not (periodo[0] <= data_lancamento <= periodo[1]):
                continue

            lancamentos.append({
                'DESCRIÇÃO': banco_identificado or 'BANCO DAYCOVAL',
                'DATA': data_lancamento.strftime('%d/%m/%Y'),
                'VALOR': round(valor, 2),
                'DÉBITO': '',
                'CRÉDITO': '',
                'HISTÓRICO': limpar_caracteres_ilegais(historico)
            })

        return lancamentos

    # Prioriza a extração simples: é a mais fiel nos modelos recentes e nos PDFs
    # antigos com texto fragmentado. O layout fica como fallback real, não é somado,
    # portanto lançamentos legítimos repetidos no mesmo dia/valor são preservados.
    texto_simples_total = '\n'.join(textos_simples)
    lancamentos_simples = processar_fonte(texto_simples_total)
    if lancamentos_simples:
        return lancamentos_simples

    textos_layout = []
    for pagina, texto_simples in zip(reader.pages, textos_simples):
        try:
            texto_layout = pagina.extract_text(extraction_mode='layout') or texto_simples
        except (TypeError, ValueError):
            texto_layout = texto_simples
        textos_layout.append(texto_layout)

    return processar_fonte('\n'.join(textos_layout))

def processar_pdf_fibra_extrato(reader, banco_identificado='BANCO FIBRA'):
    """Lê o extrato de C/C do Banco Fibra sem depender do layout posicional do PDF."""
    texto_total = '\n'.join((pagina.extract_text() or '') for pagina in reader.pages)
    texto_norm = normalizar_texto(texto_total)
    if banco_identificado != 'BANCO FIBRA' and 'banco fibra' not in texto_norm:
        return []
    if 'extrato de c/c para simples conferencia' not in texto_norm:
        return []

    linhas = [
        re.sub(r'\s+', ' ', linha.replace('Emi\x00do', 'Emitido').replace('emi\x00do', 'emitido').replace('\x00', '')).strip()
        for linha in texto_total.splitlines() if linha.strip()
    ]
    regex_data = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(.*)$')
    regex_valor = re.compile(r'R\$\s*([\d.]+,\d{2})|(?<!\d)([\d.]+,\d{2})(?!\d)')

    blocos = []
    atual = None
    for linha in linhas:
        norm = normalizar_texto(linha)
        if norm.startswith(('posicao em:', 'saldo atual:', '= disponivel:', 'saldo liquido:',
                            'lancamentos futuros:', 'tarifas pendentes:', 'previsao encargos:',
                            '= saldo provisionado:', 'fim de relatorio')):
            if atual:
                blocos.append(atual)
                atual = None
            break
        if norm.startswith('saldo '):
            if atual:
                blocos.append(atual)
                atual = None
            continue
        if norm.startswith(('pagina ', 'sujeito a alteracoes')):
            continue

        m = regex_data.match(linha)
        if m:
            if atual:
                blocos.append(atual)
            atual = {'data': m.group(1), 'linhas': [m.group(2)]}
        elif atual:
            atual['linhas'].append(linha)
    if atual:
        blocos.append(atual)

    lancamentos = []
    for bloco in blocos:
        conteudo = ' '.join(bloco['linhas'])
        norm = normalizar_texto(conteudo)
        if any(t in norm for t in ['saldo anterior', 'saldo atual', 'saldo provisionado']):
            continue

        valores = []
        for m in regex_valor.finditer(conteudo):
            token = m.group(1) or m.group(2)
            if token:
                valores.append((m, token))
        if not valores:
            continue

        m_valor, token_valor = valores[-1]
        valor_abs = abs(limpar_valor_monetario(token_valor))
        if valor_abs < 0.005:
            continue

        hist = re.sub(r'\s+', ' ', (conteudo[:m_valor.start()] + ' ' + conteudo[m_valor.end():])).strip()
        hist_norm = normalizar_texto(hist)

        if any(t in hist_norm for t in [
            'ted emitido', 'tarifa', 'debito', 'pix enviado', 'pagamento',
            'saque', 'transferencia enviada', 'ted enviado', 'doc emitido'
        ]):
            valor = -valor_abs
        elif any(t in hist_norm for t in [
            'ted recebido', 'pix recebido', 'credito', 'deposito',
            'transferencia recebida', 'recebimento'
        ]):
            valor = valor_abs
        else:
            valor = interpretar_sinal_inteligente(hist, valor_abs)

        try:
            data = datetime.strptime(bloco['data'], '%d/%m/%Y')
        except ValueError:
            continue

        lancamentos.append({
            'DESCRIÇÃO': banco_identificado or 'BANCO FIBRA',
            'DATA': data,
            'VALOR': round(valor, 2),
            'DÉBITO': '',
            'CRÉDITO': '',
            'HISTÓRICO': limpar_caracteres_ilegais(hist or 'MOVIMENTO BANCARIO')
        })

    return lancamentos

def processar_arquivo_pdf(caminho_pdf, filename_original=None):
    lancamentos = []
    try:
        reader = PdfReader(caminho_pdf, strict=False)
        # O pypdf converte o caminho em BytesIO e perde reader.stream.name.
        # Guardamos explicitamente o arquivo temporário para o fallback OCR.
        reader._razync_source_path = caminho_pdf
        texto_completo = ""
        for pagina in reader.pages:
            texto_completo += (pagina.extract_text() or "") + "\n"
            
        nome_para_identificacao = filename_original or os.path.basename(caminho_pdf)
        banco_identificado = identificar_banco_inteligente(texto_completo, nome_para_identificacao)

        # O extrato detalhado Dayconnect usa DD/MM nas linhas e informa o ano
        # somente no cabeçalho do período. Esse formato é tratado antes do
        # analisador estrutural geral e fica disponível em todos os fluxos.
        if banco_identificado == 'BANCO DAYCOVAL':
            lancamentos_daycoval = processar_pdf_daycoval_detalhado(
                reader, banco_identificado
            )
            if lancamentos_daycoval:
                return lancamentos_daycoval

        if banco_identificado in {'BANCO ITAU', 'BANCO ITAÚ'}:
            lancamentos_itau = processar_pdf_itau_detalhado(
                reader, banco_identificado
            )
            if lancamentos_itau:
                return lancamentos_itau

        if banco_identificado == 'BANCO BRADESCO':
            lancamentos_bradesco = processar_pdf_bradesco_mensal(
                reader, banco_identificado
            )
            fechamento_bradesco = getattr(reader, '_razync_balance_check', None)
            if fechamento_bradesco:
                st.session_state['ultimo_fechamento_extrato'] = fechamento_bradesco
            if lancamentos_bradesco:
                return lancamentos_bradesco
            if not texto_completo.strip():
                st.session_state['ultimo_erro_extrato'] = diagnostico_pdf_sem_lancamentos(
                    banco_identificado,
                    True,
                    bool(getattr(reader, '_razync_ocr_executado', False)),
                    str(getattr(reader, '_razync_ocr_error', '') or '')
                )

        if banco_identificado == 'BANCO FIBRA':
            lancamentos_fibra = processar_pdf_fibra_extrato(
                reader, banco_identificado
            )
            if lancamentos_fibra:
                return lancamentos_fibra

        # Primeiro tenta o analisador estrutural único, independente do banco.
        lancamentos_layout = processar_pdf_layout_universal(reader, banco_identificado)
        if lancamentos_layout:
            return lancamentos_layout

        linhas = [l.strip() for l in texto_completo.split('\n') if l.strip()]
        date_regex = re.compile(r'^(\d{2}/\d{2}/\d{4})')
        
        i = 0
        while i < len(linhas):
            linha = linhas[i]
            match_date = date_regex.match(linha)
            if match_date:
                data_str = match_date.group(1)
                bloco_linhas = [linha]
                j = i + 1
                while j < len(linhas):
                    next_linha = linhas[j]
                    if date_regex.match(next_linha) or 'SALDO' in next_linha.upper() or 'Página' in next_linha:
                        break
                    bloco_linhas.append(next_linha)
                    j += 1
                
                texto_bloco = " ".join(bloco_linhas)
                val_str, natureza_valor = extrair_valor_lancamento_pdf(texto_bloco)

                if val_str is not None:
                    v_num = limpar_valor_monetario(val_str)

                    hist = texto_bloco.replace(data_str, '', 1)
                    hist = re.sub(
                        r'(?<!\d)(?:R\$\s*)?\(?\s*[+-]?[\d\.]+,\d{2}\s*\)?\s*[CD]?(?!\d)',
                        ' ', hist, flags=re.IGNORECASE
                    )
                    hist = re.sub(r'\s+', ' ', hist).strip()

                    if not any(termo in hist.upper() for termo in ['SALDO ANTERIOR', 'SALDO FINAL', 'SALDO DO DIA']) and v_num != 0:
                        v_final = interpretar_sinal_inteligente(hist, v_num, natureza_valor)
                        lancamentos.append({'DESCRIÇÃO': banco_identificado, 'DATA': data_str, 'VALOR': v_final, 'DÉBITO': '', 'CRÉDITO': '', 'HISTÓRICO': limpar_caracteres_ilegais(hist)})
                i = j - 1
            i += 1
    except Exception as e:
        print(f"Erro no processamento PDF universal: {e}")
    return lancamentos

def processar_extrato_unificado(file_bytes, filename):
    """Leitor único de extratos usado por todas as ferramentas do Razync."""
    extensao = os.path.splitext(filename)[1].lower()
    if extensao == '.ofx':
        return processar_ofx(file_bytes, filename)
    if extensao in ['.csv', '.xlsx', '.xls']:
        return processar_planilha_universal(file_bytes, filename)
    if extensao != '.pdf':
        return []

    st.session_state.pop('ultimo_erro_extrato', None)
    st.session_state.pop('ultimo_fechamento_extrato', None)
    caminho_temporario = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temporario:
            temporario.write(file_bytes)
            caminho_temporario = temporario.name
        # O nome ORIGINAL é sempre passado. Isso evita que um arquivo temporário
        # faça o identificador perder banco/conta e cair no parser errado.
        return processar_arquivo_pdf(caminho_temporario, filename)
    finally:
        if caminho_temporario and os.path.exists(caminho_temporario):
            os.remove(caminho_temporario)

def gerar_excel_modelo_dominio(df):
    """Preenche uma cópia fiel do Modelo Domínio, preservando sua estrutura e estilos."""
    from copy import copy
    from openpyxl import load_workbook

    caminho_modelo = next(
        (caminho for caminho in ['Modelo dominio.xlsx', 'Modelo dominio(6).xlsx']
         if os.path.exists(caminho)),
        None
    )
    if not caminho_modelo:
        raise FileNotFoundError('Modelo Domínio não encontrado no sistema.')

    wb = load_workbook(caminho_modelo)
    ws = wb[wb.sheetnames[0]]

    # Localiza a linha real do cabeçalho sem presumir que seja sempre a primeira.
    cabecalho_linha = None
    mapa_colunas = {}
    for linha in range(1, min(ws.max_row, 25) + 1):
        mapa_temp = {}
        for coluna in range(1, ws.max_column + 1):
            valor = ws.cell(linha, coluna).value
            nome = normalizar_texto(str(valor or '')).strip()
            if nome:
                mapa_temp[nome] = coluna
        if 'data' in mapa_temp and 'valor' in mapa_temp and 'historico' in mapa_temp:
            cabecalho_linha = linha
            mapa_colunas = mapa_temp
            break

    if cabecalho_linha is None:
        raise ValueError('Cabeçalho do Modelo Domínio não foi localizado.')

    nomes_df = {normalizar_texto(str(c)).strip(): c for c in df.columns}
    linha_modelo = cabecalho_linha + 1

    # Guarda o estilo da primeira linha de dados do próprio modelo para replicá-lo.
    estilos = {}
    for coluna in range(1, ws.max_column + 1):
        celula = ws.cell(linha_modelo, coluna)
        estilos[coluna] = {
            'font': copy(celula.font),
            'fill': copy(celula.fill),
            'border': copy(celula.border),
            'alignment': copy(celula.alignment),
            'number_format': celula.number_format,
            'protection': copy(celula.protection),
        }

    # Remove somente conteúdos antigos da área de dados. Cabeçalho, larguras,
    # filtros, congelamentos, impressão e demais propriedades ficam intactos.
    for linha in range(cabecalho_linha + 1, ws.max_row + 1):
        for coluna in range(1, ws.max_column + 1):
            ws.cell(linha, coluna).value = None

    for indice, registro in enumerate(df.to_dict('records'), start=cabecalho_linha + 1):
        for nome_normalizado, coluna_excel in mapa_colunas.items():
            coluna_df = nomes_df.get(nome_normalizado)
            if coluna_df is None:
                continue
            valor = registro.get(coluna_df, '')
            if pd.isna(valor):
                valor = ''
            if nome_normalizado == 'data' and valor not in ('', None):
                data = pd.to_datetime(valor, dayfirst=True, errors='coerce')
                valor = data.to_pydatetime() if not pd.isna(data) else valor

            celula = ws.cell(indice, coluna_excel)
            celula.value = valor
            estilo = estilos.get(coluna_excel)
            if estilo:
                celula.font = copy(estilo['font'])
                celula.fill = copy(estilo['fill'])
                celula.border = copy(estilo['border'])
                celula.alignment = copy(estilo['alignment'])
                celula.number_format = estilo['number_format']
                celula.protection = copy(estilo['protection'])

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()

def gerar_txt_dominio(df):
    linhas_txt = []
    for _, row in df.iterrows():
        hist_limpo = limpar_caracteres_ilegais(str(row['HISTÓRICO'])).replace(';', ' ')
        linhas_txt.append(f"{row['DATA']};{row['DÉBITO'] if pd.notna(row['DÉBITO']) else ''};{row['CRÉDITO'] if pd.notna(row['CRÉDITO']) else ''};{float(row['VALOR']):.2f};{hist_limpo}\n")
    return "".join(linhas_txt)

def recuperar_xls_biff_irregular(file_bytes):
    """
    Recupera relatórios .xls antigos cujo contêiner OLE está legível, mas os
    endereços internos das planilhas estão inconsistentes. A correção ocorre
    somente em memória; o arquivo enviado pelo usuário não é alterado.
    """
    try:
        from xlrd.compdoc import CompDoc

        documento = CompDoc(file_bytes, ignore_workbook_corruption=True)
        fluxo_workbook = documento.get_named_stream('Workbook')
        if not fluxo_workbook:
            fluxo_workbook = documento.get_named_stream('Book')
        if not fluxo_workbook:
            return None

        fluxo = bytearray(fluxo_workbook)
        registros_abas = []
        posicao = 0
        fim_dos_globais = None

        # Localiza os registros BOUNDSHEET no bloco global do BIFF.
        while posicao + 4 <= len(fluxo):
            codigo, tamanho = struct.unpack_from('<HH', fluxo, posicao)
            fim_registro = posicao + 4 + tamanho
            if fim_registro > len(fluxo):
                break
            if codigo == 0x0085 and tamanho >= 4:
                registros_abas.append(posicao)
            posicao = fim_registro
            if codigo == 0x000A:
                fim_dos_globais = posicao
                break

        if not registros_abas or fim_dos_globais is None:
            return None

        # Procura os BOFs reais das planilhas, gráficos ou macros.
        inicios_reais = []
        cursor = fim_dos_globais
        while True:
            indice = fluxo.find(b'\x09\x08', cursor)
            if indice < 0:
                break
            if indice + 8 <= len(fluxo):
                tamanho = struct.unpack_from('<H', fluxo, indice + 2)[0]
                if tamanho >= 4 and indice + 4 + tamanho <= len(fluxo):
                    versao, tipo_fluxo = struct.unpack_from('<HH', fluxo, indice + 4)
                    if versao in (0x0500, 0x0600) and tipo_fluxo in (0x0010, 0x0020, 0x0040):
                        inicios_reais.append(indice)
            cursor = indice + 2

        if len(inicios_reais) < len(registros_abas):
            return None

        for registro_aba, inicio_real in zip(registros_abas, inicios_reais):
            struct.pack_into('<I', fluxo, registro_aba + 4, inicio_real)

        arquivo_recuperado = io.BytesIO(bytes(fluxo))
        xls = pd.ExcelFile(arquivo_recuperado, engine='xlrd')
        for nome_aba in xls.sheet_names:
            df_temp = pd.read_excel(
                xls, sheet_name=nome_aba, dtype=str, header=None
            )
            if df_temp is not None and not df_temp.empty and df_temp.shape[1] > 1:
                return df_temp
    except Exception:
        return None

    return None


def processar_razao_dominio(file_bytes, filename):
    df = None
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == '.xlsx':
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            for sheet in xls.sheet_names:
                df_temp = pd.read_excel(
                    xls, sheet_name=sheet, dtype=str, header=None
                )
                if df_temp is not None and not df_temp.empty and df_temp.shape[1] > 1:
                    df = df_temp
                    break
        elif ext == '.xls':
            try:
                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    dtype=str,
                    header=None,
                    engine='xlrd'
                )
            except Exception:
                # Alguns relatórios com extensão XLS são tabelas HTML.
                try:
                    tabelas = pd.read_html(io.BytesIO(file_bytes), header=None)
                    if tabelas:
                        df = tabelas[0].astype(str)
                except Exception:
                    df = None

                # Relatórios binários antigos da Domínio podem conter os dados
                # intactos, mas apontadores BIFF incorretos. Recuperamos tudo
                # em memória para evitar conversões manuais.
                if df is None or df.empty:
                    df = recuperar_xls_biff_irregular(file_bytes)
                    if df is not None and not df.empty:
                        st.session_state['razao_xls_recuperado'] = True
                    else:
                        st.session_state['erro_bof_xls'] = True
                        return None
        else:
            for enc in ['utf-8', 'latin1', 'cp1252']:
                for sep in [';', '\t', '|', ',']:
                    try:
                        df = pd.read_csv(
                            io.BytesIO(file_bytes),
                            sep=sep,
                            encoding=enc,
                            dtype=str,
                            header=None,
                            on_bad_lines='skip'
                        )
                        if df.shape[1] > 1:
                            break
                    except Exception:
                        continue
                if df is not None and df.shape[1] > 1:
                    break
    except Exception:
        return None

    if df is None or df.empty:
        return None

    header_row_idx = 0
    for idx, row in df.iterrows():
        row_str = " ".join(
            [str(v) for v in row.values if pd.notna(v)]
        ).upper()
        possui_data = 'DATA' in row_str or re.search(r'\bDT\b', row_str)
        possui_valor = any(
            termo in row_str
            for termo in ['VALOR', 'DEBITO', 'DÉBITO', 'CREDITO', 'CRÉDITO']
        )
        if possui_data and possui_valor:
            header_row_idx = idx
            break

    if header_row_idx > 0:
        df.columns = [
            str(v).strip().upper() for v in df.iloc[header_row_idx].values
        ]
        df = df.iloc[header_row_idx + 1:].copy()
    else:
        df.columns = [str(v).strip().upper() for v in df.iloc[0].values]
        df = df.iloc[1:].copy()

    df.columns = [re.sub(r'[^\w\s]', '', coluna) for coluna in df.columns]
    cols = list(df.columns)

    col_data = next(
        (c for c in cols if any(p in c for p in ['DATA', 'DT'])),
        None
    )
    col_deb = next(
        (c for c in cols if any(
            p in c for p in ['DEBITO', 'DÉBITO', 'SAIDA', 'DEB']
        )),
        None
    )
    col_cred = next(
        (c for c in cols if any(
            p in c for p in ['CREDITO', 'CRÉDITO', 'ENTRADA', 'CRE']
        )),
        None
    )
    col_val = next(
        (c for c in cols if any(p in c for p in ['VALOR', 'VL'])),
        None
    )
    col_hist = next(
        (c for c in cols if any(
            p in c
            for p in [
                'HISTORICO', 'HISTÓRICO', 'HIST', 'COMPLEMENTO',
                'LANCAMENTO', 'DESCRI'
            ]
        )),
        None
    )

    if not col_data:
        return None

    dados = []
    for _, row in df.iterrows():
        dt_raw = str(row[col_data]).strip() if pd.notna(row[col_data]) else ''

        # Datas lidas de XLS podem chegar como AAAA-MM-DD. Nesse caso não se
        # deve aplicar dayfirst, pois 2026-03-02 viraria 03/02/2026.
        if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', dt_raw):
            data_parseada = pd.to_datetime(
                dt_raw, yearfirst=True, errors='coerce'
            )
        else:
            match_dt = re.search(
                r'(?<!\d)(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})(?!\d)',
                dt_raw
            )
            data_parseada = (
                pd.to_datetime(
                    match_dt.group(1), dayfirst=True, errors='coerce'
                )
                if match_dt else pd.NaT
            )

        if pd.isna(data_parseada):
            continue
        dt_fmt = data_parseada.strftime('%d/%m/%Y')

        v_ent, v_sai = 0.0, 0.0

        if col_deb and col_cred:
            v_sai = (
                abs(limpar_valor_monetario(row[col_deb]))
                if pd.notna(row[col_deb]) else 0.0
            )
            v_ent = (
                abs(limpar_valor_monetario(row[col_cred]))
                if pd.notna(row[col_cred]) else 0.0
            )
        elif col_val and pd.notna(row[col_val]):
            val_num = limpar_valor_monetario(row[col_val])
            if val_num < 0:
                v_sai = abs(val_num)
            else:
                v_ent = val_num

        hist_str = (
            limpar_caracteres_ilegais(str(row[col_hist]).strip())
            if col_hist and pd.notna(row[col_hist])
            else 'LANCAMENTO RAZAO'
        )

        if v_ent != 0 or v_sai != 0:
            dados.append({
                'DATA': dt_fmt,
                'ENTRADAS_RAZAO': v_ent,
                'SAIDAS_RAZAO': v_sai,
                'HISTÓRICO': hist_str
            })

    if not dados:
        return None

    df_res = pd.DataFrame(dados)
    df_res['DATA_DT'] = pd.to_datetime(
        df_res['DATA'], dayfirst=True, errors='coerce'
    )
    return df_res.dropna(subset=['DATA_DT'])

def renderizar_base_inteligente_empresa(
    empresa, nome_empresa, bancos_permitidos, contas_bancarias
):
    """Base de Débito/Crédito isolada por empresa usando a mesma tabela Supabase."""
    url_base, chave_base, senha_admin = obter_config_classificacao_online()
    base = []
    erro_base = ''
    if url_base and chave_base:
        try:
            base = carregar_classificacoes_online(empresa)
        except Exception as erro:
            erro_base = str(erro)

    st.markdown(f"#### Base inteligente — {nome_empresa}")
    st.caption(
        "O aprendizado desta área é exclusivo desta empresa. Padrões de outras "
        "empresas não são usados aqui. Envie planilhas já revisadas, com DÉBITO e "
        "CRÉDITO preenchidos, para ensinar novos lançamentos."
    )
    st.caption(
        "Contas bancárias automáticas: "
        + " | ".join(
            f"{nome_banco_por_chave(banco)} {conta}"
            for banco, conta in contas_bancarias.items()
        )
    )
    if erro_base:
        st.warning(f"Não foi possível carregar a base online: {erro_base}")
    elif not url_base or not chave_base:
        st.warning("A conexão com a base online ainda não está configurada.")
    else:
        base_empresa = [item for item in base if item.get('banco') in bancos_permitidos]
        c1, c2 = st.columns(2)
        c1.metric("Padrões desta empresa", len(base_empresa))
        c2.metric(
            "Bancos com aprendizado",
            len({item.get('banco') for item in base_empresa if item.get('banco')})
        )

        arquivos_base = st.file_uploader(
            "Planilhas já classificadas",
            type=['xlsx', 'xls', 'zip'],
            accept_multiple_files=True,
            key=f"base_upload_{empresa}",
            help="Use somente arquivos revisados desta empresa."
        )
        senha_digitada = st.text_input(
            "Senha administrativa para gravar aprendizado",
            type="password",
            key=f"base_senha_{empresa}"
        ) if senha_admin else ''

        pode_gravar = bool(arquivos_base) and (
            not senha_admin or hmac.compare_digest(str(senha_digitada), str(senha_admin))
        )
        if arquivos_base and senha_admin and senha_digitada and not pode_gravar:
            st.error("Senha administrativa inválida.")

        if st.button(
            "Aprender com planilhas revisadas",
            key=f"base_aprender_{empresa}",
            disabled=not pode_gravar,
            use_container_width=True
        ):
            try:
                registros = importar_arquivos_classificados(arquivos_base, empresa)
                registros = [r for r in registros if r.get('banco') in bancos_permitidos]
                if not registros:
                    st.warning("Nenhum padrão válido foi encontrado para os bancos desta empresa.")
                else:
                    quantidade = salvar_classificacoes_online(registros, empresa)
                    st.success(f"{quantidade} padrões da {nome_empresa} foram gravados/atualizados.")
                    st.rerun()
            except Exception as erro:
                st.error(f"Não foi possível atualizar a base: {erro}")

        if not base_empresa:
            st.info("Esta empresa ainda não possui padrões aprendidos.")

        st.markdown("---")
        st.markdown("#### Classificar planilha final conciliada")
        st.caption(
            "Anexe somente a planilha final depois da conferência bancária. "
            "A classificação usa exclusivamente a Base Inteligente desta empresa."
        )
        planilha_final = st.file_uploader(
            "Planilha final com os saldos conferidos",
            type=['xlsx'],
            key=f"base_planilha_final_{empresa}"
        )
        if planilha_final:
            if erro_base:
                st.error("A base online precisa estar conectada antes da classificação.")
            elif not base_empresa:
                st.warning(
                    "A base desta empresa ainda não possui padrões. Importe primeiro "
                    "planilhas antigas já classificadas desta mesma empresa."
                )
            else:
                try:
                    arquivo_classificado, resumo = executar_com_loading(
                        "Analisando históricos e classificando as contas...",
                        classificar_planilha_final,
                        planilha_final.getvalue(),
                        planilha_final.name,
                        base_empresa,
                        contas_bancarias
                    )
                    m1, m2, m3 = st.columns(3)
                    m1.metric(
                        "Classificados automaticamente",
                        f"{int(resumo.get('automaticos', 0)):,}".replace(',', '.')
                    )
                    m2.metric(
                        "Por nome da empresa",
                        f"{int(resumo.get('por_nome_empresa', 0)):,}".replace(',', '.')
                    )
                    m3.metric(
                        "Padrões novos",
                        f"{int(resumo.get('padroes_novos', 0)):,}".replace(',', '.')
                    )
                    renderizar_revisao_inteligente(
                        arquivo_classificado,
                        planilha_final.getvalue(),
                        planilha_final.name,
                        empresa,
                        contas_bancarias,
                        senha_admin,
                        f"base_revisao_{empresa}"
                    )
                except Exception as erro_classificacao:
                    st.error(
                        "Não foi possível classificar a planilha final: "
                        f"{erro_classificacao}"
                    )

# ==============================================================================
# ORGANIZADORES ESPECÍFICOS POR EMPRESA
# ==============================================================================
def texto_celula_seguro(valor):
    if valor is None or pd.isna(valor): return ""
    if isinstance(valor, float) and valor.is_integer(): return str(int(valor))
    return limpar_caracteres_ilegais(str(valor)).strip()

def identificar_estorno_de_baixa(*campos):
    """Reconhece apenas estornos ligados a baixa, preservando outros estornos."""
    texto = normalizar_texto(" ".join(texto_celula_seguro(c) for c in campos))
    tokens = re.findall(r'[a-z0-9]+', texto)
    pos_estorno = [i for i, token in enumerate(tokens) if token.startswith(('estorn', 'revers'))]
    pos_baixa = [i for i, token in enumerate(tokens) if token.startswith('baix')]
    return any(abs(i - j) <= 6 for i in pos_estorno for j in pos_baixa)

def criar_assinatura_classificacao(historico):
    """Remove documentos variáveis e preserva natureza, empresa e observação útil."""
    texto = normalizar_texto(texto_celula_seguro(historico))
    texto = re.sub(r'\b(?:pagar|pagamento)\b', 'pago', texto)
    texto = re.sub(r'\b(?:receber|recebimento)\b', 'recebido', texto)
    natureza = (
        'pago' if re.search(r'\bpago\b', texto)
        else 'recebido' if re.search(r'\brecebido\b', texto)
        else 'outro'
    )
    empresa_match = re.search(r'empresa\s*:\s*(.*?)(?:\s+obs\s*:|$)', texto)
    empresa = empresa_match.group(1) if empresa_match else texto
    empresa = re.sub(r'[^a-z0-9]+', ' ', empresa)
    tokens_empresa = [
        token for token in empresa.split()
        if not (any(c.isdigit() for c in token) and len(token) >= 3)
    ]
    empresa = ' '.join(tokens_empresa).strip()

    observacao = ''
    observacao_match = re.search(r'\s+obs\s*:\s*(.*)$', texto)
    if observacao_match:
        observacao_bruta = re.sub(r'[^a-z0-9]+', ' ', observacao_match.group(1))
        # Números, documentos, parcelas e datas mudam a cada mês. Palavras como
        # BANCO FIBRA, TRANSFERENCIA ou DEVOLUCAO alteram a classificação e ficam.
        tokens_observacao = [
            token for token in observacao_bruta.split()
            if token.isalpha() and token not in {'doc', 'documento'}
        ]
        observacao = ' '.join(tokens_observacao).strip()

    partes = [natureza, empresa]
    if observacao:
        partes.append(observacao)
    return '|'.join(partes) if empresa else ''

def decompor_assinatura_classificacao(assinatura):
    partes = str(assinatura or '').split('|')
    natureza = partes[0] if partes else ''
    empresa = partes[1] if len(partes) > 1 else ''
    observacao = '|'.join(partes[2:]) if len(partes) > 2 else ''
    return natureza, empresa, observacao

def extrair_nome_empresa_classificacao(historico, assinatura=''):
    """
    Extrai a entidade do histórico e elimina somente ruídos que mudam por mês,
    preservando as palavras que realmente identificam o fornecedor ou cliente.
    """
    texto = normalizar_texto(texto_celula_seguro(historico))
    empresa_match = re.search(
        r'empresa\s*:\s*(.*?)(?:\s+obs\s*:|$)', texto
    )
    if empresa_match:
        empresa = empresa_match.group(1)
    elif assinatura:
        _, empresa, _ = decompor_assinatura_classificacao(assinatura)
    else:
        empresa = re.sub(
            r'^\s*(?:pago|pagar|pagamento|recebido|receber|recebimento)\b',
            ' ',
            texto
        )
        # Remove um documento variável que apareça logo após a natureza.
        empresa = re.sub(
            r'^\s*[a-z0-9./-]*\d[a-z0-9./-]*\s+',
            ' ',
            empresa
        )

    tokens = re.findall(r'[a-z0-9]+', empresa)
    termos_ruido = {
        'ltda', 'limitada', 'eireli', 'epp', 'me', 'mei', 'sa', 's', 'a',
        'sociedade', 'anonima', 'unipessoal', 'de', 'da', 'do', 'das', 'dos',
        'e', 'doc', 'documento', 'nf', 'nfe', 'nota', 'fiscal', 'pedido',
        'parcela', 'pagto', 'pago', 'recebido', 'empresa', 'obs'
    }
    tokens_validos = []
    for token in tokens:
        if token in termos_ruido or token.isdigit() or len(token) < 2:
            continue
        # Protocolos mistos longos também variam entre os períodos. Marcas
        # curtas com números, como 3M, continuam preservadas.
        if any(caractere.isdigit() for caractere in token) and len(token) >= 4:
            continue
        tokens_validos.append(token)
    return ' '.join(tokens_validos).strip()


def normalizar_nome_empresa_classificacao(nome):
    """Normaliza um nome já extraído para comparação conservadora."""
    return extrair_nome_empresa_classificacao(nome)


def calcular_similaridade_nome_empresa(nome_a, nome_b):
    """
    Compara nomes por caracteres e palavras. Nomes de uma única palavra só
    passam quando são exatamente iguais, evitando aproximações arriscadas.
    """
    a = normalizar_nome_empresa_classificacao(nome_a)
    b = normalizar_nome_empresa_classificacao(nome_b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    tokens_a, tokens_b = set(a.split()), set(b.split())
    if min(len(tokens_a), len(tokens_b)) < 2:
        return 0.0

    comuns = tokens_a & tokens_b
    if not comuns:
        return 0.0

    # Algumas palavras e siglas mudam a natureza contábil mesmo quando o nome
    # principal é igual. Elas precisam coincidir exatamente na aproximação.
    termos_criticos = {
        'antecipado', 'antecipada', 'adiantamento', 'ferias', 'rescisao',
        'salario', 'folha', 'inss', 'fgts', 'irrf', 'pis', 'cofins', 'csll',
        'icms', 'iss', 'gnre', 'imposto', 'tributo', 'tarifa', 'juros',
        'multa', 'aluguel', 'frete', 'transferencia', 'transf', 'devolucao',
        'emprestimo', 'aplicacao', 'rendimento', 'filial', 'matriz'
    }
    if (tokens_a & termos_criticos) != (tokens_b & termos_criticos):
        return 0.0

    # Siglas curtas costumam identificar empresas, estados ou tipos de operação.
    # BS e BD, por exemplo, não podem ser tratadas como simples erro de digitação.
    siglas_a = {token for token in tokens_a if len(token) <= 3}
    siglas_b = {token for token in tokens_b if len(token) <= 3}
    if siglas_a != siglas_b:
        return 0.0

    razao_caracteres = difflib.SequenceMatcher(None, a, b).ratio()
    cobertura = len(comuns) / min(len(tokens_a), len(tokens_b))
    uniao = tokens_a | tokens_b
    jaccard = len(comuns) / len(uniao) if uniao else 0.0

    # Um nome completo contido no outro é um sinal forte, desde que existam
    # pelo menos duas palavras distintivas em comum.
    if (a in b or b in a) and len(comuns) >= 2:
        return max(0.97, razao_caracteres)

    # Aceita pequenas variações de escrita ou complemento, mas nunca apenas
    # porque duas empresas compartilham uma palavra genérica.
    if len(comuns) >= 2 and cobertura >= 0.80 and razao_caracteres >= 0.72:
        return max(0.95, (razao_caracteres * 0.55) + (cobertura * 0.35) + (jaccard * 0.10))
    if len(comuns) >= 2 and cobertura >= 0.67 and razao_caracteres >= 0.90:
        return max(0.94, razao_caracteres)
    if len(comuns) >= 1 and razao_caracteres >= 0.95 and min(len(a), len(b)) >= 10:
        return razao_caracteres
    return 0.0


def encontrar_conta_por_nome_historico(historico, natureza, evidencias_nomes):
    """
    Retorna uma conta somente quando o nome foi confirmado em períodos
    diferentes, pertence à mesma natureza e não possui classificação concorrente.
    """
    nome_alvo = extrair_nome_empresa_classificacao(historico)
    if natureza not in {'pago', 'recebido'} or len(nome_alvo) < 4:
        return None, 0.0, 'nome_insuficiente'

    entidades_natureza = evidencias_nomes.get(natureza, {})
    evidencia_exata = entidades_natureza.get(nome_alvo)
    if evidencia_exata:
        if len(evidencia_exata) != 1:
            return None, 1.0, 'conflito'
        conta, evidencia = next(iter(evidencia_exata.items()))
        if len(evidencia['periodos']) >= 2:
            return conta, 1.0, 'nome_exato'
        # Um único mês não ensina uma regra contábil com segurança.
        return None, 1.0, 'evidencia_insuficiente'

    melhor_por_conta = {}
    for nome_base, contas in entidades_natureza.items():
        if len(contas) != 1:
            continue
        conta, evidencia = next(iter(contas.items()))
        # Aproximação textual exige uma conta repetida nos três meses da base.
        # Nome exatamente igual continua podendo ser confirmado em dois períodos.
        if len(evidencia['periodos']) < 3 or evidencia['ocorrencias'] < 3:
            continue
        similaridade = calcular_similaridade_nome_empresa(nome_alvo, nome_base)
        if similaridade > melhor_por_conta.get(conta, 0.0):
            melhor_por_conta[conta] = similaridade

    resultados = sorted(
        melhor_por_conta.items(), key=lambda item: item[1], reverse=True
    )
    if not resultados or resultados[0][1] < 0.94:
        return None, resultados[0][1] if resultados else 0.0, 'sem_correspondencia'

    melhor_conta, melhor_nota = resultados[0]
    segunda_nota = resultados[1][1] if len(resultados) > 1 else 0.0
    if segunda_nota >= 0.90 and (melhor_nota - segunda_nota) < 0.08:
        return None, melhor_nota, 'conflito'
    return melhor_conta, melhor_nota, 'nome_aproximado'


def obter_config_classificacao_online():
    """Obtém somente no servidor as credenciais guardadas em st.secrets."""
    try:
        secao = st.secrets.get('supabase', {})
        url = secao.get('url', '') or st.secrets.get('SUPABASE_URL', '')
        chave = secao.get('service_key', '') or st.secrets.get('SUPABASE_SERVICE_KEY', '')
        senha = secao.get('admin_password', '') or st.secrets.get(
            'CLASSIFICATION_ADMIN_PASSWORD', ''
        )
        url = str(url).strip()
        chave = str(chave).strip()
        senha = str(senha)
        placeholders_url = {'URL_DO_PROJETO_SUPABASE', 'SUA_URL_SUPABASE', 'SUPABASE_URL'}
        placeholders_chave = {
            'SERVICE_ROLE_KEY', 'SUA_SERVICE_ROLE_KEY', 'SUPABASE_SERVICE_KEY'
        }
        if url.upper() in placeholders_url or not url.startswith(('https://', 'http://')):
            url = ''
        elif url:
            partes_url = urllib.parse.urlsplit(url)
            caminho_url = partes_url.path or ''
            if '/rest/v1' in caminho_url:
                caminho_url = caminho_url.split('/rest/v1', 1)[0]
            url = urllib.parse.urlunsplit((
                partes_url.scheme,
                partes_url.netloc,
                caminho_url.rstrip('/'),
                '',
                ''
            )).rstrip('/')
        if chave.upper() in placeholders_chave:
            chave = ''
        return url, chave, senha
    except Exception:
        return '', '', ''

def requisicao_classificacao_online(caminho, metodo='GET', dados=None, prefer=''):
    url_base, chave, _ = obter_config_classificacao_online()
    if not url_base or not chave:
        raise RuntimeError(
            'A base online ainda não foi configurada com a URL e a chave reais do Supabase.'
        )
    corpo = json.dumps(dados, ensure_ascii=False).encode('utf-8') if dados is not None else None
    cabecalhos = {
        'apikey': chave,
        'Authorization': f'Bearer {chave}',
        'Content-Type': 'application/json',
    }
    if prefer:
        cabecalhos['Prefer'] = prefer
    requisicao = urllib.request.Request(
        f"{url_base.rstrip('/')}/rest/v1/{caminho}",
        data=corpo,
        headers=cabecalhos,
        method=metodo
    )
    try:
        with urllib.request.urlopen(requisicao, timeout=30) as resposta:
            conteudo = resposta.read().decode('utf-8')
            return json.loads(conteudo) if conteudo else []
    except urllib.error.HTTPError as erro:
        detalhe = erro.read().decode('utf-8', errors='ignore')
        raise RuntimeError(f"Falha na base online ({erro.code}): {detalhe[:300]}") from erro
    except urllib.error.URLError as erro:
        raise RuntimeError(f"Não foi possível acessar a base online: {erro.reason}") from erro

@st.cache_data(show_spinner=False, ttl=120, max_entries=20)
def carregar_classificacoes_online(empresa='nova_geracao'):
    registros, deslocamento, limite = [], 0, 1000
    while True:
        consulta = (
            'classificacoes_bancarias?empresa=eq.'
            + urllib.parse.quote(empresa)
            + '&select=id,empresa,banco,assinatura,debito,credito,ocorrencias,periodos,exemplo_historico'
            + f'&limit={limite}&offset={deslocamento}'
        )
        lote = requisicao_classificacao_online(consulta)
        registros.extend(lote)
        if len(lote) < limite:
            break
        deslocamento += limite
    return registros

def apagar_classificacoes_online(empresa):
    """Remove todos os padrões de uma empresa específica da base online."""
    if not empresa:
        return 0
    existentes = carregar_classificacoes_online(empresa)
    if not existentes:
        return 0
    caminho = (
        'classificacoes_bancarias?empresa=eq.'
        + urllib.parse.quote(empresa)
    )
    requisicao_classificacao_online(
        caminho,
        metodo='DELETE',
        prefer='return=minimal'
    )
    carregar_classificacoes_online.clear()
    return len(existentes)

def salvar_classificacoes_online(registros, empresa='nova_geracao'):
    if not registros:
        return 0
    existentes = {
        item['id']: item for item in carregar_classificacoes_online(empresa)
    }
    for registro in registros:
        anterior = existentes.get(registro['id'], {})
        registro['ocorrencias'] = max(
            int(registro.get('ocorrencias') or 1), int(anterior.get('ocorrencias') or 0)
        )
        registro['periodos'] = sorted(set(
            (registro.get('periodos') or []) + (anterior.get('periodos') or [])
        ))
    for inicio in range(0, len(registros), 500):
        requisicao_classificacao_online(
            'classificacoes_bancarias?on_conflict=id',
            metodo='POST',
            dados=registros[inicio:inicio + 500],
            prefer='resolution=merge-duplicates,return=minimal'
        )
    carregar_classificacoes_online.clear()
    return len(registros)

def ler_planilha_classificada(file_bytes, filename, empresa='nova_geracao'):
    """Lê planilha revisada e cria padrões exclusivos da empresa informada."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    registros = []
    banco_arquivo = identificar_chave_banco_empresa(filename)
    for nome_aba in xls.sheet_names:
        bruto = pd.read_excel(xls, sheet_name=nome_aba, header=None, dtype=object)
        indice_cabecalho = None
        for indice in range(min(30, len(bruto))):
            nomes = [normalizar_texto(texto_celula_seguro(v)).strip() for v in bruto.iloc[indice]]
            if all(nome in nomes for nome in ['data', 'valor', 'debito', 'credito']) and (
                'historico' in nomes
            ):
                indice_cabecalho = indice
                break
        if indice_cabecalho is None:
            continue
        cabecalhos = [texto_celula_seguro(v) for v in bruto.iloc[indice_cabecalho]]
        df = bruto.iloc[indice_cabecalho + 1:].copy()
        df.columns = cabecalhos
        mapa = {normalizar_texto(str(c)).strip(): c for c in df.columns}
        col_hist = mapa.get('historico')
        col_data = mapa.get('data')
        col_debito = mapa.get('debito')
        col_credito = mapa.get('credito')
        col_descricao = mapa.get('descricao')
        if col_hist is None or col_debito is None or col_credito is None:
            continue
        banco_aba = identificar_chave_banco_empresa(nome_aba)
        for _, linha in df.iterrows():
            historico = texto_celula_seguro(linha[col_hist])
            debito = texto_celula_seguro(linha[col_debito])
            credito = texto_celula_seguro(linha[col_credito])
            if not historico or not debito or not credito:
                continue
            banco_linha = (
                identificar_chave_banco_empresa(linha[col_descricao])
                if col_descricao is not None else ''
            ) or banco_aba or banco_arquivo
            assinatura = criar_assinatura_classificacao(historico)
            if banco_linha not in {'itau', 'bradesco', 'fibra', 'daycoval', 'sicredi'} or not assinatura:
                continue
            data_lancamento = (
                pd.to_datetime(linha[col_data], dayfirst=True, errors='coerce')
                if col_data is not None else pd.NaT
            )
            periodo = (
                data_lancamento.strftime('%Y-%m') if not pd.isna(data_lancamento)
                else normalizar_texto(filename)
            )
            identificador = hashlib.sha256(
                f"{empresa}|{banco_linha}|{assinatura}|{debito}|{credito}".encode('utf-8')
            ).hexdigest()
            registros.append({
                'id': identificador,
                'empresa': empresa,
                'banco': banco_linha,
                'assinatura': assinatura,
                'debito': debito,
                'credito': credito,
                'ocorrencias': 1,
                'periodos': [periodo],
                'exemplo_historico': historico[:500]
            })
    return registros

def importar_arquivos_classificados(arquivos, empresa='nova_geracao'):
    """Aceita XLSX/ZIP e mantém o aprendizado isolado por empresa."""
    registros = []
    for arquivo in arquivos:
        conteudo = arquivo.getvalue()
        nome = arquivo.name
        if nome.lower().endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(conteudo)) as pacote:
                membros = [
                    membro for membro in pacote.infolist()
                    if not membro.is_dir() and membro.filename.lower().endswith(('.xlsx', '.xls'))
                ]
                if sum(membro.file_size for membro in membros) > 60 * 1024 * 1024:
                    raise ValueError('O conteúdo descompactado ultrapassa o limite de 60 MB.')
                for membro in membros:
                    if membro.file_size > 20 * 1024 * 1024:
                        raise ValueError(f'A planilha {membro.filename} ultrapassa 20 MB.')
                    registros.extend(ler_planilha_classificada(
                        pacote.read(membro), os.path.basename(membro.filename), empresa
                    ))
        else:
            registros.extend(ler_planilha_classificada(conteudo, nome, empresa))

    agrupados = {}
    for registro in registros:
        chave = registro['id']
        if chave not in agrupados:
            agrupados[chave] = registro
        else:
            agrupados[chave]['ocorrencias'] += 1
            agrupados[chave]['periodos'] = sorted(set(
                agrupados[chave]['periodos'] + registro['periodos']
            ))
    return list(agrupados.values())

def aplicar_classificacoes_automaticas(df, banco, base_classificacoes):
    """Preenche apenas padrões repetidos e com uma única classificação conhecida."""
    resultado = df.copy()
    resultado['_CLASSIFICAÇÃO'] = 'Pendente'
    candidatos = {}
    periodos_por_assinatura = {}
    for item in base_classificacoes:
        if item.get('banco') != banco:
            continue
        assinatura = item.get('assinatura', '')
        par = (texto_celula_seguro(item.get('debito')), texto_celula_seguro(item.get('credito')))
        if assinatura and all(par):
            candidatos.setdefault(assinatura, set()).add(par)
            periodos_por_assinatura.setdefault(assinatura, set()).update(
                item.get('periodos') or []
            )
    mapa_seguro = {
        assinatura: next(iter(pares))
        for assinatura, pares in candidatos.items()
        if len(pares) == 1 and len(periodos_por_assinatura.get(assinatura, set())) >= 3
    }
    for indice, linha in resultado.iterrows():
        if texto_celula_seguro(linha.get('DÉBITO')) or texto_celula_seguro(linha.get('CRÉDITO')):
            resultado.at[indice, '_CLASSIFICAÇÃO'] = 'Já preenchido'
            continue
        assinatura = criar_assinatura_classificacao(linha.get('HISTÓRICO', ''))
        if assinatura in mapa_seguro:
            debito, credito = mapa_seguro[assinatura]
            resultado.at[indice, 'DÉBITO'] = debito
            resultado.at[indice, 'CRÉDITO'] = credito
            resultado.at[indice, '_CLASSIFICAÇÃO'] = 'Automática'
        elif assinatura in candidatos and len(candidatos[assinatura]) > 1:
            resultado.at[indice, '_CLASSIFICAÇÃO'] = 'Revisar conflito'
        elif assinatura in candidatos:
            resultado.at[indice, '_CLASSIFICAÇÃO'] = 'Revisar padrão novo'
    return resultado

def classificar_planilha_final(
    file_bytes, filename, base_classificacoes, contas_bancarias=None
):
    """Preenche Débito/Crédito somente na planilha final já conciliada."""
    from openpyxl import load_workbook

    if not filename.lower().endswith('.xlsx'):
        raise ValueError('A planilha final deve estar no formato .xlsx.')

    contas_bancarias = contas_bancarias or {
        'itau': '508', 'bradesco': '9', 'fibra': '506'
    }
    candidatos_por_banco = {}
    periodos_por_banco = {}
    evidencias_nomes_por_banco = {}
    for item in base_classificacoes:
        banco = item.get('banco', '')
        assinatura = item.get('assinatura', '')
        debito_item = texto_celula_seguro(item.get('debito'))
        credito_item = texto_celula_seguro(item.get('credito'))
        conta_banco_item = texto_celula_seguro(contas_bancarias.get(banco, ''))

        # Para padrões já aprendidos, a posição REAL da conta bancária é mais
        # confiável que palavras como "pago" ou "recebido" presentes no histórico.
        if banco in contas_bancarias and conta_banco_item:
            if credito_item == conta_banco_item and debito_item:
                natureza = 'pago'
                contrapartida = debito_item
            elif debito_item == conta_banco_item and credito_item:
                natureza = 'recebido'
                contrapartida = credito_item
            else:
                natureza = assinatura.split('|', 1)[0] if assinatura else ''
                contrapartida = ''
        else:
            natureza = assinatura.split('|', 1)[0] if assinatura else ''
            contrapartida = ''

        # Normaliza também a assinatura existente para a natureza real inferida
        # pela posição da conta do banco. Isso reaproveita a base antiga sem zerar.
        if assinatura and natureza in {'pago', 'recebido'}:
            partes_assinatura = assinatura.split('|', 1)
            sufixo_assinatura = partes_assinatura[1] if len(partes_assinatura) > 1 else ''
            assinatura = (
                f"{natureza}|{sufixo_assinatura}" if sufixo_assinatura else natureza
            )
        if banco not in contas_bancarias or not assinatura or not contrapartida:
            continue
        candidatos_por_banco.setdefault(banco, {}).setdefault(assinatura, set()).add(
            contrapartida
        )
        periodos_item = set(item.get('periodos') or [])
        periodos_por_banco.setdefault(banco, {}).setdefault(assinatura, set()).update(
            periodos_item
        )

        historico_exemplo = item.get('exemplo_historico') or ''
        nome_empresa = extrair_nome_empresa_classificacao(
            historico_exemplo, assinatura
        )
        if nome_empresa:
            por_natureza = evidencias_nomes_por_banco.setdefault(
                banco, {}
            ).setdefault(natureza, {})
            por_conta = por_natureza.setdefault(nome_empresa, {}).setdefault(
                contrapartida,
                {'periodos': set(), 'ocorrencias': 0}
            )
            por_conta['periodos'].update(periodos_item)
            try:
                ocorrencias_item = max(1, int(item.get('ocorrencias') or 1))
            except (TypeError, ValueError):
                ocorrencias_item = 1
            por_conta['ocorrencias'] += ocorrencias_item

    mapas_seguros = {}
    for banco, candidatos in candidatos_por_banco.items():
        mapas_seguros[banco] = {
            assinatura: next(iter(contas))
            for assinatura, contas in candidatos.items()
            if len(contas) == 1 and len(
                periodos_por_banco.get(banco, {}).get(assinatura, set())
            ) >= 3
        }

    def valor_conta_excel(conta):
        texto = texto_celula_seguro(conta)
        if texto.isdigit() and (texto == '0' or not texto.startswith('0')):
            return int(texto)
        return texto

    wb = load_workbook(io.BytesIO(file_bytes))
    resumo = {
        'automaticos': 0,
        'somente_banco': 0,
        'antecipados': 0,
        'por_nome_empresa': 0,
        'ja_preenchidos': 0,
        'parciais_completados': 0,
        'conflitos': 0,
        'padroes_novos': 0,
        'banco_nao_identificado': 0,
        'abas_processadas': 0,
    }
    banco_arquivo = identificar_chave_banco_empresa(filename)
    cache_similaridade = {}

    for ws in wb.worksheets:
        if 'retir' in normalizar_texto(ws.title):
            continue
        linha_cabecalho = None
        mapa_colunas = {}
        for numero_linha in range(1, min(ws.max_row, 30) + 1):
            mapa_teste = {
                normalizar_texto(texto_celula_seguro(ws.cell(numero_linha, coluna).value)).strip(): coluna
                for coluna in range(1, ws.max_column + 1)
            }
            if all(nome in mapa_teste for nome in ['historico', 'debito', 'credito']):
                linha_cabecalho = numero_linha
                mapa_colunas = mapa_teste
                break
        if linha_cabecalho is None:
            continue

        resumo['abas_processadas'] += 1
        col_hist = mapa_colunas['historico']
        col_debito = mapa_colunas['debito']
        col_credito = mapa_colunas['credito']
        col_valor = mapa_colunas.get('valor')
        col_descricao = mapa_colunas.get('descricao')
        banco_aba = identificar_chave_banco_empresa(ws.title) or banco_arquivo

        for numero_linha in range(linha_cabecalho + 1, ws.max_row + 1):
            historico = texto_celula_seguro(ws.cell(numero_linha, col_hist).value)
            if not historico:
                continue
            debito_atual = texto_celula_seguro(ws.cell(numero_linha, col_debito).value)
            credito_atual = texto_celula_seguro(ws.cell(numero_linha, col_credito).value)
            if debito_atual and credito_atual:
                resumo['ja_preenchidos'] += 1
                continue
            linha_estava_parcial = bool(debito_atual or credito_atual)

            banco_linha = (
                identificar_chave_banco_empresa(ws.cell(numero_linha, col_descricao).value)
                if col_descricao is not None else ''
            ) or banco_aba
            if banco_linha not in contas_bancarias:
                resumo['banco_nao_identificado'] += 1
                continue

            assinatura = criar_assinatura_classificacao(historico)

            # REGRA PRINCIPAL: o sinal do VALOR decide a natureza.
            # Nunca usamos uma palavra perdida no histórico para decidir se o banco
            # entra no débito ou no crédito.
            valor_linha = (
                limpar_valor_monetario(ws.cell(numero_linha, col_valor).value)
                if col_valor is not None else 0.0
            )
            if valor_linha < 0:
                natureza = 'pago'
            elif valor_linha > 0:
                natureza = 'recebido'
            else:
                natureza = assinatura.split('|', 1)[0] if assinatura else ''

            # A assinatura usada para procurar a contrapartida também recebe a
            # natureza definida pelo sinal, evitando que "pago" dentro do histórico
            # faça um recebimento buscar padrões de pagamento (e vice-versa).
            if assinatura and natureza in {'pago', 'recebido'}:
                partes_assinatura = assinatura.split('|', 1)
                sufixo_assinatura = partes_assinatura[1] if len(partes_assinatura) > 1 else ''
                assinatura = (
                    f"{natureza}|{sufixo_assinatura}" if sufixo_assinatura else natureza
                )

            conta_banco = contas_bancarias[banco_linha]
            if natureza == 'pago':
                if not credito_atual:
                    ws.cell(numero_linha, col_credito).value = valor_conta_excel(conta_banco)
                coluna_contrapartida = col_debito
            elif natureza == 'recebido':
                if not debito_atual:
                    ws.cell(numero_linha, col_debito).value = valor_conta_excel(conta_banco)
                coluna_contrapartida = col_credito
            else:
                resumo['padroes_novos'] += 1
                continue

            contrapartida_atual = texto_celula_seguro(
                ws.cell(numero_linha, coluna_contrapartida).value
            )
            if contrapartida_atual:
                resumo['automaticos'] += 1
                if linha_estava_parcial:
                    resumo['parciais_completados'] += 1
                continue

            candidatos_banco = candidatos_por_banco.get(banco_linha, {})
            mapas_banco = mapas_seguros.get(banco_linha, {})
            candidatos = candidatos_banco.get(assinatura, set())
            conta_segura = mapas_banco.get(assinatura)

            # Quando a base foi aprendida a partir de histórico cru da Autokraft,
            # a assinatura fica "outro|...". Mantemos o restante da assinatura e
            # localizamos esse padrão também.
            if not candidatos and assinatura:
                partes_assinatura = assinatura.split('|', 1)
                sufixo_assinatura = partes_assinatura[1] if len(partes_assinatura) > 1 else ''
                assinatura_outro = f"outro|{sufixo_assinatura}" if sufixo_assinatura else assinatura
                candidatos = candidatos_banco.get(assinatura_outro, set())
                conta_segura = mapas_banco.get(assinatura_outro)
            if 'antecipad' in normalizar_texto(historico):
                ws.cell(numero_linha, coluna_contrapartida).value = 532
                resumo['automaticos'] += 1
                resumo['antecipados'] += 1
                if linha_estava_parcial:
                    resumo['parciais_completados'] += 1
            elif conta_segura:
                ws.cell(numero_linha, coluna_contrapartida).value = valor_conta_excel(
                    conta_segura
                )
                resumo['automaticos'] += 1
                if linha_estava_parcial:
                    resumo['parciais_completados'] += 1
            elif len(candidatos) > 1:
                resumo['conflitos'] += 1
                resumo['somente_banco'] += 1
            else:
                chave_cache = (banco_linha, natureza, assinatura)
                if chave_cache not in cache_similaridade:
                    cache_similaridade[chave_cache] = encontrar_conta_por_nome_historico(
                        historico,
                        natureza,
                        evidencias_nomes_por_banco.get(banco_linha, {})
                    )
                conta_similar, _, motivo_nome = cache_similaridade[chave_cache]
                if conta_similar:
                    ws.cell(numero_linha, coluna_contrapartida).value = valor_conta_excel(
                        conta_similar
                    )
                    resumo['automaticos'] += 1
                    resumo['por_nome_empresa'] += 1
                    if linha_estava_parcial:
                        resumo['parciais_completados'] += 1
                else:
                    if motivo_nome == 'conflito':
                        resumo['conflitos'] += 1
                    else:
                        resumo['padroes_novos'] += 1
                    resumo['somente_banco'] += 1

    if resumo['abas_processadas'] == 0:
        raise ValueError(
            'Nenhuma aba com as colunas HISTÓRICO, DÉBITO e CRÉDITO foi encontrada.'
        )

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue(), resumo

def extrair_pendencias_revisao_inteligente(file_bytes, contas_bancarias):
    """Lista somente lançamentos que ainda precisam da conta de contrapartida."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    pendencias = []
    contas_bancarias = contas_bancarias or {}

    for ws in wb.worksheets:
        if 'retir' in normalizar_texto(ws.title):
            continue

        linha_cabecalho = None
        mapa_colunas = {}
        for numero_linha in range(1, min(ws.max_row, 30) + 1):
            mapa_teste = {
                normalizar_texto(texto_celula_seguro(ws.cell(numero_linha, coluna).value)).strip(): coluna
                for coluna in range(1, ws.max_column + 1)
            }
            if all(nome in mapa_teste for nome in ['historico', 'debito', 'credito']):
                linha_cabecalho = numero_linha
                mapa_colunas = mapa_teste
                break
        if linha_cabecalho is None:
            continue

        col_hist = mapa_colunas['historico']
        col_debito = mapa_colunas['debito']
        col_credito = mapa_colunas['credito']
        col_valor = mapa_colunas.get('valor')
        col_data = mapa_colunas.get('data')
        col_descricao = mapa_colunas.get('descricao')
        banco_aba = identificar_chave_banco_empresa(ws.title)

        for numero_linha in range(linha_cabecalho + 1, ws.max_row + 1):
            historico = texto_celula_seguro(ws.cell(numero_linha, col_hist).value)
            if not historico:
                continue

            banco = (
                identificar_chave_banco_empresa(ws.cell(numero_linha, col_descricao).value)
                if col_descricao is not None else ''
            ) or banco_aba
            if banco not in contas_bancarias:
                continue

            debito = texto_celula_seguro(ws.cell(numero_linha, col_debito).value)
            credito = texto_celula_seguro(ws.cell(numero_linha, col_credito).value)
            if debito and credito:
                continue

            assinatura = criar_assinatura_classificacao(historico)
            valor = 0.0
            if col_valor is not None:
                valor = limpar_valor_monetario(ws.cell(numero_linha, col_valor).value)

            # A Revisão Inteligente segue exatamente a mesma regra da classificação.
            if valor < 0:
                natureza = 'pago'
            elif valor > 0:
                natureza = 'recebido'
            else:
                natureza = assinatura.split('|', 1)[0] if assinatura else ''

            if natureza == 'pago':
                coluna_destino = col_debito
                coluna_banco = col_credito
                lado = 'DÉBITO'
            elif natureza == 'recebido':
                coluna_destino = col_credito
                coluna_banco = col_debito
                lado = 'CRÉDITO'
            else:
                continue

            contrapartida_atual = texto_celula_seguro(
                ws.cell(numero_linha, coluna_destino).value
            )
            if contrapartida_atual:
                continue

            data_texto = ''
            if col_data is not None:
                data_raw = ws.cell(numero_linha, col_data).value
                data_parseada = pd.to_datetime(data_raw, dayfirst=True, errors='coerce')
                if not pd.isna(data_parseada):
                    data_texto = data_parseada.strftime('%d/%m/%Y')
                else:
                    data_texto = texto_celula_seguro(data_raw)

            pendencias.append({
                'Banco': nome_banco_por_chave(banco),
                'Data': data_texto,
                'Valor': valor,
                'Histórico': historico,
                'Classificar em': lado,
                'Conta bancária': texto_celula_seguro(contas_bancarias.get(banco, '')),
                'Conta da contrapartida': '',
                '_aba': ws.title,
                '_linha': numero_linha,
                '_col_destino': coluna_destino,
                '_col_banco': coluna_banco,
                '_banco': banco,
                '_col_data': col_data or 0,
            })

    return pd.DataFrame(pendencias)


def aplicar_revisoes_inteligentes(
    file_bytes, revisoes, filename, empresa, contas_bancarias
):
    """Aplica as contas revisadas e gera somente os novos padrões confirmados pelo usuário."""
    from openpyxl import load_workbook

    if revisoes is None or revisoes.empty:
        return file_bytes, 0, []

    wb = load_workbook(io.BytesIO(file_bytes))
    registros = []
    aplicadas = 0

    def valor_conta_excel(conta):
        texto = texto_celula_seguro(conta)
        if texto.isdigit() and (texto == '0' or not texto.startswith('0')):
            return int(texto)
        return texto

    for _, item in revisoes.iterrows():
        conta_contrapartida = texto_celula_seguro(item.get('Conta da contrapartida'))
        if not conta_contrapartida:
            continue

        nome_aba = texto_celula_seguro(item.get('_aba'))
        if nome_aba not in wb.sheetnames:
            continue
        ws = wb[nome_aba]
        numero_linha = int(item.get('_linha'))
        col_destino = int(item.get('_col_destino'))
        col_banco = int(item.get('_col_banco'))
        banco = texto_celula_seguro(item.get('_banco'))
        conta_banco = texto_celula_seguro(contas_bancarias.get(banco, ''))

        if conta_banco and not texto_celula_seguro(ws.cell(numero_linha, col_banco).value):
            ws.cell(numero_linha, col_banco).value = valor_conta_excel(conta_banco)
        ws.cell(numero_linha, col_destino).value = valor_conta_excel(conta_contrapartida)
        aplicadas += 1

        # Lê o par final exatamente como ficou na planilha e aprende somente esta revisão.
        mapa_colunas = {}
        linha_cabecalho = None
        for linha_teste in range(1, min(ws.max_row, 30) + 1):
            mapa_teste = {
                normalizar_texto(texto_celula_seguro(ws.cell(linha_teste, coluna).value)).strip(): coluna
                for coluna in range(1, ws.max_column + 1)
            }
            if all(nome in mapa_teste for nome in ['historico', 'debito', 'credito']):
                mapa_colunas = mapa_teste
                linha_cabecalho = linha_teste
                break
        if linha_cabecalho is None:
            continue

        historico = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['historico']).value)
        debito = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['debito']).value)
        credito = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['credito']).value)
        assinatura = criar_assinatura_classificacao(historico)
        if not assinatura or not debito or not credito:
            continue

        periodo = normalizar_texto(filename)
        col_data = mapa_colunas.get('data')
        if col_data is not None:
            data_lancamento = pd.to_datetime(
                ws.cell(numero_linha, col_data).value, dayfirst=True, errors='coerce'
            )
            if not pd.isna(data_lancamento):
                periodo = data_lancamento.strftime('%Y-%m')

        identificador = hashlib.sha256(
            f"{empresa}|{banco}|{assinatura}|{debito}|{credito}".encode('utf-8')
        ).hexdigest()
        registros.append({
            'id': identificador,
            'empresa': empresa,
            'banco': banco,
            'assinatura': assinatura,
            'debito': debito,
            'credito': credito,
            'ocorrencias': 1,
            'periodos': [periodo],
            'exemplo_historico': historico[:500]
        })

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue(), aplicadas, registros


def renderizar_revisao_inteligente(
    arquivo_classificado,
    arquivo_original_bytes,
    filename,
    empresa,
    contas_bancarias,
    senha_admin,
    prefixo_chave,
):
    """Tela de revisão dos lançamentos que a Base Inteligente ainda não resolveu."""
    fingerprint = hashlib.sha256(
        arquivo_original_bytes + empresa.encode('utf-8')
    ).hexdigest()
    chave_fp = f'{prefixo_chave}_review_fp'
    chave_bytes = f'{prefixo_chave}_review_bytes'

    if st.session_state.get(chave_fp) != fingerprint:
        st.session_state[chave_fp] = fingerprint
        st.session_state[chave_bytes] = arquivo_classificado

    arquivo_trabalho = st.session_state.get(chave_bytes, arquivo_classificado)
    pendencias = extrair_pendencias_revisao_inteligente(
        arquivo_trabalho, contas_bancarias
    )

    st.markdown(
        """
        <div class="hc-review-box">
            <div class="hc-review-title">Revisão Inteligente</div>
            <div class="hc-review-text">
                O Razync mostra somente os lançamentos cuja contrapartida ainda não foi identificada.
                Preencha a conta correta, aplique a revisão e o novo padrão será aprendido apenas
                pela base desta empresa.
            </div>
        </div>
        <span class="hc-step-badge">1 · Base classifica</span>
        <span class="hc-step-badge">2 · Você revisa pendências</span>
        <span class="hc-step-badge">3 · Razync aprende</span>
        <span class="hc-step-badge">4 · Baixar planilha</span>
        """,
        unsafe_allow_html=True
    )

    if pendencias.empty:
        st.success("Nenhuma pendência de classificação. A planilha está pronta para download.")
    else:
        c_rev1, c_rev2 = st.columns(2)
        c_rev1.metric("Pendências para revisar", len(pendencias))
        c_rev2.metric(
            "Bancos envolvidos",
            int(pendencias['Banco'].nunique()) if 'Banco' in pendencias.columns else 0
        )
        st.caption(
            "Edite apenas a coluna Conta da contrapartida. As demais colunas servem "
            "como referência para você conferir o lançamento."
        )

        revisoes = st.data_editor(
            pendencias,
            use_container_width=True,
            hide_index=True,
            key=f'{prefixo_chave}_editor',
            disabled=[
                'Banco', 'Data', 'Valor', 'Histórico', 'Classificar em', 'Conta bancária'
            ],
            column_config={
                '_aba': None,
                '_linha': None,
                '_col_destino': None,
                '_col_banco': None,
                '_banco': None,
                '_col_data': None,
                'Conta da contrapartida': st.column_config.TextColumn(
                    'Conta da contrapartida',
                    help='Informe somente o número/código da conta contábil correta.',
                    width='medium'
                ),
                'Histórico': st.column_config.TextColumn('Histórico', width='large'),
                'Valor': st.column_config.NumberColumn('Valor', format='R$ %.2f'),
            },
            height=min(430, 92 + (len(pendencias) * 36))
        )

        senha_revisao = st.text_input(
            "Senha administrativa para salvar o aprendizado da revisão",
            type='password',
            key=f'{prefixo_chave}_senha_revisao'
        ) if senha_admin else ''

        preenchidas = int(
            revisoes['Conta da contrapartida'].fillna('').astype(str).str.strip().ne('').sum()
        )
        st.caption(f"{preenchidas} de {len(revisoes)} pendências preenchidas nesta revisão.")

        if st.button(
            "Aplicar revisões e ensinar a Base Inteligente",
            key=f'{prefixo_chave}_aplicar_revisao',
            use_container_width=True,
            disabled=preenchidas == 0
        ):
            if senha_admin and not hmac.compare_digest(str(senha_revisao), str(senha_admin)):
                st.error("Senha administrativa inválida.")
            else:
                try:
                    novo_arquivo, aplicadas, novos_padroes = executar_com_loading(
                        "Aplicando revisões e preparando o aprendizado...",
                        aplicar_revisoes_inteligentes,
                        arquivo_trabalho,
                        revisoes,
                        filename,
                        empresa,
                        contas_bancarias
                    )
                    if novos_padroes:
                        salvar_classificacoes_online(novos_padroes, empresa)
                    st.session_state[chave_bytes] = novo_arquivo
                    st.success(
                        f"{aplicadas} revisões aplicadas. "
                        f"{len(novos_padroes)} novos padrões foram enviados para a base desta empresa."
                    )
                    st.rerun()
                except Exception as erro_revisao:
                    st.error(f"Não foi possível aplicar a revisão: {erro_revisao}")

    nome_saida = os.path.splitext(filename)[0]
    st.download_button(
        "Baixar planilha classificada atual",
        data=arquivo_trabalho,
        file_name=f"{nome_saida}_Classificada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f'{prefixo_chave}_download_atual',
        use_container_width=True
    )

@st.cache_data(show_spinner=False, max_entries=12)
def processar_nova_geracao_banco(file_bytes, nome_aba, conta_esperada, descricao_banco):
    """Localiza uma conta na planilha consolidada e transforma seus lançamentos."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    conta_normalizada = re.sub(r'\D', '', conta_esperada)

    df, colunas = None, None
    for aba_candidata in xls.sheet_names:
        df_candidata = pd.read_excel(xls, sheet_name=aba_candidata, dtype=object)
        mapa = {normalizar_texto(str(col)).strip(): col for col in df_candidata.columns}
        obrigatorias = ['conta', 'data', 'valor', 'lacto', 'historico', 'doc']
        if not all(nome in mapa for nome in obrigatorias):
            continue
        contas_aba = df_candidata[mapa['conta']].apply(
            lambda valor: re.sub(r'\D', '', texto_celula_seguro(valor))
        )
        if contas_aba.eq(conta_normalizada).any():
            df = df_candidata
            colunas = mapa
            break

    if df is None or colunas is None:
        raise ValueError(
            f"A conta {conta_esperada} ({nome_aba}) não foi encontrada em nenhuma aba "
            "válida da planilha consolidada."
        )

    col_conta = colunas['conta']
    col_data = colunas['data']
    col_valor = colunas['valor']
    col_lacto = colunas['lacto']
    col_hist = colunas['historico']
    col_doc = colunas['doc']
    col_tipo = colunas.get('tipo')

    colunas_saida = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
    principais, retirados = [], []

    for _, linha in df.iterrows():
        conta = re.sub(r'\D', '', texto_celula_seguro(linha[col_conta]))
        if conta != conta_normalizada:
            continue

        data_raw = linha[col_data]
        if isinstance(data_raw, (int, float)) and not pd.isna(data_raw):
            data = pd.to_datetime(data_raw, unit='D', origin='1899-12-30', errors='coerce')
        else:
            data = pd.to_datetime(data_raw, dayfirst=True, errors='coerce')
        if pd.isna(data):
            continue

        lacto_original = texto_celula_seguro(linha[col_lacto])
        lacto_normalizado = normalizar_texto(lacto_original).strip()
        lacto = re.sub(r'\bPAGAR\b', 'PAGO', lacto_original, flags=re.IGNORECASE)
        lacto = re.sub(
            r'\b(?:RECEBER|RECEBIMENTO)\b', 'RECEBIDO', lacto, flags=re.IGNORECASE
        )

        valor_raw = linha[col_valor]
        valor_original = (
            float(valor_raw)
            if isinstance(valor_raw, (int, float)) and not pd.isna(valor_raw)
            else limpar_valor_monetario(valor_raw)
        )
        if valor_original == 0:
            continue

        tipo_normalizado = normalizar_texto(texto_celula_seguro(linha[col_tipo])) if col_tipo else ''
        if lacto_normalizado.startswith(('pagar', 'pago')):
            valor = -abs(valor_original)
        elif lacto_normalizado.startswith(('receber', 'recebido', 'recebimento')):
            valor = abs(valor_original)
        elif 'debito' in tipo_normalizado:
            valor = -abs(valor_original)
        elif 'credito' in tipo_normalizado:
            valor = abs(valor_original)
        else:
            valor = valor_original

        historico_valor_original = linha[col_hist]
        historico_origem_exato = (
            '' if historico_valor_original is None or pd.isna(historico_valor_original)
            else limpar_caracteres_ilegais(str(historico_valor_original))
        )
        historico_origem = texto_celula_seguro(historico_valor_original)
        documento = texto_celula_seguro(linha[col_doc])
        historico_final = re.sub(r'\s+', ' ', " ".join(
            parte for parte in [lacto, historico_origem, documento] if parte
        )).strip()

        registro = {
            'DESCRIÇÃO': descricao_banco,
            'DATA': data.to_pydatetime(),
            'VALOR': valor,
            'DÉBITO': '',
            'CRÉDITO': '',
            'HISTÓRICO': historico_final
        }

        if identificar_estorno_de_baixa(lacto_original, historico_origem, documento):
            registro_retirado = dict(registro)
            registro_retirado['HISTÓRICO'] = historico_origem_exato
            registro_retirado['MOTIVO'] = 'Estorno de baixa identificado'
            retirados.append(registro_retirado)
        else:
            principais.append(registro)

    if not principais and not retirados:
        raise ValueError(f"Nenhum lançamento da conta {nome_aba} {conta_esperada} foi encontrado.")

    return pd.DataFrame(principais, columns=colunas_saida), pd.DataFrame(
        retirados, columns=colunas_saida + ['MOTIVO']
    )

def processar_nova_geracao_itau(file_bytes):
    return processar_nova_geracao_banco(
        file_bytes, 'Itaú', '99549-5', 'BANCO ITAÚ'
    )

def processar_nova_geracao_bradesco(file_bytes):
    return processar_nova_geracao_banco(
        file_bytes, 'Bradesco', '451990-6', 'BANCO BRADESCO'
    )

def processar_nova_geracao_fibra(file_bytes):
    return processar_nova_geracao_banco(
        file_bytes, 'Fibra', '673947-1', 'BANCO FIBRA'
    )

def processar_nova_geracao_filial_itau(file_bytes):
    return processar_nova_geracao_banco(
        file_bytes, 'Itaú', '98002-6', 'BANCO ITAÚ'
    )

def processar_nova_geracao_filial_bradesco(file_bytes):
    return processar_nova_geracao_banco(
        file_bytes, 'Bradesco', '3084-8', 'BANCO BRADESCO'
    )

@st.cache_data(show_spinner=False, max_entries=12)
def processar_mapa_autokraft(file_bytes, filename=''):
    """Converte as abas diárias do mapa Autokraft para o Modelo Domínio."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    # Os mapas da Autokraft existem em dois padrões de nome de aba:
    # arquivos antigos usam DD.MM e arquivos mais novos usam DD-MM.
    # Aceitamos ambos sem incluir abas auxiliares de pagamentos/adiantamentos.
    abas_diarias = [
        aba for aba in xls.sheet_names
        if re.fullmatch(r'\d{2}[.-]\d{2}', str(aba).strip())
    ]
    if not abas_diarias:
        raise ValueError(
            "Nenhuma aba diária no formato DD-MM ou DD.MM foi encontrada no arquivo enviado."
        )

    ano_nome = re.search(r'(?<!\d)(20\d{2})(?!\d)', str(filename))
    ano_referencia = int(ano_nome.group(1)) if ano_nome else datetime.now().year
    colunas_saida = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
    registros = {'Itaú': [], 'Daycoval': []}
    abas_processadas = []

    for nome_aba in abas_diarias:
        df = pd.read_excel(xls, sheet_name=nome_aba, header=None, dtype=object)
        if df.empty or df.shape[1] < 6:
            continue

        data_raw = df.iloc[1, 2] if len(df.index) > 1 and df.shape[1] > 2 else None
        if isinstance(data_raw, (int, float)) and not pd.isna(data_raw):
            data_aba = pd.to_datetime(
                data_raw, unit='D', origin='1899-12-30', errors='coerce'
            )
        else:
            data_aba = pd.to_datetime(data_raw, dayfirst=True, errors='coerce')
        if pd.isna(data_aba):
            partes_data = re.split(r'[.-]', str(nome_aba).strip())
            if len(partes_data) != 2:
                continue
            dia, mes = [int(parte) for parte in partes_data]
            data_aba = pd.Timestamp(year=ano_referencia, month=mes, day=dia)

        banco_atual = None
        for _, linha in df.iterrows():
            nome_bloco = normalizar_texto(texto_celula_seguro(linha.iloc[0])).strip()
            if nome_bloco == 'itau':
                banco_atual = 'Itaú'
            elif nome_bloco == 'daycoval':
                banco_atual = 'Daycoval'

            historico_credito = texto_celula_seguro(linha.iloc[2])
            historico_debito = texto_celula_seguro(linha.iloc[4])
            texto_credito = normalizar_texto(historico_credito)
            texto_debito = normalizar_texto(historico_debito)

            if texto_credito.startswith('total de creditos') or texto_debito.startswith(
                'total de debitos'
            ):
                banco_atual = None
                continue
            if banco_atual is None:
                continue

            if historico_credito and not texto_credito.startswith('total'):
                valor_credito = abs(limpar_valor_monetario(linha.iloc[3]))
                if valor_credito:
                    historico_credito_final = limpar_caracteres_ilegais(
                        historico_credito
                    ).strip()
                    registros[banco_atual].append({
                        'DESCRIÇÃO': f'BANCO {banco_atual.upper()}',
                        'DATA': data_aba.to_pydatetime(),
                        'VALOR': valor_credito,
                        'DÉBITO': '',
                        'CRÉDITO': '',
                        'HISTÓRICO': f'Recebido: {historico_credito_final}'
                    })

            if historico_debito and not texto_debito.startswith('total'):
                valor_debito = abs(limpar_valor_monetario(linha.iloc[5]))
                if valor_debito:
                    historico_debito_final = limpar_caracteres_ilegais(
                        historico_debito
                    ).strip()
                    registros[banco_atual].append({
                        'DESCRIÇÃO': f'BANCO {banco_atual.upper()}',
                        'DATA': data_aba.to_pydatetime(),
                        'VALOR': -valor_debito,
                        'DÉBITO': '',
                        'CRÉDITO': '',
                        'HISTÓRICO': f'Pago: {historico_debito_final}'
                    })

        abas_processadas.append(str(nome_aba))

    dados_por_banco = {}
    for nome_banco, linhas in registros.items():
        df_banco = pd.DataFrame(linhas, columns=colunas_saida)
        if not df_banco.empty:
            df_banco = df_banco.sort_values('DATA', kind='stable').reset_index(drop=True)
        dados_por_banco[nome_banco] = {
            'principal': df_banco,
            'retirados': pd.DataFrame(columns=colunas_saida + ['MOTIVO'])
        }

    total_lancamentos = sum(
        len(dados['principal']) for dados in dados_por_banco.values()
    )
    if total_lancamentos == 0:
        raise ValueError(
            "As abas diárias foram encontradas, mas nenhum lançamento bancário válido foi lido."
        )
    return dados_por_banco, abas_processadas


@st.cache_data(show_spinner=False, max_entries=16)
def processar_planilha_accede_sig(file_bytes, banco_nome):
    """
    Converte planilhas SIG da ACCEDE para o Modelo Domínio.

    Regra estrutural: uma linha com DATA inicia o lançamento/grupo. Todas as linhas
    seguintes sem DATA pertencem a esse grupo até surgir uma nova DATA. Quando há
    detalhamento com valor individual, o total da linha principal não é duplicado.
    """
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    colunas_saida = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
    registros = []

    def texto_exato(valor):
        if valor is None or pd.isna(valor):
            return ''
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return limpar_caracteres_ilegais(str(valor)).strip()

    for nome_aba in xls.sheet_names:
        bruto = pd.read_excel(xls, sheet_name=nome_aba, header=None, dtype=object)
        if bruto.empty:
            continue

        idx_header = None
        nomes_header = None
        for idx in range(min(len(bruto), 30)):
            nomes = [normalizar_texto(texto_celula_seguro(v)).strip() for v in bruto.iloc[idx].tolist()]
            if all(nome in nomes for nome in ['data', 'complemento', 'entrada', 'saida']):
                idx_header = idx
                nomes_header = nomes
                break
        if idx_header is None:
            continue

        def coluna(nome):
            return nomes_header.index(nome) if nome in nomes_header else None

        c_data = coluna('data')
        c_dc = coluna('d/c')
        c_comp = coluna('complemento')
        c_conf = coluna('conf')
        c_ent = coluna('entrada')
        c_sai = coluna('saida')
        linhas = bruto.iloc[idx_header + 1:].reset_index(drop=True)
        i = 0

        while i < len(linhas):
            principal = linhas.iloc[i]
            data = pd.to_datetime(principal.iloc[c_data], dayfirst=True, errors='coerce')
            if pd.isna(data):
                i += 1
                continue

            j = i + 1
            detalhes = []
            while j < len(linhas):
                proxima_data = pd.to_datetime(linhas.iloc[j].iloc[c_data], dayfirst=True, errors='coerce')
                if not pd.isna(proxima_data):
                    break
                valores_linha = [texto_celula_seguro(v) for v in linhas.iloc[j].tolist()]
                if any(valores_linha):
                    detalhes.append(linhas.iloc[j])
                j += 1

            entrada = abs(limpar_valor_monetario(principal.iloc[c_ent])) if c_ent is not None else 0.0
            saida = abs(limpar_valor_monetario(principal.iloc[c_sai])) if c_sai is not None else 0.0
            sinal_grupo = 1 if entrada else (-1 if saida else 0)
            dc_principal = texto_exato(principal.iloc[c_dc]) if c_dc is not None else ''
            complemento = texto_exato(principal.iloc[c_comp]) if c_comp is not None else ''
            conf_principal = texto_exato(principal.iloc[c_conf]) if c_conf is not None else ''
            descricao_banco = 'BANCO ITAÚ' if normalizar_texto(banco_nome) == 'itau' else 'SICREDI'

            detalhes_validos = []
            for detalhe in detalhes:
                # Nos SIGs ACCEDE os detalhes aparecem deslocados para a esquerda:
                # [vazio/data, Conf/Documento, Valor, Favorecido/Descrição, ...].
                conf_doc = texto_exato(detalhe.iloc[1]) if len(detalhe) > 1 else ''
                valor_individual = abs(limpar_valor_monetario(detalhe.iloc[2])) if len(detalhe) > 2 else 0.0
                favorecido = texto_exato(detalhe.iloc[3]) if len(detalhe) > 3 else ''
                if valor_individual:
                    detalhes_validos.append((conf_doc, valor_individual, favorecido))

            if detalhes_validos:
                for conf_doc, valor_individual, favorecido in detalhes_validos:
                    historico = ' '.join(parte for parte in [favorecido, conf_doc] if parte).strip()
                    if not historico:
                        historico = complemento or conf_principal or dc_principal or 'MOVIMENTO BANCARIO'
                    registros.append({
                        'DESCRIÇÃO': descricao_banco,
                        'DATA': data.to_pydatetime(),
                        'VALOR': round(valor_individual * (sinal_grupo or -1), 2),
                        'DÉBITO': '',
                        'CRÉDITO': '',
                        'HISTÓRICO': historico
                    })
            else:
                valor = entrada if entrada else (-saida if saida else 0.0)
                if valor:
                    historico = complemento or conf_principal or dc_principal or 'MOVIMENTO BANCARIO'
                    registros.append({
                        'DESCRIÇÃO': descricao_banco,
                        'DATA': data.to_pydatetime(),
                        'VALOR': round(valor, 2),
                        'DÉBITO': '',
                        'CRÉDITO': '',
                        'HISTÓRICO': historico
                    })
            i = j

    df = pd.DataFrame(registros, columns=colunas_saida)
    if df.empty:
        raise ValueError(f'Nenhum lançamento válido foi encontrado na planilha SIG do {banco_nome}.')
    return df.sort_values('DATA', kind='stable').reset_index(drop=True)


def filtrar_dataframe_periodo(df, data_inicial, data_final):
    """Mantém somente os lançamentos entre as datas informadas, inclusive."""
    if df is None or df.empty:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    if 'DATA' not in df.columns:
        return df.iloc[0:0].copy()
    # Extratos brasileiros usam dia/mês/ano. Sem dayfirst=True, por exemplo,
    # 01/04/2026 seria interpretado como 4 de janeiro e sairia do filtro de abril.
    datas = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce').dt.date
    mascara = datas.between(data_inicial, data_final, inclusive='both')
    return df.loc[mascara].copy().reset_index(drop=True)

def identificar_chave_banco_empresa(valor):
    """Identifica os bancos conhecidos por descrição, aba, arquivo ou conta."""
    texto = normalizar_texto(texto_celula_seguro(valor))
    digitos = re.sub(r'\D', '', texto_celula_seguro(valor))
    if 'itau' in texto or any(conta in digitos for conta in ['995495', '980026']):
        return 'itau'
    if 'bradesco' in texto or any(conta in digitos for conta in ['4519906', '30848']):
        return 'bradesco'
    if 'fibra' in texto or '6739471' in digitos:
        return 'fibra'
    if 'daycoval' in texto:
        return 'daycoval'
    if 'sicredi' in texto:
        return 'sicredi'
    if 'banco do brasil' in texto:
        return 'banco_brasil'
    return ''

def nome_banco_por_chave(chave):
    return {
        'itau': 'Itaú', 'bradesco': 'Bradesco', 'fibra': 'Fibra',
        'daycoval': 'Daycoval', 'sicredi': 'Sicredi',
        'banco_brasil': 'Banco do Brasil'
    }.get(chave, chave)

def ler_planilha_organizada_conferencia(file_bytes, banco_alvo):
    """Lê a planilha final e retorna somente o banco escolhido para conferência."""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    colunas_base = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
    principais, retirados, bancos_encontrados = [], [], set()

    for nome_aba in xls.sheet_names:
        df_bruto = pd.read_excel(xls, sheet_name=nome_aba, header=None, dtype=object)
        if df_bruto.empty:
            continue

        indice_cabecalho = None
        for indice in range(min(len(df_bruto), 30)):
            nomes_linha = [
                normalizar_texto(texto_celula_seguro(valor)).strip()
                for valor in df_bruto.iloc[indice].tolist()
            ]
            if ('data' in nomes_linha and 'valor' in nomes_linha and
                    any(nome in nomes_linha for nome in ['historico', 'histórico'])):
                indice_cabecalho = indice
                break
        if indice_cabecalho is None:
            continue

        cabecalhos = [texto_celula_seguro(valor) for valor in df_bruto.iloc[indice_cabecalho]]
        df_aba = df_bruto.iloc[indice_cabecalho + 1:].copy()
        df_aba.columns = cabecalhos
        mapa = {normalizar_texto(str(coluna)).strip(): coluna for coluna in df_aba.columns}
        col_data = mapa.get('data')
        col_valor = mapa.get('valor')
        col_hist = mapa.get('historico')
        col_desc = mapa.get('descricao')
        col_motivo = mapa.get('motivo')
        if col_data is None or col_valor is None or col_hist is None:
            continue

        banco_aba = identificar_chave_banco_empresa(nome_aba)
        aba_retirados = 'retir' in normalizar_texto(nome_aba)
        for _, linha in df_aba.iterrows():
            banco_linha = (
                identificar_chave_banco_empresa(linha[col_desc]) if col_desc is not None else ''
            ) or banco_aba
            if banco_linha:
                bancos_encontrados.add(banco_linha)
            if banco_linha != banco_alvo:
                continue

            data_raw = linha[col_data]
            if isinstance(data_raw, (int, float)) and not pd.isna(data_raw):
                data = pd.to_datetime(data_raw, unit='D', origin='1899-12-30', errors='coerce')
            else:
                data = pd.to_datetime(data_raw, dayfirst=True, errors='coerce')
            valor = limpar_valor_monetario(linha[col_valor])
            if pd.isna(data) or valor == 0:
                continue

            descricao = texto_celula_seguro(linha[col_desc]) if col_desc is not None else ''
            if not descricao:
                descricao = {
                    'itau': 'BANCO ITAÚ', 'bradesco': 'BANCO BRADESCO',
                    'fibra': 'BANCO FIBRA', 'daycoval': 'BANCO DAYCOVAL',
                    'sicredi': 'SICREDI', 'banco_brasil': 'BANCO DO BRASIL'
                }[banco_alvo]
            historico_valor = linha[col_hist]
            historico = (
                '' if historico_valor is None or pd.isna(historico_valor)
                else limpar_caracteres_ilegais(str(historico_valor))
            )
            registro = {
                'DESCRIÇÃO': descricao,
                'DATA': data.to_pydatetime(),
                'VALOR': valor,
                'DÉBITO': '',
                'CRÉDITO': '',
                'HISTÓRICO': historico
            }
            if aba_retirados:
                registro['MOTIVO'] = (
                    texto_celula_seguro(linha[col_motivo]) if col_motivo is not None
                    else 'Estorno de baixa identificado'
                )
                retirados.append(registro)
            else:
                principais.append(registro)

    return (
        pd.DataFrame(principais, columns=colunas_base),
        pd.DataFrame(retirados, columns=colunas_base + ['MOTIVO']),
        [nome_banco_por_chave(chave) for chave in sorted(bancos_encontrados)]
    )

def gerar_excel_nova_geracao(dados_por_banco, modelo_bytes=None):
    """Gera um único arquivo com uma aba do Modelo Domínio para cada banco."""
    from openpyxl import Workbook, load_workbook

    if modelo_bytes:
        wb = load_workbook(io.BytesIO(modelo_bytes))
        ws_modelo = wb[wb.sheetnames[0]]
        if ws_modelo.max_row > 1:
            ws_modelo.delete_rows(2, ws_modelo.max_row - 1)
    else:
        wb = Workbook()
        ws_modelo = wb.active
        ws_modelo.title = 'Modelo temporário'
        ws_modelo.append(['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO'])

    cabecalhos = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
    for col, cabecalho in enumerate(cabecalhos, 1):
        ws_modelo.cell(1, col, cabecalho)

    def preparar_linha_modelo(registro, colunas):
        linha = []
        for coluna in colunas:
            valor = registro.get(coluna, '')
            if coluna == 'DATA':
                data = pd.to_datetime(valor, errors='coerce')
                valor = data.strftime('%d/%m/%Y') if not pd.isna(data) else ''
            elif pd.isna(valor):
                valor = ''
            linha.append(valor)
        return linha

    nomes_criados = []
    retirados_gerais = []
    for nome_banco, dados_banco in dados_por_banco.items():
        nome_aba = str(nome_banco)[:31]
        if nome_aba in nomes_criados:
            sufixo = 2
            while f"{nome_aba[:28]} {sufixo}" in nomes_criados:
                sufixo += 1
            nome_aba = f"{nome_aba[:28]} {sufixo}"

        ws_banco = wb.copy_worksheet(ws_modelo)
        ws_banco.title = nome_aba
        nomes_criados.append(nome_aba)

        df_principal = dados_banco.get('principal', pd.DataFrame())
        df_retirados = dados_banco.get('retirados', pd.DataFrame())
        for registro in df_principal.to_dict('records'):
            ws_banco.append(preparar_linha_modelo(registro, cabecalhos))
        if not df_retirados.empty:
            retirados_gerais.extend(df_retirados.to_dict('records'))

    wb.remove(ws_modelo)

    if retirados_gerais:
        nome_retirados = 'Lançamentos retirados'
        if nome_retirados in wb.sheetnames:
            del wb[nome_retirados]
        ws_ret = wb.create_sheet(nome_retirados)
        cabecalhos_ret = cabecalhos + ['MOTIVO']
        ws_ret.append(cabecalhos_ret)
        for registro in retirados_gerais:
            ws_ret.append(preparar_linha_modelo(registro, cabecalhos_ret))

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()

def processar_pdf_bradesco_mensal(reader, banco='BANCO BRADESCO'):
    """Lê extratos mensais Bradesco, inclusive PDFs rasterizados via OCR."""
    lancamentos = []
    data_atual = None
    partes_historico = []
    ultimo_saldo = None
    dentro_saldos_invest = False
    modo_ocr = False
    saldo_abertura = None
    indice_saldo_abertura = 0
    erro_ocr = ''

    regex_data = re.compile(r'^(\d{2}/\d{2}/\d{4})\s*[|—-]?\s*(.*)$')
    regex_moeda = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2}')
    ignorar_prefixos = (
        'extrato de:', 'agência | conta', 'agencia | conta', 'data lançamento',
        'data lancamento', 'folha ', 'extrato mensal / por período',
        'extrato mensal / por periodo', 'nome do usuário:', 'nome do usuario:',
        'data da operação:', 'data da operacao:', 'os dados acima têm como base',
        'os dados acima tem como base',
    )

    textos_paginas = [pagina.extract_text() or '' for pagina in reader.pages]

    # PDF-imagem: OCR somente quando não existe qualquer camada de texto.
    if not any(texto.strip() for texto in textos_paginas):
        modo_ocr = True
        try:
            import fitz
            import pytesseract
            from PIL import Image, ImageOps

            caminho_pdf = (
                getattr(reader, '_razync_source_path', None)
                or getattr(getattr(reader, 'stream', None), 'name', None)
            )
            if not caminho_pdf or not os.path.exists(caminho_pdf):
                erro_ocr = 'Arquivo temporário do PDF não ficou disponível para o OCR.'
                reader._razync_ocr_error = erro_ocr
            if caminho_pdf and os.path.exists(caminho_pdf):
                documento_ocr = fitz.open(caminho_pdf)
                textos_paginas = []
                for pagina_ocr in documento_ocr:
                    pix = pagina_ocr.get_pixmap(
                        matrix=fitz.Matrix(4.0, 4.0), alpha=False
                    )
                    imagem = Image.frombytes(
                        'RGB', [pix.width, pix.height], pix.samples
                    )
                    imagem = ImageOps.autocontrast(ImageOps.grayscale(imagem))
                    texto_ocr = pytesseract.image_to_string(
                        imagem,
                        lang='por',
                        config='--psm 6 -c preserve_interword_spaces=1'
                    )
                    textos_paginas.append(texto_ocr or '')
                documento_ocr.close()
        except Exception as erro:
            erro_ocr = str(erro)
            reader._razync_ocr_error = erro_ocr
            textos_paginas = textos_paginas or []

    reader._razync_ocr_executado = modo_ocr
    for texto in textos_paginas:
        for linha_bruta in texto.splitlines():
            linha = re.sub(r'\s+', ' ', linha_bruta).strip()
            if not linha:
                continue

            normalizada = normalizar_texto(linha)

            if normalizada.startswith('saldos invest facil'):
                dentro_saldos_invest = True
                partes_historico = []
                continue
            if normalizada.startswith('ultimos lancamentos'):
                dentro_saldos_invest = False
                partes_historico = []
                ultimo_saldo = None
                continue
            if normalizada.startswith(('data lancamento', 'data lançamento')):
                dentro_saldos_invest = False
                partes_historico = []
                continue
            if dentro_saldos_invest:
                continue
            if normalizada.startswith(ignorar_prefixos):
                continue
            if normalizada.startswith('nova geracao comercial') and 'cnpj:' in normalizada:
                continue
            if normalizada.startswith('total '):
                partes_historico = []
                continue

            match_data = regex_data.match(linha)
            if match_data:
                data_atual = match_data.group(1)
                linha = match_data.group(2).strip()
                normalizada = normalizar_texto(linha)
                if not linha:
                    continue

            if 'saldo anterior' in normalizada:
                moedas_saldo = regex_moeda.findall(linha)
                if moedas_saldo:
                    ultimo_saldo = limpar_valor_monetario(moedas_saldo[-1])
                    saldo_abertura = ultimo_saldo
                    indice_saldo_abertura = len(lancamentos)
                partes_historico = []
                continue

            if not data_atual:
                continue

            moedas = regex_moeda.findall(linha)
            if len(moedas) >= 2:
                valor_txt = moedas[-2]
                saldo_txt = moedas[-1]
                valor_impresso = limpar_valor_monetario(valor_txt)
                saldo_lido = limpar_valor_monetario(saldo_txt)
                valor = valor_impresso

                if ultimo_saldo is not None:
                    variacao = round(saldo_lido - ultimo_saldo, 2)
                    if modo_ocr:
                        # OCR pode perder o sinal do débito ou errar um dígito do saldo.
                        # A direção do saldo define o sinal; a magnitude impressa continua
                        # sendo usada quando a leitura do saldo não fecha exatamente.
                        sinal = -1 if variacao < 0 else 1
                        if abs(abs(variacao) - abs(valor_impresso)) <= max(
                            0.05, abs(valor_impresso) * 0.01
                        ):
                            valor = variacao
                            ultimo_saldo = saldo_lido
                        else:
                            valor = sinal * abs(valor_impresso)
                            ultimo_saldo = round(ultimo_saldo + valor, 2)
                    else:
                        if abs(abs(variacao) - abs(valor_impresso)) <= 0.02:
                            valor = variacao
                        ultimo_saldo = saldo_lido
                else:
                    ultimo_saldo = saldo_lido

                inicio_valor = linha.rfind(valor_txt)
                trecho_historico = linha[:inicio_valor].strip()
                historico = re.sub(
                    r'\s+', ' ',
                    ' '.join(
                        partes_historico
                        + ([trecho_historico] if trecho_historico else [])
                    )
                ).strip(' |—-')
                partes_historico = []

                hist_norm = normalizar_texto(historico)
                if not historico or hist_norm.startswith(('saldo ', 'total ')):
                    continue
                if abs(valor) < 0.005:
                    continue

                try:
                    data = datetime.strptime(data_atual, '%d/%m/%Y')
                except ValueError:
                    continue

                lancamentos.append({
                    'DESCRIÇÃO': banco,
                    'DATA': data,
                    'VALOR': round(valor, 2),
                    'DÉBITO': '',
                    'CRÉDITO': '',
                    'HISTÓRICO': historico,
                })
            else:
                partes_historico.append(linha)
                if len(partes_historico) > 8:
                    partes_historico = partes_historico[-8:]

    if saldo_abertura is not None and ultimo_saldo is not None:
        movimentos_validacao = [
            item.get('VALOR', 0.0) for item in lancamentos[indice_saldo_abertura:]
        ]
        reader._razync_balance_check = validar_fechamento_saldo(
            saldo_abertura, ultimo_saldo, movimentos_validacao
        )
    reader._razync_ocr_executado = modo_ocr
    if erro_ocr:
        reader._razync_ocr_error = erro_ocr
    return lancamentos

def processar_extrato_conferencia_empresa(file_bytes, filename):
    """Lê a conferência pelo mesmo motor central usado em todo o Razync."""
    termos_saldo = [
        'saldo anterior', 'saldo aplic', 'saldo invest', 'saldo total disponivel',
        'saldo movimentacao conta', 'sdo aplic aut mais ap', 'saldo final',
        'saldo do dia', 'saldo total', 'saldo disponivel', 'saldo em conta',
    ]
    filtrados = []
    for item in processar_extrato_unificado(file_bytes, filename) or []:
        historico = normalizar_texto(texto_celula_seguro(item.get('HISTÓRICO', '')))
        if any(termo in historico for termo in termos_saldo):
            continue
        valor = limpar_valor_monetario(item.get('VALOR', 0))
        if abs(valor) < 0.005:
            continue
        filtrados.append(item)
    fechamento = st.session_state.get('ultimo_fechamento_extrato')
    if fechamento and fechamento.get('disponivel') and fechamento.get('ok') is False:
        st.warning(
            'O extrato foi lido, mas o fechamento matemático do saldo apresentou '
            f"diferença de {formatar_moeda(abs(fechamento.get('diferenca', 0)))}. "
            'Revise os lançamentos antes de concluir a conciliação.'
        )
    if not filtrados:
        erro_leitura = st.session_state.get('ultimo_erro_extrato', '')
        if erro_leitura:
            raise ValueError(erro_leitura)
    return filtrados

def conciliar_empresa_com_extrato(df_planilha, lancamentos_extrato, df_retirados=None):
    """Compara movimentos por dia e faz pareamento individual por data e centavos."""
    colunas_base = ['DESCRIÇÃO', 'DATA', 'VALOR', 'HISTÓRICO']

    def preparar_dataframe(dados):
        if isinstance(dados, pd.DataFrame):
            df = dados.copy()
        else:
            df = pd.DataFrame(dados or [])
        for coluna in colunas_base:
            if coluna not in df.columns:
                df[coluna] = '' if coluna != 'VALOR' else 0.0
        df['DESCRIÇÃO'] = df['DESCRIÇÃO'].fillna('').astype(str)
        df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce').dt.normalize()
        df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce').fillna(0.0).round(2)
        df['HISTÓRICO'] = df['HISTÓRICO'].fillna('').astype(str)
        df = df.dropna(subset=['DATA'])
        df = df[df['VALOR'].abs() >= 0.005].copy()
        df['_CENTAVOS'] = (df['VALOR'] * 100).round().astype(int)
        df['_BANCO'] = df['DESCRIÇÃO'].apply(
            lambda valor: re.sub(r'\s+', ' ', normalizar_texto(valor).replace('banco', '')).strip()
        )
        return df.reset_index(drop=True)

    df_modelo = preparar_dataframe(df_planilha)
    df_extrato = preparar_dataframe(lancamentos_extrato)
    df_retirados_ok = preparar_dataframe(df_retirados if df_retirados is not None else [])

    usar_banco_na_chave = df_modelo.loc[df_modelo['_BANCO'] != '', '_BANCO'].nunique() > 1
    for dataframe in [df_modelo, df_extrato, df_retirados_ok]:
        if usar_banco_na_chave:
            dataframe['_CHAVE'] = list(zip(
                dataframe['_BANCO'], dataframe['DATA'], dataframe['_CENTAVOS']
            ))
        else:
            dataframe['_CHAVE'] = list(zip(dataframe['DATA'], dataframe['_CENTAVOS']))

    # Estornos de baixa retirados de propósito não devem gerar falso alerta.
    indices_ignorados = set()
    if not df_retirados_ok.empty and not df_extrato.empty:
        quantidades_retiradas = df_retirados_ok['_CHAVE'].value_counts().to_dict()
        for chave, quantidade in quantidades_retiradas.items():
            candidatos = df_extrato.index[
                (df_extrato['_CHAVE'] == chave) &
                df_extrato['HISTÓRICO'].apply(identificar_estorno_de_baixa)
            ].tolist()
            indices_ignorados.update(candidatos[:int(quantidade)])

    df_ignorados = df_extrato.loc[sorted(indices_ignorados)].copy() if indices_ignorados else df_extrato.iloc[0:0].copy()
    df_extrato_comparavel = df_extrato.drop(index=list(indices_ignorados)).reset_index(drop=True)

    # Pareamento um a um: lançamentos repetidos são tratados individualmente.
    disponiveis_modelo = {}
    for indice, chave in enumerate(df_modelo['_CHAVE']):
        disponiveis_modelo.setdefault(chave, []).append(indice)

    indices_modelo_pareados = set()
    indices_extrato_sem_par = []
    for indice_extrato, chave in enumerate(df_extrato_comparavel['_CHAVE']):
        candidatos = disponiveis_modelo.get(chave, [])
        if candidatos:
            indices_modelo_pareados.add(candidatos.pop(0))
        else:
            indices_extrato_sem_par.append(indice_extrato)

    indices_modelo_sem_par = [
        indice for indice in range(len(df_modelo))
        if indice not in indices_modelo_pareados
    ]

    faltando_planilha = df_extrato_comparavel.loc[indices_extrato_sem_par, colunas_base].copy()
    a_mais_planilha = df_modelo.loc[indices_modelo_sem_par, colunas_base].copy()
    ignorados = df_ignorados[colunas_base].copy()

    def resumo_diario_por_natureza(df, prefixo):
        temp = df[['DATA', 'VALOR']].copy()
        temp[f'ENTRADAS {prefixo}'] = temp['VALOR'].where(temp['VALOR'] > 0, 0.0)
        temp[f'SAÍDAS {prefixo}'] = -temp['VALOR'].where(temp['VALOR'] < 0, 0.0)
        return temp.groupby('DATA', as_index=False)[[f'ENTRADAS {prefixo}', f'SAÍDAS {prefixo}']].sum()

    ext_dia = resumo_diario_por_natureza(df_extrato_comparavel, 'EXTRATO')
    plan_dia = resumo_diario_por_natureza(df_modelo, 'PLANILHA')
    diario = pd.merge(ext_dia, plan_dia, on='DATA', how='outer').fillna(0.0).sort_values('DATA')
    diario['DIF. ENTRADAS'] = (diario['ENTRADAS PLANILHA'] - diario['ENTRADAS EXTRATO']).round(2)
    diario['DIF. SAÍDAS'] = (diario['SAÍDAS PLANILHA'] - diario['SAÍDAS EXTRATO']).round(2)
    diario['STATUS ENTRADAS'] = diario['DIF. ENTRADAS'].apply(lambda v: '✅ Batendo' if abs(v) < 0.01 else '❌ Divergente')
    diario['STATUS SAÍDAS'] = diario['DIF. SAÍDAS'].apply(lambda v: '✅ Batendo' if abs(v) < 0.01 else '❌ Divergente')
    diario['STATUS'] = diario.apply(lambda r: '✅ Batendo' if abs(r['DIF. ENTRADAS']) < 0.01 and abs(r['DIF. SAÍDAS']) < 0.01 else '❌ Divergente', axis=1)

    return diario, faltando_planilha, a_mais_planilha, ignorados

def renderizar_conferencia_autokraft(prefixo_chaves='autokraft', bancos_config=None):
    """Exibe a conferência independente da planilha final do Grupo Autokraft."""
    st.markdown("---")
    st.markdown("### Conferência com o extrato bancário")
    configs = bancos_config or [
        {'nome': 'Itaú', 'slug': 'itau'},
        {'nome': 'Daycoval', 'slug': 'daycoval'}
    ]
    nomes_bancos = [config['nome'] for config in configs]
    st.caption(
        "Envie a planilha final organizada e os extratos correspondentes. "
        "Cada banco terá seu próprio relatório diário."
    )
    conferir_todos = st.checkbox(
        "Conferir os dois bancos",
        value=False,
        key=f"{prefixo_chaves}_conferir_todos"
    )
    if conferir_todos:
        bancos_escolhidos = nomes_bancos
        st.caption("Serão apresentados relatórios separados para os bancos selecionados.")
    else:
        bancos_escolhidos = st.multiselect(
            "Bancos que serão conferidos",
            nomes_bancos,
            default=[nomes_bancos[0]],
            key=f"{prefixo_chaves}_bancos_conferencia"
        )

    if not bancos_escolhidos:
        st.info("Selecione pelo menos um banco para realizar a conferência.")
        return
    configs_escolhidas = [
        config for config in configs if config['nome'] in bancos_escolhidos
    ]

    coluna_planilha, coluna_extratos = st.columns(2)
    with coluna_planilha:
        planilha_final = st.file_uploader(
            "Planilha final organizada",
            type=['xlsx', 'xls'],
            key=f"{prefixo_chaves}_planilha_final_conferencia",
            help="Pode ser o arquivo baixado pelo organizador com uma ou duas abas bancárias."
        )
    with coluna_extratos:
        extratos = st.file_uploader(
            "Extrato(s) bancário(s)",
            type=['pdf', 'ofx', 'csv', 'xlsx', 'xls'],
            accept_multiple_files=True,
            key=(
                f"{prefixo_chaves}_extratos_conferencia_"
                + "_".join(config['slug'] for config in configs_escolhidas)
            ),
            help="Envie os extratos correspondentes ao mesmo período da planilha final."
        )

    if not planilha_final:
        st.info(
            "Envie a planilha final organizada para identificar o período e liberar a comparação."
        )
        return

    try:
        dados_planilha = {}
        bancos_detectados = set()
        datas_planilha = []
        for config in configs_escolhidas:
            df_modelo, df_retirados, bancos_arquivo = ler_planilha_organizada_conferencia(
                planilha_final.getvalue(), config['slug']
            )
            dados_planilha[config['slug']] = {
                'modelo': df_modelo,
                'retirados': df_retirados
            }
            bancos_detectados.update(bancos_arquivo)
            if not df_modelo.empty:
                datas_validas = pd.to_datetime(
                    df_modelo['DATA'], dayfirst=True, errors='coerce'
                ).dropna()
                datas_planilha.extend(datas_validas.dt.date.tolist())

        if not datas_planilha:
            st.warning("A planilha final não possui datas válidas nos bancos selecionados.")
            return

        data_minima = min(datas_planilha)
        data_maxima = max(datas_planilha)
        periodo = st.date_input(
            "Período da conferência",
            value=(data_minima, data_maxima),
            min_value=data_minima,
            max_value=data_maxima,
            format="DD/MM/YYYY",
            key=f"{prefixo_chaves}_periodo_conferencia"
        )
        if not isinstance(periodo, (tuple, list)) or len(periodo) != 2:
            st.info("Selecione também a data final para concluir o período.")
            return
        data_inicial, data_final = periodo

        dados_filtrados = {}
        for config in configs_escolhidas:
            chave = config['slug']
            dados_filtrados[chave] = {
                'modelo': filtrar_dataframe_periodo(
                    dados_planilha[chave]['modelo'], data_inicial, data_final
                ),
                'retirados': filtrar_dataframe_periodo(
                    dados_planilha[chave]['retirados'], data_inicial, data_final
                )
            }

        bancos_texto = ", ".join(sorted(bancos_detectados)) or "não identificados"
        st.success(
            f"Planilha carregada. Bancos identificados: {bancos_texto}. "
            f"Período: {data_inicial.strftime('%d/%m/%Y')} até "
            f"{data_final.strftime('%d/%m/%Y')}."
        )
        if not extratos:
            st.info("Agora envie pelo menos um extrato para gerar os relatórios.")
            return

        extratos_por_banco = {
            config['slug']: [] for config in configs_escolhidas
        }
        arquivos_nao_identificados = []
        for arquivo_extrato in extratos:
            lancamentos = executar_com_loading(
                f"Lendo {arquivo_extrato.name}...",
                processar_extrato_conferencia_empresa,
                arquivo_extrato.getvalue(),
                arquivo_extrato.name
            )
            df_extrato = filtrar_dataframe_periodo(
                pd.DataFrame(lancamentos), data_inicial, data_final
            )
            if df_extrato.empty:
                continue

            chave_nome = identificar_chave_banco_empresa(arquivo_extrato.name)
            if chave_nome in extratos_por_banco:
                extratos_por_banco[chave_nome].extend(df_extrato.to_dict('records'))
                continue

            chaves_linhas = df_extrato['DESCRIÇÃO'].apply(identificar_chave_banco_empresa)
            chaves_reconhecidas = {
                chave for chave in chaves_linhas.unique().tolist()
                if chave in extratos_por_banco
            }
            if not chaves_reconhecidas:
                if len(configs_escolhidas) == 1:
                    chave_unica = configs_escolhidas[0]['slug']
                    extratos_por_banco[chave_unica].extend(df_extrato.to_dict('records'))
                else:
                    arquivos_nao_identificados.append(arquivo_extrato.name)
                continue

            for chave in chaves_reconhecidas:
                df_banco = df_extrato[chaves_linhas.eq(chave)]
                extratos_por_banco[chave].extend(df_banco.to_dict('records'))

        if arquivos_nao_identificados:
            st.warning(
                "Não foi possível identificar o banco destes arquivos: "
                + ", ".join(arquivos_nao_identificados)
            )
        if not any(extratos_por_banco.values()):
            st.warning(
                "Nenhum lançamento dos extratos foi identificado dentro do período selecionado."
            )
            return

        abas_relatorio = st.tabs([config['nome'] for config in configs_escolhidas])
        for aba_relatorio, config in zip(abas_relatorio, configs_escolhidas):
            with aba_relatorio:
                chave = config['slug']
                nome_banco = config['nome']
                df_modelo = dados_filtrados[chave]['modelo']
                df_extrato = pd.DataFrame(extratos_por_banco[chave])
                st.markdown(f"#### Relatório — {nome_banco}")
                if df_modelo.empty:
                    st.warning(f"Não há lançamentos do {nome_banco} na planilha para o período.")
                    continue
                if df_extrato.empty:
                    st.warning(f"Nenhum extrato do {nome_banco} foi identificado para o período.")
                    continue

                diario, _, _, _ = executar_com_loading(
                    f"Conferindo os movimentos do {nome_banco}...",
                    conciliar_empresa_com_extrato,
                    df_modelo,
                    df_extrato,
                    dados_filtrados[chave]['retirados']
                )
                if diario.empty:
                    st.warning("Não existem datas válidas para realizar a conferência.")
                    continue

                dias_batendo = int((diario['STATUS'] == '✅ Batendo').sum())
                dias_divergentes = int((diario['STATUS'] == '❌ Divergente').sum())
                te = float(diario['ENTRADAS EXTRATO'].sum())
                tp = float(diario['ENTRADAS PLANILHA'].sum())
                se = float(diario['SAÍDAS EXTRATO'].sum())
                sp = float(diario['SAÍDAS PLANILHA'].sum())
                dif_ent = round(tp - te, 2)
                dif_sai = round(sp - se, 2)

                st.markdown("##### Conferência diária")
                resumo1, resumo2 = st.columns(2)
                resumo1.metric("Dias batendo", dias_batendo)
                resumo2.metric("Dias divergentes", dias_divergentes)

                exibicao = diario.copy()
                exibicao['DATA'] = exibicao['DATA'].dt.strftime('%d/%m/%Y')
                exibicao = exibicao[[
                    'DATA',
                    'ENTRADAS PLANILHA', 'ENTRADAS EXTRATO',
                    'SAÍDAS PLANILHA', 'SAÍDAS EXTRATO',
                    'STATUS'
                ]]
                exibicao.columns = [
                    'Data',
                    'Entrada Planilha', 'Entrada Extrato',
                    'Saída Planilha', 'Saída Extrato',
                    'Status'
                ]
                exibicao = formatar_dataframe_moeda_br(
                    exibicao,
                    ['Entrada Planilha', 'Entrada Extrato',
                     'Saída Planilha', 'Saída Extrato']
                )
                st.dataframe(exibicao, use_container_width=True, height=390, hide_index=True)

                if dias_divergentes == 0:
                    st.success("✅ Entradas e saídas estão batendo em todos os dias.")
                else:
                    st.warning("❌ Existem dias com divergência entre a planilha e o extrato.")
    except Exception as erro:
        st.error(f"Não foi possível realizar a conferência: {erro}")


st.markdown("""
<style>
/* Ajustes estruturais baseados na inspeção visual da Home publicada. */
section[data-testid="stSidebar"] [data-testid="stRadio"] {
    position: fixed !important;
    left: 0.9rem;
    bottom: 3.3rem;
    width: 238px;
    z-index: 5;
    padding: 0.65rem 0.55rem 0.2rem !important;
    margin: 0 !important;
    background: color-mix(in srgb, var(--rz-panel) 96%, transparent);
}
section[data-testid="stSidebar"] .stButton > button {
    margin-bottom: 0.08rem !important;
}
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0.28rem !important;
}

.rz-dashboard-intro {
    max-width: 780px;
    padding: 0.2rem 0 1.1rem;
}
.rz-dashboard-intro .rz-home-title {
    font-size: clamp(2rem, 4vw, 3rem);
}
.rz-dashboard-grid-title {
    color: var(--rz-muted);
    font-size: 0.67rem;
    font-weight: 730;
    letter-spacing: 0.105em;
    text-transform: uppercase;
    margin: 0.65rem 0 0.55rem;
}
.st-key-home_action_organizador button,
.st-key-home_action_extratos button,
.st-key-home_action_razao button {
    height: auto !important;
    min-height: 82px !important;
    max-height: none !important;
    padding: 0.9rem 1rem !important;
    margin: 0 0 0.48rem !important;
    justify-content: flex-start !important;
    align-items: center !important;
    text-align: left !important;
    white-space: pre-line !important;
    border: 1px solid var(--rz-line) !important;
    border-radius: 10px !important;
    background: transparent !important;
    box-shadow: none !important;
    color: var(--rz-muted) !important;
    font-size: 0.76rem !important;
    line-height: 1.42 !important;
}
.st-key-home_action_organizador button {
    min-height: 104px !important;
    background: var(--rz-panel) !important;
    border-color: color-mix(in srgb, var(--rz-accent) 34%, var(--rz-line)) !important;
}
.st-key-home_action_organizador button:hover,
.st-key-home_action_extratos button:hover,
.st-key-home_action_razao button:hover {
    background: var(--rz-accent-soft) !important;
    border-color: var(--rz-accent) !important;
    transform: translateX(3px) !important;
}
.st-key-home_action_organizador button p,
.st-key-home_action_extratos button p,
.st-key-home_action_razao button p {
    margin: 0 !important;
    white-space: pre-line !important;
}
.st-key-home_action_organizador button strong,
.st-key-home_action_extratos button strong,
.st-key-home_action_razao button strong {
    color: var(--rz-text) !important;
    font-size: 0.98rem !important;
    font-weight: 650 !important;
}

.rz-overview-panel {
    min-height: 286px;
    padding: 1.15rem 1.2rem;
    border: 1px solid var(--rz-line);
    border-radius: 12px;
    background: color-mix(in srgb, var(--rz-panel) 78%, transparent);
}
.rz-overview-kicker {
    color: var(--rz-muted);
    font-size: 0.66rem;
    font-weight: 730;
    letter-spacing: .105em;
    text-transform: uppercase;
    margin-bottom: 0.7rem;
}
.rz-overview-title {
    color: var(--rz-text);
    font-size: 1.08rem;
    font-weight: 650;
    letter-spacing: -0.02em;
    margin-bottom: 0.35rem;
}
.rz-overview-copy {
    color: var(--rz-muted);
    font-size: 0.8rem;
    line-height: 1.55;
    margin-bottom: 1rem;
}
.rz-overview-row {
    display: grid;
    grid-template-columns: 8px 1fr;
    gap: 0.65rem;
    align-items: start;
    padding: 0.68rem 0;
    border-top: 1px solid var(--rz-line);
}
.rz-overview-dot {
    width: 7px;
    height: 7px;
    margin-top: 0.34rem;
    border-radius: 50%;
    background: var(--rz-accent);
}
.rz-overview-row strong {
    display: block;
    color: var(--rz-text);
    font-size: 0.78rem;
    font-weight: 610;
}
.rz-overview-row span {
    display: block;
    color: var(--rz-muted);
    font-size: 0.71rem;
    line-height: 1.45;
    margin-top: 0.12rem;
}
@media (max-width: 900px) {
    section[data-testid="stSidebar"] [data-testid="stRadio"] {
        position: static !important;
        width: auto;
        margin-top: 0.8rem !important;
    }
    .rz-overview-panel { min-height: auto; }
}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
/* Alinhamento fino da Home após inspeção em 1352x615. */
.rz-dashboard-intro {
    max-width: 760px !important;
    padding-bottom: 1rem !important;
}
.rz-dashboard-intro .rz-home-title {
    font-size: clamp(1.9rem, 3.3vw, 2.65rem) !important;
    line-height: 1.06 !important;
}
.rz-dashboard-grid-title {
    height: 1.35rem;
    display: flex;
    align-items: center;
    margin: 0.55rem 0 0.45rem !important;
}

.st-key-home_action_organizador button,
.st-key-home_action_extratos button,
.st-key-home_action_razao button {
    position: relative !important;
    width: 100% !important;
    min-height: 92px !important;
    height: 92px !important;
    padding: 0.9rem 1rem 0.9rem 3.8rem !important;
    margin-bottom: 0.48rem !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    text-align: left !important;
    box-sizing: border-box !important;
}
.st-key-home_action_organizador button p,
.st-key-home_action_extratos button p,
.st-key-home_action_razao button p {
    display: block !important;
    width: 100% !important;
    max-width: none !important;
    margin: 0 !important;
    padding: 0 !important;
    text-align: left !important;
    white-space: pre-line !important;
}
.st-key-home_action_organizador button::before,
.st-key-home_action_extratos button::before,
.st-key-home_action_razao button::before {
    position: absolute;
    left: 1.15rem;
    top: 50%;
    width: 1.65rem;
    height: 1.65rem;
    display: grid;
    place-items: center;
    transform: translateY(-50%);
    border: 1px solid var(--rz-line);
    border-radius: 7px;
    background: color-mix(in srgb, var(--rz-panel) 82%, transparent);
    color: var(--rz-accent);
    font-size: 0.82rem;
    line-height: 1;
}
.st-key-home_action_organizador button::before { content: "▤"; }
.st-key-home_action_extratos button::before { content: "⇄"; }
.st-key-home_action_razao button::before { content: "✓"; }

.rz-overview-panel {
    min-height: 292px !important;
    height: 292px !important;
    padding: 1rem 1.05rem !important;
    box-sizing: border-box !important;
}
.rz-overview-copy {
    margin-bottom: 0.72rem !important;
}
.rz-overview-row {
    padding: 0.55rem 0 !important;
}

@media (max-width: 900px) {
    .st-key-home_action_organizador button,
    .st-key-home_action_extratos button,
    .st-key-home_action_razao button {
        min-height: 86px !important;
        height: 86px !important;
        padding-left: 3.45rem !important;
    }
    .rz-overview-panel {
        height: auto !important;
        min-height: 0 !important;
    }
}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
/* Grade fixa de ícones e textos da navegação lateral. */
.rz-nav-label {
    padding-left: 0.85rem !important;
    margin-top: 0.65rem !important;
}
section[data-testid="stSidebar"] [class*="st-key-sb_"] button {
    position: relative !important;
    width: 100% !important;
    min-height: 2.5rem !important;
    padding: 0.48rem 0.7rem 0.48rem 2.55rem !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    text-align: left !important;
}
section[data-testid="stSidebar"] [class*="st-key-sb_"] button p {
    display: block !important;
    width: 100% !important;
    margin: 0 !important;
    padding: 0 !important;
    text-align: left !important;
    white-space: nowrap !important;
}
section[data-testid="stSidebar"] [class*="st-key-sb_"] button::before {
    position: absolute;
    left: 0.86rem;
    top: 50%;
    width: 1rem;
    display: block;
    transform: translateY(-50%);
    color: currentColor;
    font-size: 0.78rem;
    line-height: 1;
    text-align: center;
}
section[data-testid="stSidebar"] .st-key-sb_home button::before { content: "⌂"; }
section[data-testid="stSidebar"] .st-key-sb_extratos button::before { content: "⇄"; }
section[data-testid="stSidebar"] .st-key-sb_razao button::before { content: "✓"; }
section[data-testid="stSidebar"] .st-key-sb_organizador button::before { content: "▤"; }
</style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
.rz-nav-label {
    display: block !important;
    line-height: 1.35 !important;
    margin-bottom: 0 !important;
}
.rz-nav-title-gap {
    display: block;
    width: 100%;
    height: 0.58rem;
    pointer-events: none;
}
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# CONTROLE DE ESTADO DE NAVEGAÇÃO
# ==============================================================================
if 'pagina_ativa' not in st.session_state:
    st.session_state['pagina_ativa'] = 'home'
if 'animar_transicao' not in st.session_state:
    st.session_state['animar_transicao'] = True

def mudar_pagina(nome_pagina):
    """Troca a ferramenta e anima somente o primeiro render da nova tela."""
    pagina_anterior = st.session_state.get('pagina_ativa')
    if pagina_anterior == nome_pagina:
        if nome_pagina == 'organizador':
            st.session_state['empresa_organizador'] = None
        return
    # Sempre inicia o Organizador pela escolha da empresa. A seleção permanece
    # apenas durante o trabalho atual e não reaparece ao entrar novamente.
    if nome_pagina == 'organizador':
        st.session_state['empresa_organizador'] = None
    st.session_state['pagina_ativa'] = nome_pagina
    st.session_state['animar_transicao'] = True

# ==============================================================================
# BARRA LATERAL
# ==============================================================================
st.sidebar.markdown(
    (
        '<div class="rz-nav-label">Navegação</div>'
        '<div class="rz-nav-title-gap" aria-hidden="true"></div>'
    ),
    unsafe_allow_html=True,
)

pagina_sidebar = st.session_state.get('pagina_ativa', 'home')
st.sidebar.button(
    "Início",
    use_container_width=True,
    key="sb_home",
    type="primary" if pagina_sidebar == "home" else "tertiary",
    on_click=mudar_pagina,
    args=('home',),
)
st.sidebar.button(
    "Conversor de Extratos",
    use_container_width=True,
    key="sb_extratos",
    type="primary" if pagina_sidebar == "extratos" else "tertiary",
    on_click=mudar_pagina,
    args=('extratos',),
)
st.sidebar.button(
    "Conciliação com Razão",
    use_container_width=True,
    key="sb_razao",
    type="primary" if pagina_sidebar == "razao" else "tertiary",
    on_click=mudar_pagina,
    args=('razao',),
)
st.sidebar.button(
    "Organizador de Planilhas",
    use_container_width=True,
    key="sb_organizador",
    type="primary" if pagina_sidebar == "organizador" else "tertiary",
    on_click=mudar_pagina,
    args=('organizador',),
)

if SEGURANCA_POR_SENHA_ATIVA:
    st.sidebar.markdown(
        '<div class="rz-nav-label" style="margin-top:0.9rem;">Sessão</div>',
        unsafe_allow_html=True,
    )
    st.sidebar.button(
        "Sair do sistema",
        use_container_width=True,
        key="hc_encerrar_sessao",
        type="tertiary",
        on_click=lambda: st.session_state.update({'_hc_acesso_autorizado': False}),
    )

st.sidebar.markdown(
    "<p style='font-size:10px;color:var(--hc-muted);text-align:center;"
    "position:fixed;left:1rem;bottom:1.25rem;width:230px;'>"
    "Razync · Ambiente protegido</p>",
    unsafe_allow_html=True,
)

# O marcador ativa o CSS uma única vez e desaparece nos reruns de filtros/uploads.
if st.session_state.pop('animar_transicao', False):
    st.markdown(
        '<span class="hc-page-transition-marker" aria-hidden="true"></span>',
        unsafe_allow_html=True
    )

# ==============================================================================
# TELA 1: MENU PRINCIPAL (HOME)
# ==============================================================================
if st.session_state['pagina_ativa'] == 'home':
    st.markdown(
        """
        <section class="rz-dashboard-intro" aria-labelledby="rz-home-title">
            <div class="rz-home-eyebrow">Central operacional</div>
            <div class="rz-home-title" id="rz-home-title">O que você precisa fazer hoje?</div>
            <div class="rz-home-copy">
                Acesse diretamente o fluxo necessário. Cada ferramenta mantém as regras,
                contas e formatos definidos para a operação contábil.
            </div>
        </section>
        """,
        unsafe_allow_html=True,
    )

    col_acoes, col_visao = st.columns([1.35, 0.65], gap="large")
    with col_acoes:
        st.markdown(
            '<div class="rz-dashboard-grid-title">Ações operacionais</div>',
            unsafe_allow_html=True,
        )
        st.button(
            "**Organizador de Planilhas**\nFluxos específicos, empresas e Base Inteligente.",
            key="home_action_organizador",
            use_container_width=True,
            on_click=mudar_pagina,
            args=('organizador',),
        )
        st.button(
            "**Conversor de Extratos**\nPDF, OFX, CSV e Excel para o padrão Domínio.",
            key="home_action_extratos",
            use_container_width=True,
            on_click=mudar_pagina,
            args=('extratos',),
        )
        st.button(
            "**Conciliação com Razão**\nConferência diária e identificação de divergências.",
            key="home_action_razao",
            use_container_width=True,
            on_click=mudar_pagina,
            args=('razao',),
        )

    with col_visao:
        st.markdown(
            """
            <div class="rz-dashboard-grid-title">Visão do ambiente</div>
            <section class="rz-overview-panel" aria-label="Recursos do Razync">
                <div class="rz-overview-kicker">Razync</div>
                <div class="rz-overview-title">Operação centralizada</div>
                <div class="rz-overview-copy">
                    Ferramentas bancárias e contábeis reunidas em um único fluxo de trabalho.
                </div>
                <div class="rz-overview-row">
                    <i class="rz-overview-dot"></i>
                    <div><strong>48 empresas cadastradas</strong>
                    <span>Áreas individuais preparadas para regras específicas.</span></div>
                </div>
                <div class="rz-overview-row">
                    <i class="rz-overview-dot"></i>
                    <div><strong>Arquivos bancários</strong>
                    <span>PDF, OFX, CSV, XLSX e XLS.</span></div>
                </div>
                <div class="rz-overview-row">
                    <i class="rz-overview-dot"></i>
                    <div><strong>Saída para a Domínio</strong>
                    <span>Modelo preservado, classificação e conferência.</span></div>
                </div>
            </section>
            """,
            unsafe_allow_html=True,
        )

# ==============================================================================
# TELA 2: FERRAMENTA DE CONVERSÃO DE EXTRATOS
# ==============================================================================
elif st.session_state['pagina_ativa'] == 'extratos':
    if st.button("← Início", key="btn_voltar_home", type="tertiary"):
        mudar_pagina('home')
        st.rerun()
    st.markdown(
        """
        <header class="rz-page-header">
            <div class="rz-page-kicker">Conversão bancária</div>
            <div class="rz-page-title">Conversor de Extratos</div>
            <div class="rz-page-description">
                Envie um ou mais extratos e gere arquivos prontos para importação na Domínio.
            </div>
        </header>
        """,
        unsafe_allow_html=True,
    )

    arquivos = st.file_uploader(
        "Selecione os extratos",
        type=["pdf", "ofx", "csv", "xlsx", "xls"],
        accept_multiple_files=True,
        help="Formatos aceitos: PDF, OFX, CSV, XLSX e XLS.",
    )

    if arquivos:
        try:
            colunas_dominio = ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
            df_modelo = pd.read_excel("Modelo dominio.xlsx") if os.path.exists("Modelo dominio.xlsx") else pd.DataFrame(columns=colunas_dominio)
            if 'DESCRIÇÃO' not in df_modelo.columns: df_modelo = pd.DataFrame(columns=colunas_dominio)
            
            dados_por_arquivo, todos_lancamentos_brutos = {}, []
            for arquivo in arquivos:
                file_bytes, extensao = arquivo.getvalue(), os.path.splitext(arquivo.name)[1].lower()
                lancamentos, data_ini_doc, data_fim_doc = [], None, None
                
                if extensao == '.pdf':
                    caminho_periodo = None
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_periodo:
                            temp_periodo.write(file_bytes)
                            caminho_periodo = temp_periodo.name
                        data_ini_doc, data_fim_doc = extrair_periodo_extrato(caminho_periodo)
                    finally:
                        if caminho_periodo and os.path.exists(caminho_periodo):
                            os.remove(caminho_periodo)

                lancamentos = executar_com_loading(
                    f"Analisando {arquivo.name}...",
                    processar_extrato_unificado,
                    file_bytes,
                    arquivo.name
                )
                    
                if lancamentos:
                    df_temp = pd.DataFrame(lancamentos)
                    df_temp['ARQUIVO_ORIGEM'] = arquivo.name
                    dados_por_arquivo[arquivo.name] = {'lancamentos': lancamentos, 'data_ini': data_ini_doc, 'data_fim': data_fim_doc}
                    todos_lancamentos_brutos.extend(lancamentos)

            if todos_lancamentos_brutos:
                nomes_abas = ["Visão Consolidada"] + [arq.name for arq in arquivos if arq.name in dados_por_arquivo] if len(arquivos) > 1 else [arq.name for arq in arquivos if arq.name in dados_por_arquivo]
                abas = st.tabs(nomes_abas)
                
                if len(arquivos) > 1:
                    with abas[0]:
                        st.markdown("### Resumo Consolidado")
                        df_geral_bruto = pd.DataFrame(todos_lancamentos_brutos)
                        df_geral_bruto['DATA_DT'] = pd.to_datetime(df_geral_bruto['DATA'], dayfirst=True, errors='coerce')
                        df_geral_bruto = df_geral_bruto.dropna(subset=['DATA_DT'])
                        
                        if df_geral_bruto.empty:
                            st.warning("Nenhum lançamento válido encontrado.")
                        else:
                            dt_min_geral, dt_max_geral = df_geral_bruto['DATA_DT'].min().date(), df_geral_bruto['DATA_DT'].max().date()
                            col_g1, col_g2, col_g3 = st.columns([1, 1, 1.5])
                            with col_g1: data_geral_ini = st.date_input("Data Inicial", value=dt_min_geral, min_value=dt_min_geral, max_value=dt_max_geral, format="DD/MM/YYYY", key="gen_ini")
                            with col_g2: data_geral_fim = st.date_input("Data Final", value=dt_max_geral, min_value=dt_min_geral, max_value=dt_max_geral, format="DD/MM/YYYY", key="gen_fim")
                            with col_g3: 
                                st.markdown("<label style='font-size:14px; font-weight:400; color:inherit;'>Busca rápida</label>", unsafe_allow_html=True)
                                termo_busca_geral = st.text_input("Busca rápida", placeholder="Filtrar histórico...", label_visibility="collapsed", key="gen_busca")
                            
                            df_geral_final = df_geral_bruto[(df_geral_bruto['DATA_DT'].dt.date >= data_geral_ini) & (df_geral_bruto['DATA_DT'].dt.date <= data_geral_fim)].copy()
                            if termo_busca_geral: df_geral_final = df_geral_final[df_geral_final['HISTÓRICO'].str.contains(termo_busca_geral, case=False, na=False)]
                            
                            df_geral_final = df_geral_final.drop(columns=['DATA_DT', 'ARQUIVO_ORIGEM'], errors='ignore')[df_modelo.columns]
                            df_geral_final = sanitizar_dataframe(df_geral_final)
                            
                            tot_cred_g, tot_deb_g = df_geral_final[df_geral_final['VALOR'] > 0]['VALOR'].sum(), df_geral_final[df_geral_final['VALOR'] < 0]['VALOR'].sum()
                            saldo_liq_g = tot_cred_g + tot_deb_g
                            
                            st.markdown("<br>", unsafe_allow_html=True)
                            cg1, cg2, cg3, cg4 = st.columns(4)
                            with cg1: st.markdown(f'<div class="metric-card"><div class="metric-title">Registros</div><div class="metric-value">{len(df_geral_final)}</div></div>', unsafe_allow_html=True)
                            with cg2: st.markdown(f'<div class="metric-card"><div class="metric-title">Entradas</div><div class="metric-value" style="color: #3fb950;">{formatar_moeda(tot_cred_g)}</div></div>', unsafe_allow_html=True)
                            with cg3: st.markdown(f'<div class="metric-card"><div class="metric-title">Saídas</div><div class="metric-value" style="color: #f85149;">{formatar_moeda(abs(tot_deb_g))}</div></div>', unsafe_allow_html=True)
                            with cg4:
                                color_g = "#3fb950" if saldo_liq_g >= 0 else "#f85149"
                                st.markdown(f'<div class="metric-card"><div class="metric-title">Saldo Líquido</div><div class="metric-value" style="color: {color_g};">{formatar_moeda(saldo_liq_g)}</div></div>', unsafe_allow_html=True)

                            st.markdown("<br>", unsafe_allow_html=True); st.markdown("##### Prévia Consolidada")
                            st.dataframe(formatar_dataframe_moeda_br(df_geral_final, ['VALOR']), use_container_width=True, height=280)
                            
                            st.markdown("##### Exportar")
                            cc_dl1, cc_dl2 = st.columns(2)
                            excel_modelo_g = gerar_excel_modelo_dominio(df_geral_final)
                            cc_dl1.download_button("Baixar Excel (.XLSX)", data=excel_modelo_g, file_name=f"consolidado_geral_{data_geral_ini.strftime('%d%m%Y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_excel_geral", use_container_width=True)
                            cc_dl2.download_button("Baixar TXT para Domínio", data=gerar_txt_dominio(df_geral_final), file_name=f"importacao_dominio_consolidado_{data_geral_ini.strftime('%d%m%Y')}.txt", mime="text/plain", key="dl_txt_geral", use_container_width=True)

                offset_abas = 1 if len(arquivos) > 1 else 0
                for idx_arq, arquivo in enumerate(arquivos):
                    if arquivo.name not in dados_por_arquivo: continue
                    with abas[idx_arq + offset_abas]:
                        info_arq = dados_por_arquivo[arquivo.name]
                        df_bruto = pd.DataFrame(info_arq['lancamentos'])
                        df_bruto['DATA_DT'] = pd.to_datetime(df_bruto['DATA'], dayfirst=True, errors='coerce')
                        df_bruto = df_bruto.dropna(subset=['DATA_DT'])
                        
                        if df_bruto.empty:
                            st.warning("Não há dados válidos neste arquivo.")
                            continue
                            
                        dt_min_dataset, dt_max_dataset = df_bruto['DATA_DT'].min().date(), df_bruto['DATA_DT'].max().date()
                        
                        data_ini_doc, data_fim_doc = info_arq['data_ini'], info_arq['data_fim']
                        val_ini_def = max(min(data_ini_doc.date(), dt_max_dataset), dt_min_dataset) if data_ini_doc and data_ini_doc.date() else dt_min_dataset
                        val_fim_def = max(min(data_fim_doc.date(), dt_max_dataset), dt_min_dataset) if data_fim_doc and data_fim_doc.date() else dt_max_dataset
                        if val_ini_def > val_fim_def: val_ini_def, val_fim_def = dt_min_dataset, dt_max_dataset
                        
                        with st.container():
                            col_f1, col_f2, col_f3 = st.columns([1, 1, 1.5])
                            with col_f1: data_sel_ini = st.date_input("Data Inicial", value=val_ini_def, min_value=dt_min_dataset, max_value=dt_max_dataset, format="DD/MM/YYYY", key=f"ini_{idx_arq}")
                            with col_f2: data_sel_fim = st.date_input("Data Final", value=val_fim_def, min_value=dt_min_dataset, max_value=dt_max_dataset, format="DD/MM/YYYY", key=f"fim_{idx_arq}")
                            with col_f3: 
                                st.markdown("<label style='font-size:14px; font-weight:400; color:inherit;'>Busca rápida</label>", unsafe_allow_html=True)
                                termo_busca = st.text_input("Busca rápida", placeholder="Digite para filtrar...", label_visibility="collapsed", key=f"busca_{idx_arq}")

                        df_final = df_bruto[(df_bruto['DATA_DT'].dt.date >= data_sel_ini) & (df_bruto['DATA_DT'].dt.date <= data_sel_fim)].copy()
                        if termo_busca: df_final = df_final[df_final['HISTÓRICO'].str.contains(termo_busca, case=False, na=False)]
                        
                        df_final = df_final.drop(columns=['DATA_DT', 'ARQUIVO_ORIGEM'], errors='ignore')[df_modelo.columns]
                        df_final = sanitizar_dataframe(df_final)

                        total_creditos, total_debitos = df_final[df_final['VALOR'] > 0]['VALOR'].sum(), df_final[df_final['VALOR'] < 0]['VALOR'].sum()
                        saldo_liquido = total_creditos + total_debitos
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        c1, c2, c3, c4 = st.columns(4)
                        with c1: st.markdown(f'<div class="metric-card"><div class="metric-title">Registros</div><div class="metric-value">{len(df_final)}</div></div>', unsafe_allow_html=True)
                        with c2: st.markdown(f'<div class="metric-card"><div class="metric-title">Entradas</div><div class="metric-value" style="color: #3fb950;">{formatar_moeda(total_creditos)}</div></div>', unsafe_allow_html=True)
                        with c3: st.markdown(f'<div class="metric-card"><div class="metric-title">Saídas</div><div class="metric-value" style="color: #f85149;">{formatar_moeda(abs(total_debitos))}</div></div>', unsafe_allow_html=True)
                        with c4:
                            color_liq = "#3fb950" if saldo_liquido >= 0 else "#f85149"
                            st.markdown(f'<div class="metric-card"><div class="metric-title">Saldo Líquido</div><div class="metric-value" style="color: {color_liq};">{formatar_moeda(saldo_liquido)}</div></div>', unsafe_allow_html=True)

                        st.markdown("<br>", unsafe_allow_html=True); st.markdown("##### Prévia dos Lançamentos")
                        st.dataframe(formatar_dataframe_moeda_br(df_final, ['VALOR']), use_container_width=True, height=280)
                        
                        st.markdown("##### Exportar")
                        c_dl1, c_dl2 = st.columns(2)
                        excel_modelo = gerar_excel_modelo_dominio(df_final)
                        c_dl1.download_button("Baixar Excel (.XLSX)", data=excel_modelo, file_name=f"lancamentos_{os.path.splitext(arquivo.name)[0]}_{data_sel_ini.strftime('%d%m%Y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"excel_{idx_arq}", use_container_width=True)
                        c_dl2.download_button("Baixar TXT para Domínio", data=gerar_txt_dominio(df_final), file_name=f"importacao_dominio_{os.path.splitext(arquivo.name)[0]}_{data_sel_ini.strftime('%d%m%Y')}.txt", mime="text/plain", key=f"txt_{idx_arq}", use_container_width=True)
        except Exception as e:
            st.error(f"🛑 Ocorreu um erro na aba extratos. Detalhes: {e}")

# ==============================================================================
# TELA 3: ORGANIZADOR DE PLANILHAS POR EMPRESA
# ==============================================================================
elif st.session_state['pagina_ativa'] == 'organizador':
    st.markdown(
        """
        <style>
        .rz-company-workspace {
            display: grid;
            grid-template-columns: minmax(0, 1.25fr) minmax(0, 1fr) minmax(0, 1fr);
            gap: 0;
            margin: -0.35rem 0 1.4rem;
            border-top: 1px solid rgba(148, 163, 184, 0.22);
            border-bottom: 1px solid rgba(148, 163, 184, 0.22);
        }
        .rz-company-workspace__item {
            min-width: 0;
            padding: 0.85rem 1rem 0.8rem 0;
        }
        .rz-company-workspace__item + .rz-company-workspace__item {
            padding-left: 1rem;
            border-left: 1px solid rgba(148, 163, 184, 0.18);
        }
        .rz-company-workspace__label {
            display: block;
            margin-bottom: 0.22rem;
            color: #7f93a5;
            font-size: 0.66rem;
            font-weight: 700;
            letter-spacing: 0.09em;
            line-height: 1.2;
            text-transform: uppercase;
        }
        .rz-company-workspace__value {
            display: block;
            color: #e8eef5;
            font-size: 0.84rem;
            font-weight: 600;
            line-height: 1.35;
        }
        [data-testid="stTabs"] {
            margin-top: 0.15rem;
        }
        [data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 1.7rem !important;
            border-bottom: 1px solid rgba(148, 163, 184, 0.24) !important;
        }
        [data-testid="stTabs"] button[role="tab"] {
            min-height: 2.65rem !important;
            padding: 0 0 0.7rem !important;
            border-radius: 0 !important;
            background: transparent !important;
            font-size: 0.83rem !important;
            font-weight: 600 !important;
        }
        @media (max-width: 700px) {
            .rz-company-workspace {
                grid-template-columns: 1fr;
            }
            .rz-company-workspace__item,
            .rz-company-workspace__item + .rz-company-workspace__item {
                padding: 0.65rem 0;
                border-left: 0;
            }
            .rz-company-workspace__item + .rz-company-workspace__item {
                border-top: 1px solid rgba(148, 163, 184, 0.14);
            }
            [data-testid="stTabs"] [data-baseweb="tab-list"] {
                gap: 1rem !important;
            }
            [data-testid="stTabs"] button[role="tab"] {
                font-size: 0.76rem !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True
    )
    if 'empresa_organizador' not in st.session_state:
        st.session_state['empresa_organizador'] = None
    empresa_organizador = st.session_state['empresa_organizador']
    empresa_catalogo_atual = EMPRESAS_POR_CHAVE.get(str(empresa_organizador))
    if empresa_catalogo_atual is None and empresa_organizador:
        empresas_mesma_chave = [
            empresa
            for empresas_regime in EMPRESAS_POR_REGIME.values()
            for empresa in empresas_regime
            if empresa.get('chave_sistema') == empresa_organizador
        ]
        if empresa_organizador == 'nova_geracao' and empresas_mesma_chave:
            estabelecimento_atual = st.session_state.get(
                'org_estabelecimento_nova_geracao_card', 'matriz'
            )
            empresa_catalogo_atual = next(
                (empresa for empresa in empresas_mesma_chave
                 if empresa.get('estabelecimento', 'matriz') == estabelecimento_atual),
                empresas_mesma_chave[0]
            )
        elif empresas_mesma_chave:
            empresa_catalogo_atual = empresas_mesma_chave[0]

    col_voltar, col_tit = st.columns([1.2, 8.8])
    with col_voltar:
        st.markdown("<div style='height: 4px;'></div>", unsafe_allow_html=True)
        if empresa_organizador:
            if st.button(
                "← Empresas", use_container_width=True, key="btn_voltar_empresas_org"
            ):
                st.session_state['empresa_organizador'] = None
                st.rerun()
        elif st.button("← Voltar", use_container_width=True, key="btn_voltar_home_org"):
            mudar_pagina('home')
            st.rerun()
    estabelecimento_ng_atual = st.session_state.get(
        'org_estabelecimento_nova_geracao_card', 'matriz'
    )
    titulo_nova_geracao_atual = (
        '1396 - Nova Geração Filial'
        if estabelecimento_ng_atual == 'filial'
        else '266 - Nova Geração'
    )

    with col_tit:
        st.title({
            'nova_geracao': titulo_nova_geracao_atual,
            'autokraft_industrial': '3 - Autokraft Industrial',
            'autokraft_projetos': '178 - Autokraft Projetos',
            'isa': '343 - I.S.A',
            'accede_automacao': '1000 - ACCEDE AUTOMAÇÃO',
            'accede_equipamentos': '1001 - ACCEDE EQUIPAMENTOS',
            'dias_pereira': '1529 - Dias e Pereira'
        }.get(
            empresa_organizador,
            empresa_catalogo_atual['rotulo'] if empresa_catalogo_atual else 'Organizador de Planilhas'
        ))
    st.caption({
        'nova_geracao': f'Organize, confira e classifique os movimentos da {titulo_nova_geracao_atual}.',
        'autokraft_industrial': 'Organize os mapas diários e confira os extratos da 3 - Autokraft Industrial.',
        'autokraft_projetos': 'Organize os mapas diários e confira os extratos da 178 - Autokraft Projetos.',
        'isa': 'Organize os mapas diários e confira os extratos da 343 - I.S.A.',
        'accede_automacao': 'Organize as planilhas SIG e confira Itaú e Sicredi da 1000 - ACCEDE AUTOMAÇÃO.',
        'accede_equipamentos': 'Organize as planilhas SIG e confira Itaú e Sicredi da 1001 - ACCEDE EQUIPAMENTOS.',
        'dias_pereira': 'Converta o relatório visual do Nibo da 1529 - Dias e Pereira diretamente para o Modelo Domínio.'
    }.get(
        empresa_organizador,
        (
            f"{empresa_catalogo_atual['regime'].title()} · Área cadastrada para receber ferramentas específicas."
            if empresa_catalogo_atual
            else 'Pesquise uma empresa para acessar sua área de organização e Base Inteligente.'
        )
    ))
    st.markdown("---")

    if empresa_organizador:
        regime_workspace = (
            empresa_catalogo_atual.get('regime', 'Regime cadastrado').title()
            if empresa_catalogo_atual
            else 'Regime cadastrado'
        )
        st.markdown(
            f"""
            <div class="rz-company-workspace" aria-label="Resumo da área da empresa">
                <div class="rz-company-workspace__item">
                    <span class="rz-company-workspace__label">Empresa</span>
                    <span class="rz-company-workspace__value">{empresa_catalogo_atual.get('rotulo', 'Área individual') if empresa_catalogo_atual else 'Área individual'}</span>
                </div>
                <div class="rz-company-workspace__item">
                    <span class="rz-company-workspace__label">Regime</span>
                    <span class="rz-company-workspace__value">{regime_workspace}</span>
                </div>
                <div class="rz-company-workspace__item">
                    <span class="rz-company-workspace__label">Ferramentas</span>
                    <span class="rz-company-workspace__value">Organização · Base Inteligente</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    if empresa_organizador is None:
        def _normalizar_busca_empresa(valor):
            texto = unicodedata.normalize('NFKD', str(valor))
            texto = ''.join(
                caractere for caractere in texto
                if not unicodedata.combining(caractere)
            )
            return texto.casefold().strip()

        def _abrir_empresa_catalogo(empresa_catalogo):
            chave_destino = empresa_catalogo.get(
                'chave_sistema', empresa_catalogo['chave']
            )
            if chave_destino == 'nova_geracao':
                st.session_state['org_estabelecimento_nova_geracao_card'] = (
                    empresa_catalogo.get('estabelecimento', 'matriz')
                )
            st.session_state['empresa_organizador'] = chave_destino
            st.rerun()

        empresas_catalogo_completo = sorted(
            [
                empresa
                for empresas_regime in EMPRESAS_POR_REGIME.values()
                for empresa in empresas_regime
            ],
            key=lambda item: item['codigo']
        )
        empresas_ativas = [
            empresa
            for empresa in empresas_catalogo_completo
            if empresa.get('chave_sistema')
        ]

        st.markdown(
            """
            <style>
            [class*="st-key-org_seletor_empresa"] {
                width: min(100%, 720px);
                margin: 1.15rem auto 0;
            }
            [class*="st-key-org_seletor_empresa"] [data-testid="stPopover"] > button {
                min-height: 3.25rem !important;
                padding: 0 1rem !important;
                justify-content: flex-start !important;
                color: #9aabba !important;
                background: rgba(148, 163, 184, 0.04) !important;
                border: 1px solid rgba(148, 163, 184, 0.24) !important;
                border-radius: 9px !important;
                box-shadow: none !important;
                font-size: 0.87rem !important;
                font-weight: 500 !important;
                text-align: left !important;
            }
            [class*="st-key-org_seletor_empresa"] [data-testid="stPopover"] > button:hover {
                color: var(--text-color) !important;
                border-color: rgba(34, 191, 230, 0.55) !important;
                background: rgba(34, 191, 230, 0.04) !important;
            }
            div[data-testid="stPopoverBody"] {
                width: min(680px, calc(100vw - 32px)) !important;
                max-width: min(680px, calc(100vw - 32px)) !important;
                padding: 1rem !important;
                border: 1px solid rgba(148, 163, 184, 0.24) !important;
                border-radius: 10px !important;
                background: #0b121a !important;
                box-shadow: 0 18px 46px rgba(0, 0, 0, 0.34) !important;
            }
            .rz-company-picker__eyebrow {
                margin: 0 0 0.18rem;
                color: #22bfe6;
                font-size: 0.64rem;
                font-weight: 750;
                letter-spacing: 0.1em;
                text-transform: uppercase;
            }
            .rz-company-picker__title {
                margin: 0 0 0.75rem;
                color: var(--text-color);
                font-size: 0.95rem;
                font-weight: 680;
            }
            [class*="st-key-org_busca_no_popover"] [data-testid="stTextInput"] input {
                min-height: 3rem;
                color: var(--text-color) !important;
                background: #0d1721 !important;
                border: 1px solid rgba(34, 191, 230, 0.62) !important;
                border-radius: 7px !important;
                box-shadow: none !important;
                font-size: 0.84rem !important;
            }
            [class*="st-key-org_busca_no_popover"] [data-testid="stTextInput"] input:focus {
                border-color: #22bfe6 !important;
                box-shadow: 0 0 0 2px rgba(34, 191, 230, 0.08) !important;
            }
            [class*="st-key-org_resultados_em_tabela"] {
                margin-top: 0.55rem;
            }
            [class*="st-key-org_resultados_em_tabela"] [data-testid="stDataFrame"] {
                border: 1px solid rgba(148, 163, 184, 0.2);
                border-radius: 7px;
                overflow: hidden;
            }
            @media (max-width: 640px) {
                [class*="st-key-org_seletor_empresa"] {
                    width: 100%;
                    margin-top: 0.75rem;
                }
                div[data-testid="stPopoverBody"] {
                    padding: 0.8rem !important;
                }
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        with st.container(key='org_seletor_empresa'):
            with st.popover(
                '⌕   Pesquisar empresa',
                use_container_width=True,
            ):
                st.markdown(
                    '<p class="rz-company-picker__eyebrow">Empresas</p>'
                    '<p class="rz-company-picker__title">Pesquisar empresa</p>',
                    unsafe_allow_html=True,
                )
                with st.container(key='org_busca_no_popover'):
                    termo_busca_empresas = st.text_input(
                        'Pesquisar empresa',
                        placeholder='Digite o código ou nome',
                        key='org_busca_empresas_catalogo',
                        label_visibility='collapsed',
                    )
                termo_normalizado = _normalizar_busca_empresa(termo_busca_empresas)

                if not termo_normalizado:
                    st.caption('Digite para localizar uma empresa.')
                else:
                    empresas_encontradas = []
                    for empresa_catalogo in empresas_catalogo_completo:
                        alvo = _normalizar_busca_empresa(
                            f"{empresa_catalogo['codigo']} {empresa_catalogo['nome']}"
                        )
                        if termo_normalizado in alvo:
                            empresas_encontradas.append(empresa_catalogo)

                    if not empresas_encontradas:
                        st.caption('Nenhuma empresa encontrada.')
                    else:
                        empresas_exibidas = empresas_encontradas[:8]
                        dados_resultados_empresa = pd.DataFrame([
                            {
                                'Código': str(empresa_catalogo['codigo']),
                                'Empresa': empresa_catalogo['nome'],
                                'Regime': empresa_catalogo.get(
                                    'regime', 'Não informado'
                                ).title(),
                            }
                            for empresa_catalogo in empresas_exibidas
                        ])
                        with st.container(key='org_resultados_em_tabela'):
                            evento_resultado_empresa = st.dataframe(
                                dados_resultados_empresa,
                                hide_index=True,
                                use_container_width=True,
                                height=min(
                                    360,
                                    38 + (len(dados_resultados_empresa) * 42)
                                ),
                                column_config={
                                    'Código': st.column_config.TextColumn(
                                        'Código', width='small'
                                    ),
                                    'Empresa': st.column_config.TextColumn(
                                        'Empresa', width='large'
                                    ),
                                    'Regime': st.column_config.TextColumn(
                                        'Regime', width='medium'
                                    ),
                                },
                                on_select='rerun',
                                selection_mode='single-row',
                                key='org_tabela_resultados_empresa',
                            )
                        linhas_resultado = evento_resultado_empresa.selection.rows
                        if linhas_resultado:
                            indice_resultado = linhas_resultado[0]
                            if 0 <= indice_resultado < len(empresas_exibidas):
                                _abrir_empresa_catalogo(
                                    empresas_exibidas[indice_resultado]
                                )
                        if len(empresas_encontradas) > 8:
                            st.caption('Continue digitando para refinar a pesquisa.')

    if empresa_catalogo_atual and not empresa_catalogo_atual.get('chave_sistema'):
        st.markdown(f"### {empresa_catalogo_atual['rotulo']}")
        st.caption(f"Regime tributário: {empresa_catalogo_atual['regime'].title()}")
        st.info(
            "Empresa cadastrada no Razync. As ferramentas específicas desta empresa "
            "ainda não foram configuradas."
        )

    if st.session_state['empresa_organizador'] in {
        'autokraft_industrial', 'autokraft_projetos', 'isa'
    }:
        configuracao_empresa_autokraft = CONFIGURACOES_AUTOKRAFT[
            st.session_state['empresa_organizador']
        ]
        empresa_autokraft = configuracao_empresa_autokraft['empresa']
        slug_empresa_autokraft = configuracao_empresa_autokraft["slug"]

        aba_operacoes_autokraft, aba_base_autokraft = st.tabs([
            "Organizar arquivos",
            "Base Inteligente"
        ])

        with aba_base_autokraft:
            renderizar_base_inteligente_empresa(
                slug_empresa_autokraft,
                empresa_autokraft,
                {'itau', 'daycoval'},
                configuracao_empresa_autokraft['contas_bancarias']
            )

        with aba_operacoes_autokraft:
            st.caption(
            f"Ferramentas ativas para {empresa_autokraft}. O sistema lê automaticamente "
            "cada aba diária, ignora saldos e totais e separa os lançamentos por banco."
        )
            bancos_autokraft = st.multiselect(
                "Bancos para organizar",
                ["Itaú", "Daycoval"],
                default=["Itaú", "Daycoval"],
                key=f"org_bancos_{slug_empresa_autokraft}"
            )
            arquivo_autokraft = st.file_uploader(
                f"Envie o mapa bancário da {empresa_autokraft}",
                type=['xlsx', 'xls'],
                key=f"upload_mapa_{slug_empresa_autokraft}",
                help="O arquivo pode conter todas as abas diárias do mês."
            )

            if arquivo_autokraft is not None:
                try:
                    dados_autokraft, abas_autokraft = processar_mapa_autokraft(
                        arquivo_autokraft.getvalue(), arquivo_autokraft.name
                    )
                    datas_disponiveis = []
                    for dados_banco in dados_autokraft.values():
                        df_banco = dados_banco['principal']
                        if not df_banco.empty:
                            datas_disponiveis.extend(
                                pd.to_datetime(df_banco['DATA'], errors='coerce').dropna().dt.date.tolist()
                            )
                    if not datas_disponiveis:
                        raise ValueError("Nenhuma data válida foi localizada nas abas diárias.")

                    data_min_autokraft = min(datas_disponiveis)
                    data_max_autokraft = max(datas_disponiveis)
                    col_data_ak1, col_data_ak2 = st.columns(2)
                    with col_data_ak1:
                        data_ini_autokraft = st.date_input(
                            "Data inicial",
                            value=data_min_autokraft,
                            min_value=data_min_autokraft,
                            max_value=data_max_autokraft,
                            format="DD/MM/YYYY",
                            key=f"data_ini_{slug_empresa_autokraft}"
                        )
                    with col_data_ak2:
                        data_fim_autokraft = st.date_input(
                            "Data final",
                            value=data_max_autokraft,
                            min_value=data_min_autokraft,
                            max_value=data_max_autokraft,
                            format="DD/MM/YYYY",
                            key=f"data_fim_{slug_empresa_autokraft}"
                        )

                    if data_ini_autokraft > data_fim_autokraft:
                        st.warning("A data inicial deve ser anterior ou igual à data final.")
                    elif not bancos_autokraft:
                        st.warning("Selecione pelo menos um banco para gerar a planilha.")
                    else:
                        dados_filtrados_autokraft = {}
                        for nome_banco in bancos_autokraft:
                            df_filtrado = filtrar_dataframe_periodo(
                                dados_autokraft[nome_banco]['principal'],
                                data_ini_autokraft,
                                data_fim_autokraft
                            )
                            dados_filtrados_autokraft[nome_banco] = {
                                'principal': df_filtrado,
                                'retirados': pd.DataFrame()
                            }

                        df_resumo_autokraft = pd.concat(
                            [dados['principal'] for dados in dados_filtrados_autokraft.values()],
                            ignore_index=True
                        )
                        total_autokraft = len(df_resumo_autokraft)
                        entradas_autokraft = df_resumo_autokraft.loc[
                            df_resumo_autokraft['VALOR'] > 0, 'VALOR'
                        ].sum() if not df_resumo_autokraft.empty else 0
                        saidas_autokraft = abs(df_resumo_autokraft.loc[
                            df_resumo_autokraft['VALOR'] < 0, 'VALOR'
                        ].sum()) if not df_resumo_autokraft.empty else 0

                        met_ak1, met_ak2, met_ak3 = st.columns(3)
                        with met_ak1:
                            st.metric("Lançamentos", total_autokraft)
                        with met_ak2:
                            st.metric("Entradas", formatar_moeda(entradas_autokraft))
                        with met_ak3:
                            st.metric("Saídas", formatar_moeda(saidas_autokraft))

                        st.caption(
                            f"{len(abas_autokraft)} abas diárias reconhecidas, de "
                            f"{data_min_autokraft.strftime('%d/%m/%Y')} a "
                            f"{data_max_autokraft.strftime('%d/%m/%Y')}."
                        )
                        if df_resumo_autokraft.empty:
                            st.warning("Não há lançamentos para os bancos e período escolhidos.")
                        else:
                            arquivo_final_autokraft = gerar_excel_nova_geracao(
                                dados_filtrados_autokraft
                            )
                            st.download_button(
                                "Baixar planilha no Modelo Domínio",
                                data=arquivo_final_autokraft,
                                file_name=(
                                    f"{configuracao_empresa_autokraft['arquivo']}_"
                                    f"{data_ini_autokraft.strftime('%d%m%Y')}_a_"
                                    f"{data_fim_autokraft.strftime('%d%m%Y')}.xlsx"
                                ),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"download_{slug_empresa_autokraft}"
                            )
                except Exception as erro_autokraft:
                    st.error(
                        f"Não foi possível processar o mapa de {empresa_autokraft}: "
                        f"{erro_autokraft}"
                    )

            st.markdown(f"#### Conferência — {empresa_autokraft}")
            renderizar_conferencia_autokraft(slug_empresa_autokraft)


    if st.session_state['empresa_organizador'] in {'accede_automacao', 'accede_equipamentos'}:
        chave_accede = st.session_state['empresa_organizador']
        config_accede = CONFIGURACOES_ACCEDE[chave_accede]
        empresa_accede = config_accede['empresa']
        slug_accede = config_accede['slug']

        aba_operacoes_accede, aba_base_accede = st.tabs([
            'Organizar arquivos',
            'Base Inteligente'
        ])

        with aba_base_accede:
            renderizar_base_inteligente_empresa(
                slug_accede,
                empresa_accede,
                {'itau', 'sicredi'},
                config_accede['contas_bancarias']
            )

        with aba_operacoes_accede:
            st.caption(
                'Envie as planilhas SIG do Itaú e/ou Sicredi. Linhas sem DATA abaixo '
                'de um lançamento são tratadas como detalhamento do mesmo grupo.'
            )
            col_itau_accede, col_sicredi_accede = st.columns(2)
            with col_itau_accede:
                arquivo_itau_accede = st.file_uploader(
                    'Planilha SIG — Itaú',
                    type=['xlsx', 'xls'],
                    key=f'{slug_accede}_sig_itau'
                )
            with col_sicredi_accede:
                arquivo_sicredi_accede = st.file_uploader(
                    'Planilha SIG — Sicredi',
                    type=['xlsx', 'xls'],
                    key=f'{slug_accede}_sig_sicredi'
                )

            dados_accede = {}
            try:
                if arquivo_itau_accede is not None:
                    dados_accede['Itaú'] = {
                        'principal': executar_com_loading(
                            'Organizando a planilha SIG do Itaú...',
                            processar_planilha_accede_sig,
                            arquivo_itau_accede.getvalue(),
                            'Itaú'
                        ),
                        'retirados': pd.DataFrame()
                    }
                if arquivo_sicredi_accede is not None:
                    dados_accede['Sicredi'] = {
                        'principal': executar_com_loading(
                            'Organizando a planilha SIG do Sicredi...',
                            processar_planilha_accede_sig,
                            arquivo_sicredi_accede.getvalue(),
                            'Sicredi'
                        ),
                        'retirados': pd.DataFrame()
                    }

                if dados_accede:
                    df_accede = pd.concat(
                        [dados['principal'] for dados in dados_accede.values()],
                        ignore_index=True
                    ).sort_values(['DATA', 'DESCRIÇÃO'], kind='stable').reset_index(drop=True)
                    datas_accede = pd.to_datetime(df_accede['DATA'], errors='coerce').dropna().dt.date
                    data_min_accede = min(datas_accede)
                    data_max_accede = max(datas_accede)

                    met_ac1, met_ac2, met_ac3 = st.columns(3)
                    met_ac1.metric('Lançamentos', len(df_accede))
                    met_ac2.metric(
                        'Entradas',
                        formatar_moeda(df_accede.loc[df_accede['VALOR'] > 0, 'VALOR'].sum())
                    )
                    met_ac3.metric(
                        'Saídas',
                        formatar_moeda(abs(df_accede.loc[df_accede['VALOR'] < 0, 'VALOR'].sum()))
                    )
                    st.caption(
                        f'Período identificado: {data_min_accede.strftime("%d/%m/%Y")} a '
                        f'{data_max_accede.strftime("%d/%m/%Y")}.'
                    )

                    modelo_bytes_accede = None
                    for caminho_modelo in ['Modelo dominio.xlsx', 'Modelo dominio(6).xlsx']:
                        if os.path.exists(caminho_modelo):
                            with open(caminho_modelo, 'rb') as modelo_arquivo:
                                modelo_bytes_accede = modelo_arquivo.read()
                            break
                    arquivo_final_accede = gerar_excel_nova_geracao(
                        dados_accede, modelo_bytes_accede
                    )
                    st.download_button(
                        'Baixar planilha no Modelo Domínio',
                        data=arquivo_final_accede,
                        file_name=(
                            f"{config_accede['arquivo']}_"
                            f"{data_min_accede.strftime('%d%m%Y')}_a_"
                            f"{data_max_accede.strftime('%d%m%Y')}.xlsx"
                        ),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        key=f'{slug_accede}_download_modelo'
                    )
            except Exception as erro_accede:
                st.error(f'Não foi possível processar as planilhas da ACCEDE: {erro_accede}')

            st.markdown(f'#### Conferência — {empresa_accede}')
            renderizar_conferencia_autokraft(
                slug_accede,
                bancos_config=[
                    {'nome': 'Itaú', 'slug': 'itau'},
                    {'nome': 'Sicredi', 'slug': 'sicredi'}
                ]
            )

    if st.session_state['empresa_organizador'] == 'nova_geracao':
        st.markdown("<div class='ng-area-label'>Área da empresa</div>", unsafe_allow_html=True)
        if 'org_estabelecimento_nova_geracao_card' not in st.session_state:
            st.session_state['org_estabelecimento_nova_geracao_card'] = 'matriz'

        st.markdown(
            """
            <style>
            .ng-area-label {
                font-size: 12px;
                opacity: .72;
                margin: 0 0 5px 1px;
            }
            .st-key-ng_card_matriz button,
            .st-key-ng_card_filial button {
                width: 100% !important;
                height: 52px !important;
                min-height: 52px !important;
                max-height: 52px !important;
                padding: 6px 10px !important;
                border-radius: 8px !important;
                border: 1px solid #12324a !important;
                background: #050b12 !important;
                box-shadow: none !important;
                transform: none !important;
                font-size: 12px !important;
                line-height: 1.25 !important;
                font-weight: 600 !important;
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
                text-align: center !important;
                white-space: normal !important;
            }
            .st-key-ng_card_matriz button:hover,
            .st-key-ng_card_filial button:hover {
                background: #081725 !important;
                border-color: #1d6f9b !important;
                box-shadow: none !important;
                transform: none !important;
            }
            .st-key-ng_card_matriz_ativo button,
            .st-key-ng_card_filial_ativo button {
                width: 100% !important;
                height: 52px !important;
                min-height: 52px !important;
                max-height: 52px !important;
                padding: 6px 10px !important;
                border-radius: 8px !important;
                border: 1px solid #1d6f9b !important;
                background: #0b1f33 !important;
                box-shadow: none !important;
                transform: none !important;
                font-size: 12px !important;
                line-height: 1.25 !important;
                font-weight: 700 !important;
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
                text-align: center !important;
                white-space: normal !important;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        col_matriz, col_filial, col_restante = st.columns([0.19, 0.19, 0.62], gap='small')
        selecionado_ng = st.session_state['org_estabelecimento_nova_geracao_card']

        with col_matriz:
            chave_card_matriz = 'ng_card_matriz_ativo' if selecionado_ng == 'matriz' else 'ng_card_matriz'
            if st.button('266 - Nova Geração Matriz', key=chave_card_matriz, use_container_width=True):
                st.session_state['org_estabelecimento_nova_geracao_card'] = 'matriz'
                st.rerun()

        with col_filial:
            chave_card_filial = 'ng_card_filial_ativo' if selecionado_ng == 'filial' else 'ng_card_filial'
            if st.button('1396 - Nova Geração Filial', key=chave_card_filial, use_container_width=True):
                st.session_state['org_estabelecimento_nova_geracao_card'] = 'filial'
                st.rerun()

        chave_estabelecimento = st.session_state['org_estabelecimento_nova_geracao_card']
        nome_estabelecimento_nova = (
            '1396 - Nova Geração Filial' if chave_estabelecimento == 'filial' else '266 - Nova Geração Matriz'
        )
        empresa_base_nova = (
            'nova_geracao_filial'
            if chave_estabelecimento == 'filial'
            else 'nova_geracao_matriz'
        )
        nome_base_nova = (
            '1396 - Nova Geração Filial'
            if chave_estabelecimento == 'filial'
            else '266 - Nova Geração Matriz'
        )
        if chave_estabelecimento == 'filial':
            contas_dominio_estabelecimento = {'itau': '515', 'bradesco': '514'}
            configuracoes_bancos = {
                "Itaú - Conta 98002-6": {
                    "nome": "Itaú", "conta": "98002-6", "slug": "itau",
                    "processador": processar_nova_geracao_filial_itau
                },
                "Bradesco - Conta 3084-8": {
                    "nome": "Bradesco", "conta": "3084-8", "slug": "bradesco",
                    "processador": processar_nova_geracao_filial_bradesco
                }
            }
            st.caption(
                "1396 - Nova Geração Filial selecionada — Itaú 98002-6 usa a conta 515 e "
                "Bradesco 3084-8 usa a conta 514 na classificação."
            )
        else:
            contas_dominio_estabelecimento = {
                'itau': '508', 'bradesco': '9', 'fibra': '506'
            }
            configuracoes_bancos = {
                "Itaú - Conta 99549-5": {
                    "nome": "Itaú", "conta": "99549-5", "slug": "itau",
                    "processador": processar_nova_geracao_itau
                },
                "Bradesco - Conta 451990-6": {
                    "nome": "Bradesco", "conta": "451990-6", "slug": "bradesco",
                    "processador": processar_nova_geracao_bradesco
                },
                "Fibra - Conta 673947-1": {
                    "nome": "Fibra", "conta": "673947-1", "slug": "fibra",
                    "processador": processar_nova_geracao_fibra
                }
            }
            st.caption(
                "Matriz selecionada — Itaú usa a conta 508, Bradesco a conta 9 "
                "e Fibra a conta 506 na classificação."
            )
        url_base_classificacao, chave_base_classificacao, senha_admin_classificacao = (
            obter_config_classificacao_online()
        )
        base_classificacoes = []
        erro_base_classificacoes = ''
        if url_base_classificacao and chave_base_classificacao:
            try:
                base_classificacoes = carregar_classificacoes_online(empresa_base_nova)
            except Exception as erro_base:
                erro_base_classificacoes = str(erro_base)

        # Migração: a antiga base compartilhada não deve mais alimentar Matriz ou Filial.
        # Com a service role já configurada, ela é apagada automaticamente na primeira
        # abertura após esta atualização. As duas novas bases começam vazias.
        if url_base_classificacao and chave_base_classificacao:
            try:
                if not st.session_state.get('_nova_geracao_base_legada_verificada'):
                    apagar_classificacoes_online('nova_geracao')
                    st.session_state['_nova_geracao_base_legada_verificada'] = True
            except Exception as erro_limpeza_legada:
                st.session_state['_nova_geracao_erro_limpeza_legada'] = str(
                    erro_limpeza_legada
                )

        aba_operacoes, aba_base_inteligente = st.tabs([
            "Organizar arquivos",
            "Base Inteligente"
        ])

        with aba_base_inteligente:
            if erro_base_classificacoes:
                st.error(erro_base_classificacoes)
            elif url_base_classificacao and chave_base_classificacao:
                st.success(
                    f"{nome_base_nova}: {len(base_classificacoes)} padrões disponíveis."
                )
            else:
                st.warning(
                    "A base online ainda não foi configurada com os valores reais nos Secrets "
                    "do Streamlit. Substitua URL_DO_PROJETO_SUPABASE e SERVICE_ROLE_KEY pelos "
                    "dados do seu projeto Supabase. "
                    "A organização continuará funcionando, mas sem preencher Débito e Crédito."
                )
            st.caption(
                f"Base exclusiva de {nome_base_nova}. Matriz e Filial não compartilham mais "
                "nenhum padrão. Envie apenas planilhas antigas já classificadas desta área."
            )
            if st.session_state.get('_nova_geracao_erro_limpeza_legada'):
                st.warning(
                    "A separação das bases já está ativa, mas a base antiga compartilhada "
                    "não pôde ser apagada automaticamente: "
                    + st.session_state['_nova_geracao_erro_limpeza_legada']
                )
            arquivos_aprendizado = st.file_uploader(
                "Planilhas classificadas para ensinar o sistema",
                type=['xlsx', 'xls', 'zip'],
                accept_multiple_files=True,
                key=f'org_base_classificada_nova_{chave_estabelecimento}'
            )
            senha_aprendizado = st.text_input(
                "Senha administrativa para atualizar a base",
                type='password',
                key=f'org_senha_base_classificada_nova_{chave_estabelecimento}'
            )
            if st.button(
                "Importar classificações",
                key=f'org_importar_base_classificada_nova_{chave_estabelecimento}',
                use_container_width=True
            ):
                if not url_base_classificacao or not chave_base_classificacao:
                    st.error("Configure primeiro a conexão da base online nos Secrets do Streamlit.")
                elif not senha_admin_classificacao:
                    st.error("Configure a senha administrativa nos Secrets do Streamlit.")
                elif not hmac.compare_digest(senha_aprendizado, senha_admin_classificacao):
                    st.error("Senha administrativa incorreta.")
                elif not arquivos_aprendizado:
                    st.warning("Envie pelo menos uma planilha classificada ou arquivo ZIP.")
                else:
                    try:
                        novos_registros = executar_com_loading(
                            "Lendo os padrões das planilhas...",
                            importar_arquivos_classificados,
                            arquivos_aprendizado,
                            empresa_base_nova
                        )
                        quantidade_salva = executar_com_loading(
                            "Atualizando a base inteligente...",
                            salvar_classificacoes_online,
                            novos_registros,
                            empresa_base_nova
                        )
                        st.success(
                            f"{nome_base_nova} atualizada com {quantidade_salva} padrões de classificação."
                        )
                        st.rerun()
                    except Exception as erro_importacao:
                        st.error(f"Não foi possível atualizar a base: {erro_importacao}")

            st.markdown("---")
            st.markdown("#### Classificar planilha final conciliada")
            st.caption(
                "Anexe somente a planilha final, depois que a conferência bancária estiver "
                "concluída e os saldos estiverem batendo. O arquivo original não será alterado."
            )
            planilha_final_classificacao = st.file_uploader(
                "Planilha final com os saldos conferidos",
                type=['xlsx'],
                key=f'org_planilha_final_classificacao_nova_{chave_estabelecimento}'
            )
            if planilha_final_classificacao:
                if erro_base_classificacoes:
                    st.error("A base online precisa estar conectada antes da classificação.")
                elif not base_classificacoes:
                    st.warning(
                        "A base ainda não possui padrões. Importe primeiro as planilhas dos "
                        "meses já classificados."
                    )
                else:
                    try:
                        arquivo_classificado, resumo_classificacao = executar_com_loading(
                            "Analisando históricos e classificando as contas...",
                            classificar_planilha_final,
                            planilha_final_classificacao.getvalue(),
                            planilha_final_classificacao.name,
                            base_classificacoes,
                            contas_dominio_estabelecimento
                        )
                        coluna_classificados, coluna_532 = st.columns(2)
                        with coluna_classificados:
                            st.metric(
                                "Classificados automaticamente",
                                f"{int(resumo_classificacao['automaticos']):,}".replace(',', '.')
                            )
                        with coluna_532:
                            st.metric(
                                "Antecipados — conta 532",
                                f"{int(resumo_classificacao['antecipados']):,}".replace(',', '.')
                            )
                        renderizar_revisao_inteligente(
                            arquivo_classificado,
                            planilha_final_classificacao.getvalue(),
                            planilha_final_classificacao.name,
                            empresa_base_nova,
                            contas_dominio_estabelecimento,
                            senha_admin_classificacao,
                            'org_revisao_nova_' + chave_estabelecimento
                        )
                    except Exception as erro_classificacao_final:
                        st.error(
                            "Não foi possível classificar a planilha final: "
                            f"{erro_classificacao_final}"
                        )

        with aba_operacoes:
            st.markdown("---")
            st.markdown("### Organizar planilha bancária")

            banco_padrao = next(iter(configuracoes_bancos))
            bancos_empresa = st.multiselect(
                "Bancos",
                list(configuracoes_bancos.keys()),
                default=[banco_padrao],
                key=f"org_banco_nova_geracao_{chave_estabelecimento}"
            )
            if not bancos_empresa:
                st.info("Selecione pelo menos um banco para organizar a planilha.")

            configs_selecionadas = [configuracoes_bancos[banco] for banco in bancos_empresa]
            nomes_bancos = ", ".join(config['nome'] for config in configs_selecionadas)
            st.caption(
                f"O sistema localizará automaticamente as contas de {nomes_bancos} dentro da "
                "planilha consolidada pelas colunas CONTA, DATA, VALOR, LACTO, HISTORICO e DOC."
            )
            arquivo_empresa = st.file_uploader(
                f"Envie a planilha bancária da 266 - Nova Geração — {nome_estabelecimento_nova}",
                type=["xlsx", "xls"],
                key=f"org_upload_nova_geracao_multibanco_{chave_estabelecimento}"
            )

            if arquivo_empresa and configs_selecionadas:
                try:
                    bytes_empresa = arquivo_empresa.getvalue()
                    dados_processados = []
                    for config in configs_selecionadas:
                        df_banco, df_banco_retirados = executar_com_loading(
                            f"Organizando os lançamentos do {config['nome']}...",
                            config['processador'],
                            bytes_empresa
                        )
                        dados_processados.append((config, df_banco, df_banco_retirados))

                    datas_disponiveis = pd.concat(
                        [dados[1][['DATA']] for dados in dados_processados if not dados[1].empty],
                        ignore_index=True
                    )
                    datas_disponiveis['DATA'] = pd.to_datetime(
                        datas_disponiveis['DATA'], errors='coerce'
                    )
                    datas_disponiveis = datas_disponiveis.dropna(subset=['DATA'])
                    if datas_disponiveis.empty:
                        raise ValueError("Nenhuma data válida foi encontrada nos bancos selecionados.")

                    data_minima = datas_disponiveis['DATA'].min().date()
                    data_maxima = datas_disponiveis['DATA'].max().date()
                    chave_periodo = (
                        f"org_periodo_nova_{chave_estabelecimento}_"
                        f"{data_minima.isoformat()}_{data_maxima.isoformat()}_"
                        + "_".join(config['slug'] for config in configs_selecionadas)
                    )
                    st.markdown("### Período dos lançamentos")
                    periodo_selecionado = st.date_input(
                        "Selecione a data inicial e a data final",
                        value=(data_minima, data_maxima),
                        min_value=data_minima,
                        max_value=data_maxima,
                        format="DD/MM/YYYY",
                        key=chave_periodo
                    )
                    if not isinstance(periodo_selecionado, (tuple, list)) or len(periodo_selecionado) != 2:
                        raise ValueError(
                            "Selecione também a data final para concluir o período."
                        )
                    data_inicial, data_final = periodo_selecionado
                    if data_inicial > data_final:
                        raise ValueError(
                            "A data inicial não pode ser maior que a data final."
                        )

                    st.caption(
                        f"Serão considerados os lançamentos de {data_inicial.strftime('%d/%m/%Y')} "
                        f"até {data_final.strftime('%d/%m/%Y')}."
                    )

                    modelos_por_banco, retirados_por_banco = [], []
                    dados_exportacao_por_banco = {}
                    for config, df_banco_completo, df_banco_retirados_completo in dados_processados:
                        df_banco = filtrar_dataframe_periodo(
                            df_banco_completo, data_inicial, data_final
                        )
                        df_banco_retirados = filtrar_dataframe_periodo(
                            df_banco_retirados_completo, data_inicial, data_final
                        )
                        modelos_por_banco.append(df_banco)
                        retirados_por_banco.append(df_banco_retirados)
                        dados_exportacao_por_banco[config['nome']] = {
                            'principal': df_banco.sort_values('DATA', kind='stable').reset_index(drop=True),
                            'retirados': df_banco_retirados.sort_values('DATA', kind='stable').reset_index(drop=True)
                            if not df_banco_retirados.empty else df_banco_retirados
                        }

                    df_org = pd.concat(modelos_por_banco, ignore_index=True)
                    if df_org.empty:
                        raise ValueError(
                            "Nenhum lançamento foi encontrado no período selecionado."
                        )
                    df_org = df_org.sort_values(
                        ['DATA', 'DESCRIÇÃO'], kind='stable'
                    ).reset_index(drop=True)
                    df_retirados = pd.concat(retirados_por_banco, ignore_index=True)
                    if not df_retirados.empty:
                        df_retirados = df_retirados.sort_values(
                            ['DATA', 'DESCRIÇÃO'], kind='stable'
                        ).reset_index(drop=True)
                    modelo_org_bytes = None
                    for caminho_modelo in ['Modelo dominio.xlsx', 'Modelo dominio(6).xlsx']:
                        if os.path.exists(caminho_modelo):
                            with open(caminho_modelo, 'rb') as arquivo_modelo:
                                modelo_org_bytes = arquivo_modelo.read()
                            break
                    arquivo_final = executar_com_loading(
                        "Gerando a planilha final...",
                        gerar_excel_nova_geracao,
                        dados_exportacao_por_banco,
                        modelo_org_bytes
                    )

                    total_entradas = df_org.loc[df_org['VALOR'] > 0, 'VALOR'].sum()
                    total_saidas = df_org.loc[df_org['VALOR'] < 0, 'VALOR'].sum()
                    saldo_liquido = total_entradas + total_saidas

                    st.markdown("<br>", unsafe_allow_html=True)
                    m1, m2, m3, m4 = st.columns(4)
                    with m1: st.markdown(f'<div class="metric-card"><div class="metric-title">Modelo principal</div><div class="metric-value">{len(df_org)}</div></div>', unsafe_allow_html=True)
                    with m2: st.markdown(f'<div class="metric-card"><div class="metric-title">Retirados</div><div class="metric-value">{len(df_retirados)}</div></div>', unsafe_allow_html=True)
                    with m3: st.markdown(f'<div class="metric-card"><div class="metric-title">Entradas</div><div class="metric-value" style="color: #3fb950;">{formatar_moeda(total_entradas)}</div></div>', unsafe_allow_html=True)
                    with m4:
                        cor_saldo = "#3fb950" if saldo_liquido >= 0 else "#f85149"
                        st.markdown(f'<div class="metric-card"><div class="metric-title">Saldo líquido</div><div class="metric-value" style="color: {cor_saldo};">{formatar_moeda(saldo_liquido)}</div></div>', unsafe_allow_html=True)

                    tab_principal, tab_retirados = st.tabs(["Modelo principal", "Lançamentos retirados"])
                    with tab_principal:
                        previa = df_org.copy()
                        previa['DATA'] = pd.to_datetime(previa['DATA']).dt.strftime('%d/%m/%Y')
                        st.dataframe(formatar_dataframe_moeda_br(previa, ['VALOR']), use_container_width=True, height=320)
                    with tab_retirados:
                        if df_retirados.empty:
                            st.info("Nenhum estorno de baixa foi identificado neste arquivo.")
                        else:
                            previa_ret = df_retirados.copy()
                            previa_ret['DATA'] = pd.to_datetime(previa_ret['DATA']).dt.strftime('%d/%m/%Y')
                            st.dataframe(formatar_dataframe_moeda_br(previa_ret, ['VALOR']), use_container_width=True, height=280)

                    nome_saida_banco = (
                        configs_selecionadas[0]['nome']
                        if len(configs_selecionadas) == 1
                        else f"Separado_{len(configs_selecionadas)}_Bancos"
                    )
                    st.download_button(
                        "Baixar Modelo Domínio com abas por banco (.XLSX)",
                        data=arquivo_final,
                        file_name=(
                            f"Nova_Geracao_{nome_estabelecimento_nova}_{nome_saida_banco}_"
                            f"{data_inicial.strftime('%d%m%Y')}_a_{data_final.strftime('%d%m%Y')}_"
                            "Modelo_Dominio.xlsx"
                        ),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_org_nova_multibanco_{chave_estabelecimento}",
                        use_container_width=True
                    )

                    st.markdown("---")
                except Exception as e:
                    st.error(f"Não foi possível organizar a planilha: {e}")


            st.markdown("---")
            st.markdown("### Conferência com o extrato bancário")
            st.caption(
                "Esta área funciona de forma independente. Envie a planilha final organizada "
                "e os extratos bancários que deseja comparar."
            )

            todas_configs_conferencia = list(configuracoes_bancos.values())
            nomes_disponiveis_conferencia = [
                config['nome'] for config in todas_configs_conferencia
            ]
            chave_grupo_conferencia = "_".join(
                config['slug'] for config in todas_configs_conferencia
            )
            conferir_todos_bancos = st.checkbox(
                "Conferir todos os bancos disponíveis",
                value=False,
                key=(
                    f"org_conferir_todos_indep_nova_{chave_estabelecimento}_"
                    f"{chave_grupo_conferencia}"
                ),
                disabled=len(nomes_disponiveis_conferencia) == 1
            )
            if conferir_todos_bancos:
                bancos_conferencia = nomes_disponiveis_conferencia
                st.caption("Cada banco terá seu próprio relatório de conferência.")
            else:
                bancos_conferencia = st.multiselect(
                    "Bancos que serão conferidos",
                    nomes_disponiveis_conferencia,
                    default=nomes_disponiveis_conferencia[:1],
                    key=(
                        f"org_bancos_conferencia_indep_nova_{chave_estabelecimento}_"
                        f"{chave_grupo_conferencia}"
                    ),
                    help="Selecione um ou vários bancos para conferir."
                )

            if not bancos_conferencia:
                st.info("Selecione pelo menos um banco para realizar a conferência.")
            else:
                configs_conferencia = [
                    config for config in todas_configs_conferencia
                    if config['nome'] in bancos_conferencia
                ]

                col_planilha_conf, col_extratos_conf = st.columns(2)
                with col_planilha_conf:
                    planilha_atualizada_conferencia = st.file_uploader(
                        "Planilha final organizada",
                        type=["xlsx", "xls"],
                        key=(
                            "org_planilha_conferencia_indep_nova_"
                            + chave_estabelecimento
                        ),
                        help=(
                            "Pode conter um ou vários bancos. Somente os bancos "
                            "selecionados acima serão utilizados."
                        )
                    )
                with col_extratos_conf:
                    extratos_conferencia = st.file_uploader(
                        "Extrato(s) bancário(s)",
                        type=["pdf", "ofx", "csv", "xlsx", "xls"],
                        accept_multiple_files=True,
                        key=(
                            f"org_extratos_conferencia_indep_nova_"
                            f"{chave_estabelecimento}_"
                            + "_".join(config['slug'] for config in configs_conferencia)
                        ),
                        help="Envie os extratos correspondentes ao mesmo período da planilha."
                    )

                if not planilha_atualizada_conferencia:
                    st.info(
                        "Envie a planilha final organizada para identificar automaticamente "
                        "o período e liberar a comparação."
                    )
                else:
                    try:
                        dados_brutos_conferencia = {}
                        bancos_detectados_geral = set()
                        datas_planilha_conferencia = []

                        for config in configs_conferencia:
                            df_atualizada, df_retirados_atualizada, bancos_detectados = (
                                ler_planilha_organizada_conferencia(
                                    planilha_atualizada_conferencia.getvalue(),
                                    config['slug']
                                )
                            )
                            bancos_detectados_geral.update(bancos_detectados)
                            dados_brutos_conferencia[config['slug']] = {
                                'modelo': df_atualizada,
                                'retirados': df_retirados_atualizada
                            }
                            if not df_atualizada.empty and 'DATA' in df_atualizada.columns:
                                datas_validas_banco = pd.to_datetime(
                                    df_atualizada['DATA'], dayfirst=True, errors='coerce'
                                ).dropna()
                                datas_planilha_conferencia.extend(
                                    datas_validas_banco.dt.date.tolist()
                                )

                        if not datas_planilha_conferencia:
                            st.warning(
                                "A planilha final não possui datas válidas nos bancos selecionados."
                            )
                        else:
                            data_minima_conferencia = min(datas_planilha_conferencia)
                            data_maxima_conferencia = max(datas_planilha_conferencia)
                            periodo_conferencia = st.date_input(
                                "Período da conferência",
                                value=(
                                    data_minima_conferencia,
                                    data_maxima_conferencia
                                ),
                                min_value=data_minima_conferencia,
                                max_value=data_maxima_conferencia,
                                format="DD/MM/YYYY",
                                key=(
                                    f"org_periodo_conferencia_indep_nova_"
                                    f"{chave_estabelecimento}_{chave_grupo_conferencia}"
                                )
                            )

                            if (
                                not isinstance(periodo_conferencia, (tuple, list))
                                or len(periodo_conferencia) != 2
                            ):
                                st.info(
                                    "Selecione também a data final para concluir o período."
                                )
                            else:
                                data_inicial_conferencia, data_final_conferencia = (
                                    periodo_conferencia
                                )
                                dados_conferencia_por_banco = {}
                                bancos_sem_dados = []

                                for config in configs_conferencia:
                                    chave = config['slug']
                                    dados_conferencia_por_banco[chave] = {
                                        'modelo': filtrar_dataframe_periodo(
                                            dados_brutos_conferencia[chave]['modelo'],
                                            data_inicial_conferencia,
                                            data_final_conferencia
                                        ),
                                        'retirados': filtrar_dataframe_periodo(
                                            dados_brutos_conferencia[chave]['retirados'],
                                            data_inicial_conferencia,
                                            data_final_conferencia
                                        )
                                    }
                                    if dados_conferencia_por_banco[chave]['modelo'].empty:
                                        bancos_sem_dados.append(config['nome'])

                                bancos_texto = (
                                    ", ".join(sorted(bancos_detectados_geral))
                                    if bancos_detectados_geral else "não identificados"
                                )
                                st.success(
                                    f"Planilha carregada. Bancos identificados: {bancos_texto}. "
                                    f"Período: {data_inicial_conferencia.strftime('%d/%m/%Y')} "
                                    f"até {data_final_conferencia.strftime('%d/%m/%Y')}."
                                )
                                if bancos_sem_dados:
                                    st.warning(
                                        "Sem lançamentos no período para: "
                                        + ", ".join(bancos_sem_dados)
                                    )

                                if not extratos_conferencia:
                                    st.info(
                                        "Agora envie pelo menos um extrato bancário para "
                                        "gerar os relatórios."
                                    )
                                else:
                                    extratos_por_banco = {
                                        config['slug']: [] for config in configs_conferencia
                                    }
                                    arquivos_nao_identificados = []

                                    for extrato_conferencia in extratos_conferencia:
                                        lancamentos_arquivo = executar_com_loading(
                                            f"Lendo {extrato_conferencia.name}...",
                                            processar_extrato_conferencia_empresa,
                                            extrato_conferencia.getvalue(),
                                            extrato_conferencia.name
                                        )
                                        df_arquivo = filtrar_dataframe_periodo(
                                            pd.DataFrame(lancamentos_arquivo),
                                            data_inicial_conferencia,
                                            data_final_conferencia
                                        )
                                        if df_arquivo.empty:
                                            continue

                                        chave_pelo_nome = identificar_chave_banco_empresa(
                                            extrato_conferencia.name
                                        )
                                        if chave_pelo_nome in extratos_por_banco:
                                            extratos_por_banco[chave_pelo_nome].extend(
                                                df_arquivo.to_dict('records')
                                            )
                                            continue

                                        chaves_arquivo = df_arquivo['DESCRIÇÃO'].apply(
                                            identificar_chave_banco_empresa
                                        )
                                        chaves_reconhecidas = set(
                                            chave for chave in chaves_arquivo.unique().tolist()
                                            if chave
                                        )
                                        if not chaves_reconhecidas:
                                            if len(configs_conferencia) == 1:
                                                chave_unica = configs_conferencia[0]['slug']
                                                extratos_por_banco[chave_unica].extend(
                                                    df_arquivo.to_dict('records')
                                                )
                                            else:
                                                arquivos_nao_identificados.append(
                                                    extrato_conferencia.name
                                                )
                                            continue

                                        for config in configs_conferencia:
                                            chave = config['slug']
                                            df_banco_extrato = df_arquivo[
                                                chaves_arquivo.eq(chave)
                                            ]
                                            if not df_banco_extrato.empty:
                                                extratos_por_banco[chave].extend(
                                                    df_banco_extrato.to_dict('records')
                                                )

                                    if arquivos_nao_identificados:
                                        st.warning(
                                            "Não foi possível identificar o banco destes arquivos: "
                                            + ", ".join(arquivos_nao_identificados)
                                        )

                                    if not any(extratos_por_banco.values()):
                                        st.warning(
                                            "Nenhum lançamento dos extratos foi identificado "
                                            "dentro do período selecionado."
                                        )
                                    else:
                                        abas_bancos = st.tabs([
                                            config['nome']
                                            for config in configs_conferencia
                                        ])
                                        for aba_banco, config in zip(
                                            abas_bancos, configs_conferencia
                                        ):
                                            with aba_banco:
                                                chave = config['slug']
                                                nome_banco = config['nome']
                                                df_modelo_banco = (
                                                    dados_conferencia_por_banco[chave]['modelo']
                                                )
                                                df_retirados_banco = (
                                                    dados_conferencia_por_banco[chave]['retirados']
                                                )
                                                df_extrato_banco = pd.DataFrame(
                                                    extratos_por_banco[chave]
                                                )

                                                st.markdown(
                                                    f"#### Relatório — {nome_banco}"
                                                )
                                                if df_modelo_banco.empty:
                                                    st.warning(
                                                        f"Não há lançamentos do {nome_banco} "
                                                        "na planilha para o período."
                                                    )
                                                    continue
                                                if df_extrato_banco.empty:
                                                    st.warning(
                                                        f"Nenhum extrato do {nome_banco} "
                                                        "foi identificado para o período."
                                                    )
                                                    continue

                                                diario, _, _, _ = executar_com_loading(
                                                    f"Conferindo os movimentos do {nome_banco}...",
                                                    conciliar_empresa_com_extrato,
                                                    df_modelo_banco,
                                                    df_extrato_banco,
                                                    df_retirados_banco
                                                )
                                                if diario.empty:
                                                    st.warning(
                                                        "Não existem datas válidas para realizar "
                                                        "a conferência."
                                                    )
                                                    continue

                                                periodo_inicial = (
                                                    diario['DATA'].min().strftime('%d/%m/%Y')
                                                )
                                                periodo_final = (
                                                    diario['DATA'].max().strftime('%d/%m/%Y')
                                                )
                                                dias_batendo = int(
                                                    (diario['STATUS'] == '✅ Batendo').sum()
                                                )
                                                dias_divergentes = int(
                                                    (diario['STATUS'] == '❌ Divergente').sum()
                                                )
                                                st.info(
                                                    f"Período analisado: {periodo_inicial} "
                                                    f"até {periodo_final}"
                                                )

                                                c1, c2 = st.columns(2)
                                                with c1:
                                                    st.markdown(
                                                        '<div class="metric-card">'
                                                        '<div class="metric-title">Dias batendo</div>'
                                                        f'<div class="metric-value" style="color: #3fb950;">'
                                                        f'{dias_batendo}</div></div>',
                                                        unsafe_allow_html=True
                                                    )
                                                with c2:
                                                    st.markdown(
                                                        '<div class="metric-card">'
                                                        '<div class="metric-title">Dias divergentes</div>'
                                                        f'<div class="metric-value" style="color: #f85149;">'
                                                        f'{dias_divergentes}</div></div>',
                                                        unsafe_allow_html=True
                                                    )

                                                if dias_divergentes == 0:
                                                    st.success(
                                                        "Conferência concluída: todos os dias "
                                                        "estão batendo."
                                                    )
                                                else:
                                                    st.warning(
                                                        "Foram encontradas diferenças nos "
                                                        "totais diários."
                                                    )

                                                exibicao_diaria = diario[[
                                                    'DATA', 'ENTRADAS PLANILHA', 'ENTRADAS EXTRATO', 'DIF. ENTRADAS',
                                                    'SAÍDAS PLANILHA', 'SAÍDAS EXTRATO', 'DIF. SAÍDAS', 'STATUS'
                                                ]].copy()
                                                exibicao_diaria['DATA'] = exibicao_diaria['DATA'].dt.strftime('%d/%m/%Y')
                                                exibicao_diaria.columns = [
                                                    'Data', 'Entrada Planilha', 'Entrada Extrato', 'Diferença Entradas',
                                                    'Saída Planilha', 'Saída Extrato', 'Diferença Saídas', 'Status'
                                                ]
                                                exibicao_diaria = formatar_dataframe_moeda_br(
                                                    exibicao_diaria,
                                                    ['Entrada Planilha', 'Entrada Extrato', 'Diferença Entradas', 'Saída Planilha', 'Saída Extrato', 'Diferença Saídas']
                                                )
                                                st.dataframe(
                                                    exibicao_diaria,
                                                    use_container_width=True,
                                                    height=390,
                                                    hide_index=True
                                                )
                    except Exception as erro_conferencia_independente:
                        st.error(
                            "Não foi possível realizar a conferência: "
                            f"{erro_conferencia_independente}"
                        )

    # --- Ferramenta exclusiva 1529: Nibo -> Modelo Dominio ---
    if st.session_state['empresa_organizador'] == 'dias_pereira':
        contas_dias_pereira = {'itau': '508', 'banco_brasil': '8'}
        bancos_dias_pereira = {
            'Itaú · Conta contábil 508': {
                'slug': 'itau', 'descricao': 'BANCO ITAÚ', 'arquivo': 'Itau', 'aba': 'Itaú'
            },
            'Banco do Brasil · Conta contábil 8': {
                'slug': 'banco_brasil', 'descricao': 'BANCO DO BRASIL', 'arquivo': 'Banco_do_Brasil', 'aba': 'Banco do Brasil'
            },
        }

        aba_nibo, aba_base_dias = st.tabs([
            'Organizar arquivos',
            'Base Inteligente'
        ])

        with aba_base_dias:
            renderizar_base_inteligente_empresa(
                'dias_pereira',
                '1529 - Dias e Pereira',
                {'itau', 'banco_brasil'},
                contas_dias_pereira
            )

        with aba_nibo:
            st.markdown('### Nibo → Modelo Domínio')
            st.caption(
                'Selecione o banco e envie o PDF de Contas & Extratos exportado pelo Nibo. '
                'O Razync organiza os movimentos e aplica a Base Inteligente exclusiva da 1529.'
            )

            bancos_nibo_selecionados = st.multiselect(
                'Bancos deste processamento Nibo',
                list(bancos_dias_pereira.keys()),
                default=[list(bancos_dias_pereira.keys())[0]],
                key='dias_pereira_bancos_nibo',
                help='Você pode selecionar Itaú, Banco do Brasil ou os dois bancos ao mesmo tempo.'
            )

            arquivos_nibo_por_banco = {}
            if not bancos_nibo_selecionados:
                st.info('Selecione pelo menos um banco para continuar.')
            else:
                st.caption(
                    'Envie um PDF para cada banco selecionado. Quando os dois forem enviados, '
                    'o Razync gera um único Modelo Domínio consolidado.'
                )
                for banco_nibo_rotulo in bancos_nibo_selecionados:
                    config_banco_nibo = bancos_dias_pereira[banco_nibo_rotulo]
                    arquivo_nibo = st.file_uploader(
                        f"Extrato Nibo em PDF — {banco_nibo_rotulo}",
                        type=['pdf'],
                        key=f"dias_pereira_extrato_nibo_{config_banco_nibo['slug']}",
                        help='Use o relatório mensal de Contas & Extratos do Nibo.'
                    )
                    if arquivo_nibo is not None:
                        arquivos_nibo_por_banco[banco_nibo_rotulo] = arquivo_nibo

            todos_arquivos_nibo_enviados = (
                bool(bancos_nibo_selecionados)
                and len(arquivos_nibo_por_banco) == len(bancos_nibo_selecionados)
            )

            if todos_arquivos_nibo_enviados:
                try:
                    quadros_nibo = []
                    quadros_nibo_por_slug = {}
                    configs_nibo_processados = []
                    with st.spinner('Lendo e organizando os relatórios Nibo...'):
                        for banco_nibo_rotulo in bancos_nibo_selecionados:
                            config_banco_nibo = bancos_dias_pereira[banco_nibo_rotulo]
                            arquivo_nibo = arquivos_nibo_por_banco[banco_nibo_rotulo]
                            df_nibo = processar_extrato_nibo_pdf(arquivo_nibo.getvalue())
                            df_banco_nibo = df_nibo[
                                ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
                            ].copy()
                            df_banco_nibo['DESCRIÇÃO'] = config_banco_nibo['descricao']
                            quadros_nibo.append(df_banco_nibo)
                            quadros_nibo_por_slug[config_banco_nibo['slug']] = df_banco_nibo
                            configs_nibo_processados.append(config_banco_nibo)

                    df_export_nibo = pd.concat(quadros_nibo, ignore_index=True)
                    df_export_nibo['_DATA_ORDEM'] = pd.to_datetime(
                        df_export_nibo['DATA'], dayfirst=True, errors='coerce'
                    )
                    df_export_nibo = (
                        df_export_nibo
                        .sort_values(['_DATA_ORDEM', 'DESCRIÇÃO'], kind='stable')
                        .drop(columns=['_DATA_ORDEM'])
                        .reset_index(drop=True)
                    )
                    datas_nibo = pd.to_datetime(
                        df_export_nibo['DATA'], dayfirst=True, errors='coerce'
                    )
                    entradas_nibo = df_export_nibo.loc[
                        df_export_nibo['VALOR'] > 0, 'VALOR'
                    ].sum()
                    saidas_nibo = abs(df_export_nibo.loc[
                        df_export_nibo['VALOR'] < 0, 'VALOR'
                    ].sum())

                    col_nibo_1, col_nibo_2, col_nibo_3 = st.columns(3)
                    col_nibo_1.metric(
                        'Lançamentos', f'{len(df_export_nibo):,}'.replace(',', '.')
                    )
                    col_nibo_2.metric(
                        'Entradas',
                        f'R$ {entradas_nibo:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                    )
                    col_nibo_3.metric(
                        'Saídas',
                        f'R$ {saidas_nibo:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                    )

                    st.success('Relatório(s) Nibo organizado(s) com sucesso.')
                    st.dataframe(
                        df_export_nibo[['DESCRIÇÃO', 'DATA', 'VALOR', 'HISTÓRICO']],
                        use_container_width=True,
                        hide_index=True
                    )

                    datas_validas = datas_nibo.dropna()
                    bancos_nome_arquivo = '_'.join(
                        config['arquivo'] for config in configs_nibo_processados
                    )
                    if not datas_validas.empty:
                        nome_nibo = (
                            f"1529_Dias_Pereira_{bancos_nome_arquivo}_"
                            f"{datas_validas.min().strftime('%m_%Y')}.xlsx"
                        )
                    else:
                        nome_nibo = (
                            f"1529_Dias_Pereira_{bancos_nome_arquivo}_Modelo_Dominio.xlsx"
                        )

                    def _bytes_excel_nibo(arquivo_excel):
                        if isinstance(arquivo_excel, (bytes, bytearray)):
                            return bytes(arquivo_excel)
                        if hasattr(arquivo_excel, 'getvalue'):
                            return arquivo_excel.getvalue()
                        return bytes(arquivo_excel)

                    def _combinar_modelos_nibo_em_abas(modelos_por_banco):
                        from copy import copy as copiar_estilo_nibo
                        from openpyxl import Workbook, load_workbook

                        wb_saida = Workbook()
                        wb_saida.remove(wb_saida.active)

                        for nome_aba, arquivo_banco in modelos_por_banco:
                            wb_origem = load_workbook(io.BytesIO(_bytes_excel_nibo(arquivo_banco)))
                            ws_origem = wb_origem.active
                            ws_destino = wb_saida.create_sheet(title=nome_aba[:31])

                            for row in ws_origem.iter_rows():
                                for celula in row:
                                    nova = ws_destino[celula.coordinate]
                                    nova.value = celula.value
                                    if celula.has_style:
                                        nova._style = copiar_estilo_nibo(celula._style)
                                    if celula.number_format:
                                        nova.number_format = celula.number_format
                                    if celula.font:
                                        nova.font = copiar_estilo_nibo(celula.font)
                                    if celula.fill:
                                        nova.fill = copiar_estilo_nibo(celula.fill)
                                    if celula.border:
                                        nova.border = copiar_estilo_nibo(celula.border)
                                    if celula.alignment:
                                        nova.alignment = copiar_estilo_nibo(celula.alignment)
                                    if celula.protection:
                                        nova.protection = copiar_estilo_nibo(celula.protection)

                            for chave, dimensao in ws_origem.column_dimensions.items():
                                ws_destino.column_dimensions[chave].width = dimensao.width
                                ws_destino.column_dimensions[chave].hidden = dimensao.hidden
                                ws_destino.column_dimensions[chave].bestFit = dimensao.bestFit

                            for indice, dimensao in ws_origem.row_dimensions.items():
                                ws_destino.row_dimensions[indice].height = dimensao.height
                                ws_destino.row_dimensions[indice].hidden = dimensao.hidden

                            for intervalo in ws_origem.merged_cells.ranges:
                                ws_destino.merge_cells(str(intervalo))

                            ws_destino.freeze_panes = ws_origem.freeze_panes
                            ws_destino.sheet_format = copiar_estilo_nibo(ws_origem.sheet_format)
                            ws_destino.sheet_properties = copiar_estilo_nibo(ws_origem.sheet_properties)
                            ws_destino.page_margins = copiar_estilo_nibo(ws_origem.page_margins)
                            ws_destino.page_setup = copiar_estilo_nibo(ws_origem.page_setup)
                            ws_destino.print_options = copiar_estilo_nibo(ws_origem.print_options)
                            ws_destino.sheet_view.showGridLines = ws_origem.sheet_view.showGridLines
                            if ws_origem.auto_filter.ref:
                                ws_destino.auto_filter.ref = ws_origem.auto_filter.ref

                        saida = io.BytesIO()
                        wb_saida.save(saida)
                        saida.seek(0)
                        return saida.getvalue()

                    modelos_nibo_por_banco = []
                    resumo_nibo = {'automaticos': 0, 'somente_banco': 0}
                    base_dias_pereira = []
                    erro_base_nibo = ''
                    try:
                        base_dias_pereira = carregar_classificacoes_online('dias_pereira')
                    except Exception as erro_base:
                        erro_base_nibo = str(erro_base)

                    for config_banco_nibo in configs_nibo_processados:
                        df_banco_nibo = quadros_nibo_por_slug[config_banco_nibo['slug']]
                        excel_banco_nibo = gerar_excel_modelo_dominio(df_banco_nibo)
                        arquivo_banco_nibo = excel_banco_nibo
                        resumo_banco_nibo = {}

                        if not erro_base_nibo:
                            try:
                                base_banco_nibo = [
                                    item for item in base_dias_pereira
                                    if item.get('banco') == config_banco_nibo['slug']
                                ]
                                arquivo_banco_nibo, resumo_banco_nibo = classificar_planilha_final(
                                    excel_banco_nibo,
                                    nome_nibo,
                                    base_banco_nibo,
                                    contas_dias_pereira
                                )
                            except Exception as erro_classificacao_banco:
                                st.info(
                                    f"A aba {config_banco_nibo['aba']} foi gerada normalmente, "
                                    'mas a Base Inteligente não pôde ser aplicada nela agora: '
                                    f'{erro_classificacao_banco}'
                                )

                        resumo_nibo['automaticos'] += int(
                            resumo_banco_nibo.get('automaticos', 0) or 0
                        )
                        resumo_nibo['somente_banco'] += int(
                            resumo_banco_nibo.get('somente_banco', 0) or 0
                        )
                        modelos_nibo_por_banco.append(
                            (config_banco_nibo['aba'], arquivo_banco_nibo)
                        )

                    if erro_base_nibo:
                        st.info(
                            'O Modelo Domínio foi gerado normalmente em abas separadas por banco, '
                            'mas a Base Inteligente não pôde ser carregada agora: '
                            f'{erro_base_nibo}'
                        )

                    arquivo_saida_nibo = _combinar_modelos_nibo_em_abas(
                        modelos_nibo_por_banco
                    )

                    if any(resumo_nibo.values()):
                        c_auto_1, c_auto_2 = st.columns(2)
                        c_auto_1.metric(
                            'Classificados automaticamente',
                            f"{int(resumo_nibo.get('automaticos', 0)):,}".replace(',', '.')
                        )
                        c_auto_2.metric(
                            'Pendentes de contrapartida',
                            f"{int(resumo_nibo.get('somente_banco', 0)):,}".replace(',', '.')
                        )

                    st.download_button(
                        'Baixar Modelo Domínio por banco',
                        data=arquivo_saida_nibo,
                        file_name=nome_nibo,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True,
                        key='dias_pereira_download_modelo_dominio'
                    )
                except Exception as erro_nibo:
                    st.error(f'Não foi possível processar o(s) relatório(s) Nibo: {erro_nibo}')

# ==============================================================================
# TELA 4: CONCILIAÇÃO COM O RAZÃO DA DOMÍNIO
# ==============================================================================
elif st.session_state['pagina_ativa'] == 'razao':
    if st.button("← Início", key="btn_voltar_home_razao", type="tertiary"):
        mudar_pagina('home')
        st.rerun()
    st.markdown(
        """
        <header class="rz-page-header">
            <div class="rz-page-kicker">Conferência contábil</div>
            <div class="rz-page-title">Conciliação com Razão</div>
            <div class="rz-page-description">
                Compare o extrato bancário com o Razão da Domínio e identifique
                diferenças diárias com a natureza contábil correta.
            </div>
        </header>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("""<div class="aviso-banner"><p><strong>Formatos aceitos:</strong> CSV, XLSX e XLS antigo da Domínio. Quando necessário, o sistema recupera e normaliza o arquivo automaticamente antes da leitura.</p></div>""", unsafe_allow_html=True)

    st.markdown("##### 📁 Arquivos de Importação")
    col_up1, col_up2 = st.columns(2)
    with col_up1: arq_extrato = st.file_uploader("1º - Envie o Extrato (PDF, OFX, Excel, CSV)", type=["pdf", "ofx", "csv", "xlsx", "xls"], key="up_extrato")
    with col_up2: arq_razao = st.file_uploader("2º - Envie o Razão exportado (CSV, XLSX ou XLS)", type=["csv", "xlsx", "xls"], key="up_razao")

    if arq_extrato and arq_razao:
        try:
            ext_bytes, ext_ext = arq_extrato.getvalue(), os.path.splitext(arq_extrato.name)[1].lower()
            lancamentos_ext = executar_com_loading(
                "Analisando o extrato bancário...",
                processar_extrato_unificado,
                ext_bytes,
                arq_extrato.name
            )
                
            raz_bytes, raz_name = arq_razao.getvalue(), arq_razao.name
            
            for chave_estado_xls in ['erro_bof_xls', 'razao_xls_recuperado']:
                st.session_state.pop(chave_estado_xls, None)

            df_razao_bruto = executar_com_loading(
                "Lendo e preparando o Razão da Domínio...",
                processar_razao_dominio,
                raz_bytes,
                raz_name
            )

            if st.session_state.get('razao_xls_recuperado', False):
                st.success(
                    "Arquivo XLS antigo recuperado e convertido automaticamente "
                    "em memória. A conciliação pode continuar normalmente."
                )

            if st.session_state.get('erro_bof_xls', False):
                st.error(
                    "Não foi possível recuperar este arquivo XLS. Como alternativa, "
                    "salve-o como CSV ou XLSX e envie novamente."
                )
                st.stop()

            if lancamentos_ext and df_razao_bruto is not None and not df_razao_bruto.empty:
                # ---------------- PREPARAÇÃO DOS DADOS ----------------
                df_ext = pd.DataFrame(lancamentos_ext)
                df_ext['DATA_DT'] = pd.to_datetime(df_ext['DATA'], dayfirst=True, errors='coerce')
                df_ext = df_ext.dropna(subset=['DATA_DT'])
                
                df_ext['ENTRADAS_EXTRATO'] = df_ext['VALOR'].apply(lambda x: x if x > 0 else 0.0)
                df_ext['SAIDAS_EXTRATO'] = df_ext['VALOR'].apply(lambda x: abs(x) if x < 0 else 0.0)
                df_ext_agregado = df_ext.groupby('DATA_DT')[['ENTRADAS_EXTRATO', 'SAIDAS_EXTRATO']].sum().reset_index()
                
                df_razao_bruto['DATA_DT'] = pd.to_datetime(df_razao_bruto['DATA'], dayfirst=True, errors='coerce')
                df_razao_bruto = df_razao_bruto.dropna(subset=['DATA_DT'])
                df_razao_agregado = df_razao_bruto.groupby('DATA_DT')[['ENTRADAS_RAZAO', 'SAIDAS_RAZAO']].sum().reset_index()

                for col in ['ENTRADAS_EXTRATO', 'SAIDAS_EXTRATO']:
                    if col not in df_ext_agregado.columns: df_ext_agregado[col] = 0.0
                for col in ['ENTRADAS_RAZAO', 'SAIDAS_RAZAO']:
                    if col not in df_razao_agregado.columns: df_razao_agregado[col] = 0.0

                df_conciliacao = pd.merge(df_ext_agregado, df_razao_agregado, on='DATA_DT', how='outer').fillna(0.0)
                df_conciliacao = df_conciliacao.sort_values('DATA_DT')
                df_conciliacao['DATA_EXIBICAO'] = df_conciliacao['DATA_DT'].dt.strftime('%d/%m/%Y')

                if df_conciliacao.empty:
                    st.warning("⚠️ Não conseguimos cruzar as datas. Verifique se os arquivos contêm datas válidas.")
                    st.stop()

                # ---------------- FILTRO DE PERÍODO ----------------
                st.markdown("---")
                st.markdown("##### 📅 Filtro de Período da Conciliação")
                
                dt_min_geral, dt_max_geral = df_conciliacao['DATA_DT'].min().date(), df_conciliacao['DATA_DT'].max().date()
                col_p1, col_p2 = st.columns(2)
                with col_p1: data_ini_filtro = st.date_input("Data Inicial", value=dt_min_geral, min_value=dt_min_geral, max_value=dt_max_geral, format="DD/MM/YYYY", key="raz_ini")
                with col_p2: data_fim_filtro = st.date_input("Data Final", value=dt_max_geral, min_value=dt_min_geral, max_value=dt_max_geral, format="DD/MM/YYYY", key="raz_fim")
                
                if data_ini_filtro > data_fim_filtro:
                    st.warning("⚠️ A data inicial não pode ser maior que a data final.")
                    data_ini_filtro, data_fim_filtro = dt_min_geral, dt_max_geral

                df_conciliacao = df_conciliacao[(df_conciliacao['DATA_DT'].dt.date >= data_ini_filtro) & (df_conciliacao['DATA_DT'].dt.date <= data_fim_filtro)].copy()

                if df_conciliacao.empty:
                    st.info("Nenhuma movimentação no período selecionado.")
                    st.stop()

                # ---------------- CÁLCULOS DE DIFERENÇAS ----------------
                df_conciliacao = df_conciliacao.sort_values('DATA_DT')
                # Natureza espelhada entre banco e contabilidade:
                # SAÍDA no extrato deve bater com ENTRADA/DÉBITO no Razão.
                # ENTRADA no extrato deve bater com SAÍDA/CRÉDITO no Razão.
                df_conciliacao['DIF_SAIDAS_EXT_ENTRADAS_RAZAO'] = (
                    df_conciliacao['ENTRADAS_RAZAO'] - df_conciliacao['SAIDAS_EXTRATO']
                )
                df_conciliacao['DIF_ENTRADAS_EXT_SAIDAS_RAZAO'] = (
                    df_conciliacao['SAIDAS_RAZAO'] - df_conciliacao['ENTRADAS_EXTRATO']
                )
                
                df_conciliacao['STATUS'] = df_conciliacao.apply(
                    lambda row: "✅ Batendo" if (
                        abs(row['DIF_SAIDAS_EXT_ENTRADAS_RAZAO']) < 0.01
                        and abs(row['DIF_ENTRADAS_EXT_SAIDAS_RAZAO']) < 0.01
                    ) else "❌ Divergente",
                    axis=1
                )
                
                # ---------------- 4 CARDS RESUMO ----------------
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("### 📊 Resultado da Conferência Diária")
                
                tot_ent_ext = df_conciliacao['ENTRADAS_EXTRATO'].sum()
                tot_sai_ext = df_conciliacao['SAIDAS_EXTRATO'].sum()
                tot_ent_raz = df_conciliacao['ENTRADAS_RAZAO'].sum()
                tot_sai_raz = df_conciliacao['SAIDAS_RAZAO'].sum()
                
                rc1, rc2, rc3, rc4 = st.columns(4)
                with rc1: st.markdown(f'<div class="metric-card"><div class="metric-title">Saídas do Extrato</div><div class="metric-value" style="color: #f85149;">{formatar_moeda(abs(tot_sai_ext))}</div></div>', unsafe_allow_html=True)
                with rc2: st.markdown(f'<div class="metric-card"><div class="metric-title">Entradas/Débitos do Razão</div><div class="metric-value" style="color: #f85149;">{formatar_moeda(tot_ent_raz)}</div></div>', unsafe_allow_html=True)
                with rc3: st.markdown(f'<div class="metric-card"><div class="metric-title">Entradas do Extrato</div><div class="metric-value" style="color: #3fb950;">{formatar_moeda(tot_ent_ext)}</div></div>', unsafe_allow_html=True)
                with rc4: st.markdown(f'<div class="metric-card"><div class="metric-title">Saídas/Créditos do Razão</div><div class="metric-value" style="color: #3fb950;">{formatar_moeda(tot_sai_raz)}</div></div>', unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                
                # ---------------- TABELA DE EXIBIÇÃO ----------------
                df_exibicao = df_conciliacao[[
                    'DATA_EXIBICAO',
                    'SAIDAS_EXTRATO', 'ENTRADAS_RAZAO', 'DIF_SAIDAS_EXT_ENTRADAS_RAZAO',
                    'ENTRADAS_EXTRATO', 'SAIDAS_RAZAO', 'DIF_ENTRADAS_EXT_SAIDAS_RAZAO',
                    'STATUS'
                ]].copy()
                df_exibicao.columns = [
                    'Data',
                    'Saídas Extrato (R$)', 'Entradas/Débitos Razão (R$)', 'Dif. Saída Ext. x Entrada Razão (R$)',
                    'Entradas Extrato (R$)', 'Saídas/Créditos Razão (R$)', 'Dif. Entrada Ext. x Saída Razão (R$)',
                    'Status'
                ]
                colunas_monetarias_conciliacao = [
                    'Saídas Extrato (R$)', 'Entradas/Débitos Razão (R$)',
                    'Dif. Saída Ext. x Entrada Razão (R$)',
                    'Entradas Extrato (R$)', 'Saídas/Créditos Razão (R$)',
                    'Dif. Entrada Ext. x Saída Razão (R$)'
                ]
                
                st.dataframe(
                    formatar_dataframe_moeda_br(df_exibicao, colunas_monetarias_conciliacao),
                    use_container_width=True, 
                    height=380
                )

                # ---------------- EXPORTAÇÃO EXCEL BLINDADA ----------------
                st.markdown("---")
                st.markdown("##### 📥 Exportar Relatório de Conciliação")
                st.caption("Faça o download da conferência completa em formato Excel.")
                
                buf_audit = io.BytesIO()
                with pd.ExcelWriter(buf_audit, engine='openpyxl') as writer:
                    df_exib_excel = df_exibicao.copy()
                    for col in colunas_monetarias_conciliacao:
                        df_exib_excel[col] = df_exib_excel[col].apply(formatar_moeda)
                        
                    sanitizar_dataframe(df_exib_excel).to_excel(writer, sheet_name="Resumo Geral", index=False)
                    
                    df_divergencias = df_conciliacao[df_conciliacao['STATUS'] == '❌ Divergente'].copy()
                    if not df_divergencias.empty:
                        df_div_export = df_divergencias[[
                            'DATA_EXIBICAO',
                            'SAIDAS_EXTRATO', 'ENTRADAS_RAZAO', 'DIF_SAIDAS_EXT_ENTRADAS_RAZAO',
                            'ENTRADAS_EXTRATO', 'SAIDAS_RAZAO', 'DIF_ENTRADAS_EXT_SAIDAS_RAZAO'
                        ]].copy()
                        df_div_export.columns = [
                            'Data',
                            'Saidas Extrato', 'Entradas Debitos Razao', 'Diferenca Saida Ext x Entrada Razao',
                            'Entradas Extrato', 'Saidas Creditos Razao', 'Diferenca Entrada Ext x Saida Razao'
                        ]
                        for col in df_div_export.columns[1:]: df_div_export[col] = df_div_export[col].apply(formatar_moeda)
                        sanitizar_dataframe(df_div_export).to_excel(writer, sheet_name="Dias Divergentes", index=False)

                st.download_button(
                    label="Baixar Relatório em Excel (.XLSX)", 
                    data=buf_audit.getvalue(), 
                    file_name=f"Analise_Conciliacao_{data_ini_filtro.strftime('%d%m%Y')}_a_{data_fim_filtro.strftime('%d%m%Y')}.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    use_container_width=False
                )

            else:
                st.warning("⚠️ Não conseguimos extrair as linhas contábeis válidas. Verifique se os arquivos contêm Data e Valor.")
        
        except Exception as e:
            st.error("Não foi possível concluir o cruzamento dos dados. Verifique os arquivos enviados e tente novamente.")
