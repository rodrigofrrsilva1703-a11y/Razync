import io
import zipfile

import pandas as pd
from openpyxl import Workbook, load_workbook

import razync.eletro_forte_francesinhas as francesinhas


TEXTO_508 = """
Nome do cliente Agência/Conta Carteira Emitido em Folha
ELETRO FORTE C E LTDA ME 0020/10531-8 SIMPLES 04/08/26 1
157 000477870 00147.991 SIKA SA 0757 02/08/26 5.663,40 L 03/08 5.663,40
157 000477871 00147.992 EMPRESA IGNORADA 0757 02/08/26 100,00 E 03/08 0,00
"""

TEXTO_509 = """
Nome do cliente Agência/Conta Carteira Emitido em Folha
ELETRO FORTE C E LTDA ME 3392/18153-7 CONTRATUAL 06/08/26 1
157 000564495 00148.518A SCA SERVICOS LTDA 0691 05/08/26 853,60 L 05/08 05 450,00 403,60
"""


def _zip_teste():
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as arquivo:
        arquivo.writestr("FRANC 508.pdf", b"pdf-508")
        arquivo.writestr("FRANC 509.pdf", b"pdf-509")
    return buffer.getvalue()


def test_zip_filtra_l_usa_emitido_em_e_valor_creditado(monkeypatch):
    monkeypatch.setattr(
        francesinhas,
        "_texto_pdf",
        lambda conteudo: TEXTO_508 if conteudo == b"pdf-508" else TEXTO_509,
    )
    dados, avisos = francesinhas.processar_zip_francesinhas(_zip_teste())
    assert avisos == []
    assert list(dados["DÉBITO"]) == ["508", "509"]
    assert list(dados["VALOR"]) == [5663.40, 403.60]
    assert list(dados["HISTÓRICO"]) == ["SIKA SA", "SCA SERVICOS LTDA"]
    assert list(pd.to_datetime(dados["DATA"]).dt.strftime("%d/%m/%Y")) == [
        "04/08/2026", "06/08/2026"
    ]


def test_gera_excel_unico_com_abas_por_conta(monkeypatch):
    monkeypatch.setattr(
        francesinhas,
        "_texto_pdf",
        lambda conteudo: TEXTO_508 if conteudo == b"pdf-508" else TEXTO_509,
    )
    dados, _ = francesinhas.processar_zip_francesinhas(_zip_teste())
    wb_modelo = Workbook()
    ws = wb_modelo.active
    ws.append(francesinhas.COLUNAS_MODELO)
    ws.append(["", None, 0.0, "", "", ""])
    modelo = io.BytesIO()
    wb_modelo.save(modelo)

    saida = francesinhas.gerar_excel_francesinhas(modelo.getvalue(), dados)
    wb = load_workbook(io.BytesIO(saida), data_only=False)
    assert wb.sheetnames == ["Francesinhas - Itau 508", "Francesinhas - Itau 509"]
    assert wb["Francesinhas - Itau 508"]["D2"].value == 508
    assert wb["Francesinhas - Itau 509"]["C2"].value == 403.60


def test_zip_rejeita_caminho_inseguro():
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as arquivo:
        arquivo.writestr("../indevido.pdf", b"pdf")
    try:
        francesinhas.processar_zip_francesinhas(buffer.getvalue())
    except ValueError as erro:
        assert "inseguro" in str(erro)
    else:
        raise AssertionError("ZIP inseguro deveria ser rejeitado")

