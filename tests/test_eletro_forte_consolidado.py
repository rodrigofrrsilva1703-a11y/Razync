import io

import pandas as pd
from openpyxl import Workbook, load_workbook

from razync.eletro_forte import (
    COLUNAS_MODELO,
    corrigir_datas_com_francesinhas,
    gerar_consolidado_bancos_eletro_forte,
)


def _linha(data, valor, debito, credito, historico):
    return pd.DataFrame([{
        "DESCRIÇÃO": "BANCO",
        "DATA": pd.Timestamp(data),
        "VALOR": valor,
        "DÉBITO": debito,
        "CRÉDITO": credito,
        "HISTÓRICO": historico,
    }], columns=COLUNAS_MODELO)


def _modelo():
    wb = Workbook()
    ws = wb.active
    ws.append(COLUNAS_MODELO)
    ws.append(["", None, 0.0, "", "", ""])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_consolida_tres_origens_em_abas_por_banco():
    despesas = {
        "8": _linha("2026-08-03", -100, "10", "8", "Pago: despesa BB"),
        "508": _linha("2026-08-04", -200, "20", "508", "Pago: despesa Itaú"),
    }
    fornecedores = {
        "8": _linha("2026-08-05", -300, "30", "8", "Pago: fornecedor BB"),
        "509": _linha("2026-08-06", -400, "40", "509", "Pago: fornecedor Itaú"),
    }
    recebidos = {
        "8": _linha("2026-08-07", 500, "8", "50", "Recebido: cliente BB"),
        "508": _linha("2026-08-08", 600, "508", "60", "Recebido: cliente Itaú"),
    }

    saida = gerar_consolidado_bancos_eletro_forte(
        _modelo(), despesas, fornecedores, recebidos
    )
    wb = load_workbook(io.BytesIO(saida), data_only=False)

    assert wb.sheetnames == [
        "Banco do Brasil - 8", "Itau - 508", "Itau - 509"
    ]
    assert wb["Banco do Brasil - 8"].max_row == 4
    assert wb["Itau - 508"].max_row == 3
    assert wb["Itau - 509"].max_row == 2
    assert wb["Banco do Brasil - 8"]["C2"].value == -100
    assert wb["Banco do Brasil - 8"]["C4"].value == 500


def test_consolidado_mantem_conta_zero_separada():
    revisar = {"0": _linha("2026-08-10", -75, "166", "0", "Pago: revisar")}
    saida = gerar_consolidado_bancos_eletro_forte(
        _modelo(), {}, revisar, {}
    )
    wb = load_workbook(io.BytesIO(saida), data_only=False)
    assert wb.sheetnames == ["Revisar - 0"]


def test_francesinha_corrige_data_sem_criar_ou_duplicar_lancamento():
    recebidos = {
        "508": pd.concat([
            _linha("2026-08-03", 5663.40, "508", "10", "Recebido: 001 - SIKA SA"),
            _linha("2026-08-03", 100.00, "508", "20", "Recebido: OUTRO CLIENTE"),
        ], ignore_index=True)
    }
    francesinhas = pd.DataFrame([{
        "DATA": pd.Timestamp("2026-08-04"),
        "VALOR": 5663.40,
        "DÉBITO": "508",
        "HISTÓRICO": "Recebido: SIKA SA",
    }])

    corrigidos, resumo, pendencias = corrigir_datas_com_francesinhas(
        recebidos, francesinhas
    )

    assert len(corrigidos["508"]) == 2
    assert pd.Timestamp(corrigidos["508"].iloc[0]["DATA"]) == pd.Timestamp("2026-08-04")
    assert pd.Timestamp(corrigidos["508"].iloc[1]["DATA"]) == pd.Timestamp("2026-08-03")
    assert resumo == {"corrigidos": 1, "nao_encontrados": 0, "ambiguos": 0}
    assert pendencias.empty


def test_francesinha_sem_correspondencia_e_preservada_com_aviso():
    recebidos = {
        "509": _linha(
            "2026-08-09", 500.00, "509", "30", "Recebido: CLIENTE EXISTENTE"
        )
    }
    francesinhas = pd.DataFrame([{
        "DATA": pd.Timestamp("2026-08-10"),
        "VALOR": 704.26,
        "DÉBITO": "509",
        "HISTÓRICO": "Recebido: CLIENTE AUSENTE",
    }])

    corrigidos, resumo, pendencias = corrigir_datas_com_francesinhas(
        recebidos, francesinhas
    )

    assert pd.Timestamp(corrigidos["509"].iloc[0]["DATA"]) == pd.Timestamp("2026-08-09")
    assert resumo["corrigidos"] == 0
    assert resumo["nao_encontrados"] == 1
    assert len(pendencias) == 1
