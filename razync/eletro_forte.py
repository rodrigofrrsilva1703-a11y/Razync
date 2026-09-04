"""Processamento específico da empresa 242 - Eletro Forte.

Os relatórios recebidos usam principalmente extensão .xls em formato HTML exportado
pelo sistema do cliente. Cada arquivo final preserva a aba principal do relatório e
acrescenta abas montadas sobre o Modelo Domínio padrão do Razync.
"""
from __future__ import annotations

import html
import io
import re
import unicodedata
from copy import copy
from difflib import SequenceMatcher
from typing import Dict

import pandas as pd


CONTAS_ELETRO_FORTE = {
    "8": "Banco BB · Conta 8",
    "508": "Itaú · 105318 · Conta 508",
    "509": "Itaú · 181537 · Conta 509",
    "0": "Revisar · Conta 0",
}

DESCRICOES_BANCOS = {
    "8": "BANCO DO BRASIL",
    "508": "BANCO ITAÚ",
    "509": "BANCO ITAÚ",
    "0": "REVISAR CONTA 0",
}

COLUNAS_MODELO = ["DESCRIÇÃO", "DATA", "VALOR", "DÉBITO", "CRÉDITO", "HISTÓRICO"]


def _texto_celula(valor: str) -> str:
    valor = re.sub(r"<[^>]+>", " ", valor or "")
    return " ".join(html.unescape(valor).replace("\xa0", " ").split()).strip()


def _ler_tabela_html(conteudo: bytes) -> pd.DataFrame:
    texto = conteudo.decode("cp1252", errors="replace")
    linhas = re.findall(r"<tr[^>]*>(.*?)</tr>", texto, flags=re.I | re.S)
    matriz = []
    for linha in linhas:
        celulas = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", linha, flags=re.I | re.S)
        if celulas:
            matriz.append([_texto_celula(celula) for celula in celulas])
    if len(matriz) < 2:
        raise ValueError("Não foi possível localizar a tabela de lançamentos no arquivo.")
    cabecalho = [str(c).strip() for c in matriz[0]]
    largura = len(cabecalho)
    dados = [(linha + [""] * largura)[:largura] for linha in matriz[1:]]
    return pd.DataFrame(dados, columns=cabecalho)


def ler_aba_principal(conteudo: bytes, nome_arquivo: str = "") -> pd.DataFrame:
    """Lê a primeira aba/tabela exatamente como recebida para preservá-la no arquivo final."""
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".xlsx"):
        try:
            return pd.read_excel(io.BytesIO(conteudo), sheet_name=0, dtype=object)
        except Exception:
            pass
    if nome.endswith(".xls"):
        try:
            return _ler_tabela_html(conteudo)
        except Exception:
            try:
                return pd.read_excel(io.BytesIO(conteudo), sheet_name=0, dtype=object, engine="xlrd")
            except Exception:
                pass
    try:
        return _ler_tabela_html(conteudo)
    except Exception as erro:
        raise ValueError("Não foi possível preservar a aba principal do relatório.") from erro


def _norm_coluna(valor: str) -> str:
    texto = str(valor).upper()
    substituicoes = {
        "É": "E", "Ê": "E", "Í": "I", "Ó": "O", "Ô": "O",
        "Á": "A", "Ã": "A", "À": "A", "Ú": "U", "Ç": "C",
    }
    for origem, destino in substituicoes.items():
        texto = texto.replace(origem, destino)
    return re.sub(r"[^A-Z0-9]", "", texto)


def _valor_br(valor) -> float:
    texto = str(valor or "").strip().replace("R$", "").replace(" ", "")
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _conta(valor) -> str:
    texto = str(valor or "").strip()
    texto = re.sub(r"\.0$", "", texto)
    texto = re.sub(r"\D", "", texto)
    return texto.lstrip("0") or "0"


def _data(valor, ano_referencia: int) -> pd.Timestamp:
    texto = str(valor or "").strip()
    if re.fullmatch(r"\d{1,2}/\d{1,2}", texto):
        texto = f"{texto}/{ano_referencia}"
    return pd.to_datetime(texto, dayfirst=True, errors="coerce")


def _padronizar(conteudo: bytes, ano_referencia: int) -> pd.DataFrame:
    bruto = _ler_tabela_html(conteudo)
    mapa = {_norm_coluna(col): col for col in bruto.columns}
    col_data = mapa.get("DATA") or mapa.get("DATAPAGTO")
    col_debito = mapa.get("DEBITO")
    col_credito = mapa.get("CREDITO")
    col_valor = mapa.get("VALOR")
    col_historico = mapa.get("HISTORICO")
    faltantes = [nome for nome, col in {
        "DATA": col_data, "DÉBITO": col_debito, "CRÉDITO": col_credito,
        "VALOR": col_valor, "HISTÓRICO": col_historico,
    }.items() if not col]
    if faltantes:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltantes))

    saida = pd.DataFrame({
        "DATA": bruto[col_data].map(lambda v: _data(v, ano_referencia)),
        "DÉBITO": bruto[col_debito].map(_conta),
        "CRÉDITO": bruto[col_credito].map(_conta),
        "VALOR": bruto[col_valor].map(_valor_br),
        "HISTÓRICO": bruto[col_historico].astype(str).str.strip(),
    })
    return saida.dropna(subset=["DATA"]).reset_index(drop=True)


def _aplicar_regra_movimento(df: pd.DataFrame, tipo: str) -> pd.DataFrame:
    """Aplica sinal e prefixo de histórico conforme a origem do relatório."""
    saida = df.copy()
    historico = saida["HISTÓRICO"].fillna("").astype(str).str.strip()
    if tipo in {"despesa", "fornecedor"}:
        saida["VALOR"] = -pd.to_numeric(saida["VALOR"], errors="coerce").fillna(0.0).abs()
        saida["HISTÓRICO"] = historico.map(
            lambda texto: texto if texto.lower().startswith("pago: ") else f"Pago: {texto}"
        )
    elif tipo == "recebido":
        saida["VALOR"] = pd.to_numeric(saida["VALOR"], errors="coerce").fillna(0.0).abs()
        saida["HISTÓRICO"] = historico.map(
            lambda texto: texto if texto.lower().startswith("recebido: ") else f"Recebido: {texto}"
        )
    return saida


def _com_descricao(df: pd.DataFrame, coluna_banco: str) -> pd.DataFrame:
    saida = df.copy()
    saida.insert(
        0,
        "DESCRIÇÃO",
        saida[coluna_banco].astype(str).map(DESCRICOES_BANCOS).fillna("MOVIMENTO BANCÁRIO"),
    )
    return saida[COLUNAS_MODELO]


def _separar_por_conta(df: pd.DataFrame, coluna_conta: str) -> Dict[str, pd.DataFrame]:
    resultado: Dict[str, pd.DataFrame] = {}
    ordem = ["8", "508", "509", "0"]
    contas_presentes = list(dict.fromkeys(df[coluna_conta].astype(str).tolist()))
    for conta in ordem + [c for c in contas_presentes if c not in ordem]:
        parte = df.loc[df[coluna_conta].astype(str) == conta].copy()
        if not parte.empty:
            resultado[conta] = _com_descricao(parte.reset_index(drop=True), coluna_conta)
    return resultado


def processar_despesas(conteudo: bytes, ano_referencia: int) -> Dict[str, pd.DataFrame]:
    """Despesa: converte 001→8 e 002→508, força valor negativo e prefixa Pago:."""
    df = _padronizar(conteudo, ano_referencia)
    df["CRÉDITO"] = df["CRÉDITO"].replace({"1": "8", "2": "508"})
    df = _aplicar_regra_movimento(df, "despesa")
    return _separar_por_conta(df, "CRÉDITO")


def processar_fornecedores(conteudo: bytes, ano_referencia: int) -> Dict[str, pd.DataFrame]:
    """Fornecedor: separa por CRÉDITO, força valor negativo e prefixa Pago:."""
    df = _aplicar_regra_movimento(_padronizar(conteudo, ano_referencia), "fornecedor")
    return _separar_por_conta(df, "CRÉDITO")


def processar_recebidos(conteudo: bytes, ano_referencia: int) -> Dict[str, pd.DataFrame]:
    """Recebido: separa por DÉBITO, mantém valor positivo e prefixa Recebido:."""
    df = _aplicar_regra_movimento(_padronizar(conteudo, ano_referencia), "recebido")
    return _separar_por_conta(df, "DÉBITO")


def _nome_para_correspondencia(valor: str) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = texto.encode("ascii", "ignore").decode().upper()
    texto = re.sub(r"^RECEBIDO:\s*", "", texto)
    texto = re.sub(r"^\d+\s*-\s*", "", texto)
    return re.sub(r"[^A-Z0-9]", "", texto)


def corrigir_datas_com_francesinhas(
    recebidos: Dict[str, pd.DataFrame] | None,
    francesinhas: pd.DataFrame,
    limite_dias: int = 7,
) -> tuple[Dict[str, pd.DataFrame], dict, pd.DataFrame]:
    """Corrige somente a data de recebimentos conciliados com liquidações L."""
    corrigidos = {
        str(conta): df.copy().reset_index(drop=True)
        for conta, df in (recebidos or {}).items()
    }
    resumo = {"corrigidos": 0, "nao_encontrados": 0, "ambiguos": 0}
    pendencias = []
    usados = set()

    for _, francesinha in francesinhas.sort_values("DATA", kind="stable").iterrows():
        conta = str(francesinha.get("DÉBITO", "")).strip()
        grupo = corrigidos.get(conta)
        data_correta = pd.to_datetime(francesinha.get("DATA"), errors="coerce")
        valor_correto = round(float(francesinha.get("VALOR", 0) or 0), 2)
        nome_francesinha = _nome_para_correspondencia(
            francesinha.get("HISTÓRICO", "")
        )
        candidatos = []

        if grupo is not None and not grupo.empty and not pd.isna(data_correta):
            for indice, linha in grupo.iterrows():
                identificador = (conta, int(indice))
                if identificador in usados:
                    continue
                valor_linha = round(float(linha.get("VALOR", 0) or 0), 2)
                if valor_linha != valor_correto:
                    continue
                data_linha = pd.to_datetime(linha.get("DATA"), errors="coerce")
                if pd.isna(data_linha):
                    continue
                diferenca_dias = (data_correta.normalize() - data_linha.normalize()).days
                if not 0 <= diferenca_dias <= limite_dias:
                    continue
                similaridade = SequenceMatcher(
                    None,
                    nome_francesinha,
                    _nome_para_correspondencia(linha.get("HISTÓRICO", "")),
                ).ratio()
                candidatos.append(
                    (similaridade, -diferenca_dias, int(indice), data_linha)
                )

        candidatos.sort(reverse=True)
        motivo = ""
        if not candidatos or candidatos[0][0] < 0.35:
            resumo["nao_encontrados"] += 1
            motivo = "Correspondência não encontrada"
        else:
            melhor = candidatos[0]
            empate = (
                len(candidatos) > 1
                and candidatos[1][0] == melhor[0]
                and candidatos[1][1] == melhor[1]
            )
            if empate:
                resumo["ambiguos"] += 1
                motivo = "Mais de uma correspondência possível"
            else:
                indice = melhor[2]
                grupo.at[indice, "DATA"] = data_correta
                usados.add((conta, indice))
                resumo["corrigidos"] += 1

        if motivo:
            pendencias.append({
                "CONTA": conta,
                "DATA FRANCESINHA": data_correta,
                "VALOR": valor_correto,
                "PAGADOR": francesinha.get("HISTÓRICO", ""),
                "MOTIVO": motivo,
            })

    return corrigidos, resumo, pd.DataFrame(pendencias)


def inferir_ano_recebidos(conteudo: bytes) -> int | None:
    try:
        bruto = _ler_tabela_html(conteudo)
    except Exception:
        return None
    for valor in bruto.astype(str).to_numpy().ravel():
        achou = re.search(r"\b\d{1,2}/\d{1,2}/(20\d{2})\b", valor)
        if achou:
            return int(achou.group(1))
    return None


def _nome_aba(prefixo: str, conta: str) -> str:
    nomes = {
        "8": "BB 8",
        "508": "Itau 508",
        "509": "Itau 509",
        "0": "Revisar 0",
    }
    return f"{prefixo} - {nomes.get(conta, conta)}"[:31]


def _localizar_cabecalho_modelo(ws):
    esperadas = {_norm_coluna(c): c for c in COLUNAS_MODELO}
    for linha in range(1, min(ws.max_row, 25) + 1):
        mapa = {}
        for coluna in range(1, ws.max_column + 1):
            chave = _norm_coluna(ws.cell(linha, coluna).value or "")
            if chave:
                mapa[chave] = coluna
        if all(chave in mapa for chave in esperadas):
            return linha, mapa
    raise ValueError("O cabeçalho padrão do Modelo Domínio não foi localizado.")


def _preencher_aba_modelo(ws, df: pd.DataFrame):
    cabecalho_linha, mapa_colunas = _localizar_cabecalho_modelo(ws)
    linha_modelo = cabecalho_linha + 1
    estilos = {}
    for coluna in range(1, ws.max_column + 1):
        celula = ws.cell(linha_modelo, coluna)
        estilos[coluna] = {
            "font": copy(celula.font),
            "fill": copy(celula.fill),
            "border": copy(celula.border),
            "alignment": copy(celula.alignment),
            "number_format": celula.number_format,
            "protection": copy(celula.protection),
        }

    for linha in range(cabecalho_linha + 1, ws.max_row + 1):
        for coluna in range(1, ws.max_column + 1):
            ws.cell(linha, coluna).value = None

    for indice, registro in enumerate(df[COLUNAS_MODELO].to_dict("records"), start=linha_modelo):
        for nome in COLUNAS_MODELO:
            coluna_excel = mapa_colunas[_norm_coluna(nome)]
            valor = registro.get(nome, "")
            if pd.isna(valor):
                valor = ""
            if nome == "DATA" and valor not in ("", None):
                data = pd.to_datetime(valor, dayfirst=True, errors="coerce")
                valor = data.to_pydatetime() if not pd.isna(data) else valor
            elif nome in {"DÉBITO", "CRÉDITO"} and str(valor).isdigit():
                valor = int(valor)
            elif nome == "VALOR":
                valor = float(valor)

            celula = ws.cell(indice, coluna_excel)
            celula.value = valor
            estilo = estilos.get(coluna_excel)
            if estilo:
                celula.font = copy(estilo["font"])
                celula.fill = copy(estilo["fill"])
                celula.border = copy(estilo["border"])
                celula.alignment = copy(estilo["alignment"])
                celula.number_format = estilo["number_format"]
                celula.protection = copy(estilo["protection"])


def _preencher_aba_principal(ws, df_principal: pd.DataFrame):
    ws.title = "Principal"
    for col_idx, nome in enumerate(df_principal.columns, start=1):
        ws.cell(1, col_idx).value = str(nome)
    for row_idx, linha in enumerate(df_principal.itertuples(index=False, name=None), start=2):
        for col_idx, valor in enumerate(linha, start=1):
            if pd.isna(valor):
                valor = ""
            ws.cell(row_idx, col_idx).value = valor
    if df_principal.shape[1]:
        for col_idx, nome in enumerate(df_principal.columns, start=1):
            largura = max(12, min(48, max(len(str(nome)), 10) + 2))
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = largura
    ws.freeze_panes = "A2"


def gerar_modelo_dominio_eletro_forte(
    conteudo_original: bytes,
    nome_original: str,
    modelo_bytes: bytes,
    despesas: Dict[str, pd.DataFrame] | None,
    fornecedores: Dict[str, pd.DataFrame] | None,
    recebidos: Dict[str, pd.DataFrame] | None,
) -> bytes:
    """Preserva a aba principal e cria abas bancárias copiadas do Modelo Domínio real."""
    from openpyxl import load_workbook

    if not modelo_bytes:
        raise FileNotFoundError("Modelo Domínio não encontrado no sistema.")

    principal = ler_aba_principal(conteudo_original, nome_original)
    wb = load_workbook(io.BytesIO(modelo_bytes))
    template = wb[wb.sheetnames[0]]

    ws_principal = wb.create_sheet("Principal", 0)
    _preencher_aba_principal(ws_principal, principal)

    conjuntos = []
    for conta, df in (despesas or {}).items():
        conjuntos.append((_nome_aba("Despesa", conta), df))
    for conta, df in (fornecedores or {}).items():
        conjuntos.append((_nome_aba("Fornecedor", conta), df))
    for conta, df in (recebidos or {}).items():
        conjuntos.append((_nome_aba("Recebido", conta), df))

    if not conjuntos:
        raise ValueError("Nenhum lançamento válido foi encontrado nos arquivos enviados.")

    for nome, df in conjuntos:
        ws = wb.copy_worksheet(template)
        ws.title = nome[:31]
        _preencher_aba_modelo(ws, df)

    wb.remove(template)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_consolidado_bancos_eletro_forte(
    modelo_bytes: bytes,
    despesas: Dict[str, pd.DataFrame] | None,
    fornecedores: Dict[str, pd.DataFrame] | None,
    recebidos: Dict[str, pd.DataFrame] | None,
) -> bytes:
    """Reúne os três relatórios em um único Modelo Domínio, por conta bancária."""
    from openpyxl import load_workbook

    if not modelo_bytes:
        raise FileNotFoundError("Modelo Domínio não encontrado no sistema.")

    grupos: Dict[str, list[pd.DataFrame]] = {}
    for conjunto in (despesas or {}, fornecedores or {}, recebidos or {}):
        for conta, df in conjunto.items():
            if df is not None and not df.empty:
                grupos.setdefault(str(conta), []).append(df.copy())

    if not grupos:
        raise ValueError("Nenhum lançamento válido foi encontrado nos três relatórios.")

    wb = load_workbook(io.BytesIO(modelo_bytes))
    template = wb[wb.sheetnames[0]]
    nomes_abas = {
        "8": "Banco do Brasil - 8",
        "508": "Itau - 508",
        "509": "Itau - 509",
        "0": "Revisar - 0",
    }
    ordem_contas = ["8", "508", "509", "0"]
    restantes = sorted(conta for conta in grupos if conta not in ordem_contas)

    for conta in ordem_contas + restantes:
        partes = grupos.get(conta)
        if not partes:
            continue
        consolidado = pd.concat(partes, ignore_index=True)
        consolidado["DATA"] = pd.to_datetime(
            consolidado["DATA"], dayfirst=True, errors="coerce"
        )
        consolidado = consolidado.sort_values(
            ["DATA", "HISTÓRICO"], kind="stable", na_position="last"
        ).reset_index(drop=True)
        ws = wb.copy_worksheet(template)
        ws.title = nomes_abas.get(conta, f"Conta - {conta}")[:31]
        _preencher_aba_modelo(ws, consolidado)

    wb.remove(template)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
