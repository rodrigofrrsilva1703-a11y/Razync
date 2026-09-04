"""Conversão das francesinhas Itaú da empresa 242 para o Modelo Domínio."""
from __future__ import annotations

import hashlib
import io
import re
import zipfile
from copy import copy
from pathlib import PurePosixPath

import pandas as pd
from pypdf import PdfReader


COLUNAS_MODELO = ["DESCRIÇÃO", "DATA", "VALOR", "DÉBITO", "CRÉDITO", "HISTÓRICO"]
CONTAS_POR_CONTA_ITAU = {"10531-8": "508", "18153-7": "509"}
LIMITE_PDFS = 300
LIMITE_DESCOMPACTADO = 40 * 1024 * 1024


def _valor_br(texto: str) -> float:
    return float(str(texto).replace(".", "").replace(",", "."))


def _texto_pdf(conteudo: bytes) -> str:
    leitor = PdfReader(io.BytesIO(conteudo))
    return "\n".join(
        pagina.extract_text(extraction_mode="layout") or ""
        for pagina in leitor.pages
    )


def processar_francesinha_pdf(conteudo: bytes, nome_arquivo: str = "") -> pd.DataFrame:
    """Extrai apenas liquidações L e usa a data de emissão do relatório."""
    texto = _texto_pdf(conteudo)
    conta_encontrada = re.search(r"\b\d{4}/(10531-8|18153-7)\b", texto)
    if not conta_encontrada:
        raise ValueError(f"Conta Itaú 10531-8/18153-7 não encontrada em {nome_arquivo}.")
    conta_itau = conta_encontrada.group(1)
    conta_dominio = CONTAS_POR_CONTA_ITAU[conta_itau]

    emissao = re.search(
        rf"\b\d{{4}}/{re.escape(conta_itau)}\s+\S+\s+(\d{{2}}/\d{{2}}/\d{{2,4}})\b",
        texto,
    )
    if not emissao:
        raise ValueError(f"Data 'Emitido em' não encontrada em {nome_arquivo}.")
    data_emissao = pd.to_datetime(emissao.group(1), dayfirst=True, errors="raise")

    padrao_linha = re.compile(
        r"^\s*\d{3}\s+\S+\s+\S+\s+"
        r"(?P<pagador>.+?)\s+\d{4}\s+\d{2}/\d{2}/\d{2}\s+"
        r"(?P<valor>[\d.]+,\d{2})\s+L\s+\d{2}/\d{2}"
        r"(?:\s+\d{2}\s+[\d.]+,\d{2})?\s+"
        r"(?P<credito>[\d.]+,\d{2})\s*$"
    )
    registros = []
    for linha in texto.splitlines():
        encontrado = padrao_linha.match(linha)
        if not encontrado:
            continue
        valor_creditado = _valor_br(encontrado.group("credito"))
        if valor_creditado <= 0:
            continue
        registros.append({
            "DESCRIÇÃO": "BANCO ITAÚ",
            "DATA": data_emissao,
            "VALOR": valor_creditado,
            "DÉBITO": conta_dominio,
            "CRÉDITO": "",
            "HISTÓRICO": (
                "Recebido: " + " ".join(encontrado.group("pagador").split())
            ),
            "ARQUIVO": nome_arquivo,
            "CONTA_ITAU": conta_itau,
        })
    return pd.DataFrame(
        registros,
        columns=COLUNAS_MODELO + ["ARQUIVO", "CONTA_ITAU"],
    )


def processar_zip_francesinhas(conteudo_zip: bytes) -> tuple[pd.DataFrame, list[str]]:
    """Processa, de uma só vez, todos os PDFs válidos de um ZIP."""
    resultados = []
    avisos = []
    hashes = set()
    with zipfile.ZipFile(io.BytesIO(conteudo_zip)) as arquivo_zip:
        membros = [m for m in arquivo_zip.infolist() if not m.is_dir()]
        pdfs = []
        total = 0
        for membro in membros:
            caminho = PurePosixPath(membro.filename.replace("\\", "/"))
            if caminho.is_absolute() or ".." in caminho.parts:
                raise ValueError("O ZIP contém um caminho de arquivo inseguro.")
            if caminho.suffix.lower() != ".pdf":
                avisos.append(f"Ignorado (não é PDF): {caminho.name}")
                continue
            total += membro.file_size
            if total > LIMITE_DESCOMPACTADO:
                raise ValueError("O conteúdo descompactado excede o limite de 40 MB.")
            pdfs.append(membro)
        if not pdfs:
            raise ValueError("Nenhum PDF foi encontrado no ZIP.")
        if len(pdfs) > LIMITE_PDFS:
            raise ValueError(f"O ZIP pode conter no máximo {LIMITE_PDFS} PDFs.")

        for membro in pdfs:
            conteudo = arquivo_zip.read(membro)
            assinatura = hashlib.sha256(conteudo).hexdigest()
            if assinatura in hashes:
                avisos.append(f"Ignorado (PDF duplicado): {PurePosixPath(membro.filename).name}")
                continue
            hashes.add(assinatura)
            try:
                df = processar_francesinha_pdf(
                    conteudo, PurePosixPath(membro.filename).name
                )
                if df.empty:
                    avisos.append(
                        f"Sem liquidações L: {PurePosixPath(membro.filename).name}"
                    )
                else:
                    resultados.append(df)
            except Exception as erro:
                avisos.append(f"Não processado: {PurePosixPath(membro.filename).name} - {erro}")

    if not resultados:
        raise ValueError("Nenhuma liquidação com Hist. L foi encontrada nos PDFs.")
    consolidado = pd.concat(resultados, ignore_index=True)
    consolidado = consolidado.sort_values(
        ["DATA", "DÉBITO", "ARQUIVO", "HISTÓRICO"], kind="stable"
    ).reset_index(drop=True)
    return consolidado, avisos


def _normalizar_coluna(valor: str) -> str:
    texto = str(valor).upper()
    for origem, destino in {"Ç": "C", "Ã": "A", "É": "E", "Ó": "O"}.items():
        texto = texto.replace(origem, destino)
    return re.sub(r"[^A-Z0-9]", "", texto)


def _preencher_modelo(ws, dados: pd.DataFrame) -> None:
    cabecalho = None
    mapa = {}
    for linha in range(1, min(ws.max_row, 25) + 1):
        atual = {
            _normalizar_coluna(ws.cell(linha, coluna).value or ""): coluna
            for coluna in range(1, ws.max_column + 1)
        }
        if all(_normalizar_coluna(nome) in atual for nome in COLUNAS_MODELO):
            cabecalho, mapa = linha, atual
            break
    if cabecalho is None:
        raise ValueError("Cabeçalho do Modelo Domínio não encontrado.")

    linha_modelo = cabecalho + 1
    estilos = {}
    for coluna in range(1, ws.max_column + 1):
        celula = ws.cell(linha_modelo, coluna)
        estilos[coluna] = (
            copy(celula.font), copy(celula.fill), copy(celula.border),
            copy(celula.alignment), celula.number_format, copy(celula.protection),
        )
    for linha in range(cabecalho + 1, ws.max_row + 1):
        for coluna in range(1, ws.max_column + 1):
            ws.cell(linha, coluna).value = None

    for numero, registro in enumerate(
        dados[COLUNAS_MODELO].to_dict("records"), start=linha_modelo
    ):
        for nome in COLUNAS_MODELO:
            coluna = mapa[_normalizar_coluna(nome)]
            valor = registro[nome]
            if nome == "DATA":
                valor = pd.Timestamp(valor).to_pydatetime()
            elif nome == "VALOR":
                valor = float(valor)
            elif nome in {"DÉBITO", "CRÉDITO"} and str(valor).isdigit():
                valor = int(valor)
            celula = ws.cell(numero, coluna)
            celula.value = valor
            fonte, preenchimento, borda, alinhamento, formato, protecao = estilos[coluna]
            celula.font, celula.fill, celula.border = copy(fonte), copy(preenchimento), copy(borda)
            celula.alignment, celula.number_format = copy(alinhamento), formato
            celula.protection = copy(protecao)


def gerar_excel_francesinhas(modelo_bytes: bytes, dados: pd.DataFrame) -> bytes:
    """Gera um único Excel, com abas separadas para as contas 508 e 509."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(modelo_bytes))
    template = wb[wb.sheetnames[0]]
    abas_criadas = 0
    for conta, nome_aba in (("508", "Francesinhas - Itau 508"), ("509", "Francesinhas - Itau 509")):
        parte = dados.loc[dados["DÉBITO"].astype(str) == conta].copy()
        if parte.empty:
            continue
        ws = wb.copy_worksheet(template)
        ws.title = nome_aba
        _preencher_modelo(ws, parte)
        abas_criadas += 1
    if not abas_criadas:
        raise ValueError("Nenhum lançamento das contas 508 ou 509 foi encontrado.")
    wb.remove(template)
    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()

def corrigir_datas_com_francesinhas(recebidos, francesinhas, limite_dias=7):
    """Carrega a correção sob demanda para não bloquear a inicialização do app."""
    from razync.eletro_forte import (
        corrigir_datas_com_francesinhas as executar_correcao,
    )

    return executar_correcao(recebidos, francesinhas, limite_dias)
