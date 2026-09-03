"""Processamento específico da empresa 1211 - GZ Importadora e Exportadora.

Fonte principal: extrato Itaú (conta Domínio 508).
Fonte auxiliar: relatório 'Boletos baixados e liquidados'.
Os lançamentos agregados 'BOLETOS RECEBIDOS' são substituídos por boletos
individuais liquidados, com histórico 'Recebido: NOME DO PAGADOR'.
"""
from __future__ import annotations

import io
import re
from copy import copy
from dataclasses import dataclass
from typing import List

import pandas as pd
from pypdf import PdfReader


CONTA_ITAU_GZ = "508"
COLUNAS_MODELO = ["DESCRIÇÃO", "DATA", "VALOR", "DÉBITO", "CRÉDITO", "HISTÓRICO"]


@dataclass
class BoletoGZ:
    pagador: str
    vencimento: pd.Timestamp
    liquidacao: pd.Timestamp
    valor: float
    status: str
    usado: bool = False


def _normalizar_espacos(texto: str) -> str:
    return re.sub(r"\s+", " ", str(texto or "")).strip()


def _moeda_br(valor) -> float:
    texto = str(valor or "").replace("R$", "").replace("\xa0", " ").strip()
    texto = texto.replace(" ", "")
    if not texto:
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _texto_pdf(conteudo: bytes) -> str:
    reader = PdfReader(io.BytesIO(conteudo), strict=False)
    return "\n".join((pagina.extract_text() or "") for pagina in reader.pages)


def _ocr_pdf(conteudo: bytes) -> str:
    import fitz
    import pytesseract
    from PIL import Image

    doc = fitz.open(stream=conteudo, filetype="pdf")
    textos = []
    for pagina in doc:
        pix = pagina.get_pixmap(matrix=fitz.Matrix(2.2, 2.2), alpha=False)
        imagem = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        try:
            texto = pytesseract.image_to_string(imagem, lang="por")
        except Exception:
            texto = pytesseract.image_to_string(imagem)
        textos.append(texto)
    return "\n".join(textos)


def _limpar_historico_extrato(texto: str) -> str:
    texto = _normalizar_espacos(texto)
    texto = re.sub(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", " ", texto)
    texto = re.sub(r"\b\d{3}\.\d{3}\.\d{3}-\d{2}\b", " ", texto)
    return _normalizar_espacos(texto).strip(" -")


def ler_extrato_itau_gz(conteudo: bytes) -> pd.DataFrame:
    """Extrai movimentos do extrato Itaú e ignora linhas de saldo."""
    texto = _texto_pdf(conteudo)
    if "GZ IMPORTADORA" not in texto.upper() and "0099343-5" not in texto:
        raise ValueError("O PDF enviado não parece ser o extrato Itaú da GZ.")

    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    blocos = []
    atual = None
    padrao_data = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(.*)$")
    for linha in linhas:
        match = padrao_data.match(linha)
        if match:
            if atual:
                blocos.append(atual)
            atual = {"data": match.group(1), "partes": [match.group(2)]}
        elif atual:
            atual["partes"].append(linha)
    if atual:
        blocos.append(atual)

    registros = []
    regex_moeda = re.compile(r"(?<!\d)([-+]?\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)")
    for bloco in blocos:
        conteudo_bloco = _normalizar_espacos(" ".join(bloco["partes"]))
        norm = conteudo_bloco.upper()
        if norm.startswith("SALDO ") or "SALDO TOTAL DISPONÍVEL" in norm or "SALDO EM CONTA" in norm:
            continue
        moedas = list(regex_moeda.finditer(conteudo_bloco))
        if not moedas:
            continue
        moeda = moedas[-1]
        valor = _moeda_br(moeda.group(1))
        historico = _limpar_historico_extrato(conteudo_bloco[:moeda.start()])
        if not historico or abs(valor) < 0.005:
            continue
        data = pd.to_datetime(bloco["data"], dayfirst=True, errors="coerce")
        if pd.isna(data):
            continue
        registros.append({
            "DESCRIÇÃO": "BANCO ITAÚ",
            "DATA": data,
            "VALOR": round(valor, 2),
            "DÉBITO": CONTA_ITAU_GZ if valor > 0 else "",
            "CRÉDITO": CONTA_ITAU_GZ if valor < 0 else "",
            "HISTÓRICO": historico,
        })
    if not registros:
        raise ValueError("Nenhum lançamento foi reconhecido no extrato Itaú da GZ.")
    return pd.DataFrame(registros, columns=COLUNAS_MODELO)


def ler_boletos_liquidados_gz(conteudo: bytes) -> List[BoletoGZ]:
    """Lê o relatório de boletos; usa OCR quando o PDF é imagem."""
    texto = _texto_pdf(conteudo)
    if len(texto.strip()) < 100:
        texto = _ocr_pdf(conteudo)
    texto = texto.replace("|", " ")
    boletos: List[BoletoGZ] = []
    padrao = re.compile(
        r"^\s*(.*?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+"
        r"([\d.]+,\d{2})\s+157\b.*?\b(Baixado|Liquidado)\b",
        re.I,
    )
    for linha in texto.splitlines():
        linha = _normalizar_espacos(linha)
        match = padrao.search(linha)
        if not match:
            continue
        pagador = _normalizar_espacos(match.group(1)).strip(" -")
        vencimento = pd.to_datetime(match.group(2), dayfirst=True, errors="coerce")
        liquidacao = pd.to_datetime(match.group(3), dayfirst=True, errors="coerce")
        valor = _moeda_br(match.group(4))
        status = match.group(5).title()
        if not pagador or pd.isna(liquidacao) or valor <= 0:
            continue
        boletos.append(BoletoGZ(pagador, vencimento, liquidacao, round(valor, 2), status))
    if not boletos:
        raise ValueError("Nenhum boleto foi reconhecido no relatório de boletos liquidados.")
    return boletos


def _subset_exato(indices: list[int], boletos: List[BoletoGZ], alvo: float) -> list[int]:
    alvo_cent = int(round(alvo * 100))
    dp = {0: []}
    for indice in indices:
        valor = int(round(boletos[indice].valor * 100))
        for soma, escolhidos in list(dp.items())[::-1]:
            nova = soma + valor
            if nova > alvo_cent or nova in dp:
                continue
            dp[nova] = escolhidos + [indice]
            if nova == alvo_cent:
                return dp[nova]
    return []


def processar_gz(extrato_bytes: bytes, boletos_bytes: bytes):
    """Substitui agregados de boletos por liquidações individuais e gera diagnóstico."""
    extrato = ler_extrato_itau_gz(extrato_bytes)
    boletos = ler_boletos_liquidados_gz(boletos_bytes)
    data_ini = extrato["DATA"].min().normalize()
    data_fim = extrato["DATA"].max().normalize()

    saida = []
    diagnosticos = []
    for _, linha in extrato.iterrows():
        historico = str(linha["HISTÓRICO"] or "")
        if "BOLETOS RECEBIDOS" not in historico.upper():
            saida.append(linha.to_dict())
            continue

        data_extrato = pd.to_datetime(linha["DATA"]).normalize()
        alvo = abs(float(linha["VALOR"]))
        elegiveis = [
            i for i, boleto in enumerate(boletos)
            if not boleto.usado
            and boleto.status.lower() == "liquidado"
            and data_ini <= boleto.liquidacao.normalize() <= data_extrato
        ]
        soma_elegivel = round(sum(boletos[i].valor for i in elegiveis), 2)
        escolhidos = elegiveis if abs(soma_elegivel - alvo) <= 0.02 else _subset_exato(elegiveis, boletos, alvo)
        soma_escolhida = round(sum(boletos[i].valor for i in escolhidos), 2)
        bate = bool(escolhidos) and abs(soma_escolhida - alvo) <= 0.02

        diagnosticos.append({
            "DATA_EXTRATO": data_extrato,
            "TOTAL_EXTRATO": alvo,
            "TOTAL_BOLETOS": soma_escolhida if escolhidos else soma_elegivel,
            "DIFERENÇA": round((soma_escolhida if escolhidos else soma_elegivel) - alvo, 2),
            "QTD_BOLETOS": len(escolhidos),
            "STATUS": "Batendo" if bate else "Divergente",
        })

        if not bate:
            saida.append(linha.to_dict())
            continue

        for indice in escolhidos:
            boleto = boletos[indice]
            boleto.usado = True
            saida.append({
                "DESCRIÇÃO": "BANCO ITAÚ",
                "DATA": boleto.liquidacao,
                "VALOR": boleto.valor,
                "DÉBITO": CONTA_ITAU_GZ,
                "CRÉDITO": "",
                "HISTÓRICO": f"Recebido: {boleto.pagador}",
            })

    df_saida = pd.DataFrame(saida, columns=COLUNAS_MODELO)
    df_saida["DATA"] = pd.to_datetime(df_saida["DATA"], dayfirst=True, errors="coerce")
    df_saida = df_saida.dropna(subset=["DATA"]).sort_values("DATA", kind="stable").reset_index(drop=True)

    nao_usados = [
        boleto for boleto in boletos
        if not boleto.usado
        and boleto.status.lower() == "liquidado"
        and data_ini <= boleto.liquidacao.normalize() <= data_fim
    ]
    df_nao_usados = pd.DataFrame([
        {"DATA": b.liquidacao, "PAGADOR": b.pagador, "VALOR": b.valor, "STATUS": b.status}
        for b in nao_usados
    ])
    df_diag = pd.DataFrame(diagnosticos)
    resumo = {
        "periodo_inicio": data_ini,
        "periodo_fim": data_fim,
        "agregados": len(diagnosticos),
        "agregados_batendo": int((df_diag["STATUS"] == "Batendo").sum()) if not df_diag.empty else 0,
        "agregados_divergentes": int((df_diag["STATUS"] == "Divergente").sum()) if not df_diag.empty else 0,
        "boletos_nao_usados": len(nao_usados),
        "total_extrato": round(float(extrato["VALOR"].sum()), 2),
        "total_modelo": round(float(df_saida["VALOR"].sum()), 2),
    }
    return df_saida, df_diag, df_nao_usados, resumo


def gerar_modelo_dominio_gz(df: pd.DataFrame, modelo_bytes: bytes) -> bytes:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(modelo_bytes))
    template = None
    cabecalho = None
    mapa = None
    normalizar = lambda v: re.sub(r"[^A-Z0-9]", "", str(v or "").upper().translate(str.maketrans("ÁÀÃÂÉÊÍÓÔÕÚÇ", "AAAAEEIOOOUC")))
    esperadas = [normalizar(c) for c in COLUNAS_MODELO]
    for ws in wb.worksheets:
        for linha in range(1, min(ws.max_row, 25) + 1):
            mapa_temp = {normalizar(ws.cell(linha, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(linha, c).value is not None}
            if all(nome in mapa_temp for nome in esperadas):
                template, cabecalho, mapa = ws, linha, mapa_temp
                break
        if template is not None:
            break
    if template is None:
        raise ValueError("Cabeçalho do Modelo Domínio não localizado.")

    linha_modelo = cabecalho + 1
    estilos = {}
    for c in range(1, template.max_column + 1):
        cel = template.cell(linha_modelo, c)
        estilos[c] = (copy(cel.font), copy(cel.fill), copy(cel.border), copy(cel.alignment), cel.number_format, copy(cel.protection))
    for r in range(cabecalho + 1, template.max_row + 1):
        for c in range(1, template.max_column + 1):
            template.cell(r, c).value = None

    for r, registro in enumerate(df[COLUNAS_MODELO].to_dict("records"), start=linha_modelo):
        for nome in COLUNAS_MODELO:
            c = mapa[normalizar(nome)]
            valor = registro.get(nome, "")
            if pd.isna(valor):
                valor = ""
            if nome == "DATA" and valor not in ("", None):
                valor = pd.to_datetime(valor).to_pydatetime()
            elif nome in {"DÉBITO", "CRÉDITO"} and str(valor).isdigit():
                valor = int(valor)
            elif nome == "VALOR":
                valor = float(valor)
            cel = template.cell(r, c)
            cel.value = valor
            fonte, fill, border, alinhamento, formato, protecao = estilos[c]
            cel.font, cel.fill, cel.border, cel.alignment = copy(fonte), copy(fill), copy(border), copy(alinhamento)
            cel.number_format, cel.protection = formato, copy(protecao)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
