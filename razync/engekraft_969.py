"""Processamento da empresa 969 - Engekraft Automação LTDA - EPP.

Fonte: extrato empresarial Itaú em PDF. Conta Domínio: 508.
"""
from __future__ import annotations

import io
import re
from copy import copy

import pandas as pd
from pypdf import PdfReader

CONTA_ITAU_969 = "508"
COLUNAS_MODELO = ["DESCRIÇÃO", "DATA", "VALOR", "DÉBITO", "CRÉDITO", "HISTÓRICO"]


def _espacos(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _moeda(v):
    s = str(v or "").replace("R$", "").replace(" ", "").strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _texto_pdf(conteudo: bytes) -> str:
    reader = PdfReader(io.BytesIO(conteudo), strict=False)
    return "\n".join((p.extract_text() or "") for p in reader.pages)


def _limpar_historico(texto: str) -> str:
    texto = _espacos(texto)
    texto = re.sub(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", " ", texto)
    texto = re.sub(r"\b\d{3}\.\d{3}\.\d{3}-\d{2}\b", " ", texto)
    prefixos_operacao = [
        r"BOLETO\s+PAGO(?:\s+[A-Z0-9./&-]+){0,3}\s+",
        r"PIX\s+ENVIADO\s+",
        r"PIX\s+RECEBIDO(?:\s+[A-Z0-9/.-]+)?\s+",
        r"PAGAMENTOS?\s+TRANSF\s+CC\s+ITAU\s+",
        r"PAGAMENTOS?\s+TRIB\s+MUNICIPAL\s+",
        r"PAGAMENTOS?\s+TRIB\s+COD\s+BARRAS\s+",
        r"RECEBIMENTOS?\s+",
    ]
    for padrao in prefixos_operacao:
        novo = re.sub(r"^" + padrao, "", texto, flags=re.I)
        if novo != texto:
            texto = novo
            break
    texto = _espacos(texto).strip(" -")
    palavras = texto.split()
    if len(palavras) >= 2:
        metade = len(palavras) // 2
        if len(palavras) % 2 == 0 and palavras[:metade] == palavras[metade:]:
            texto = " ".join(palavras[metade:])
    return texto


def processar_extrato_engekraft_969(conteudo: bytes) -> pd.DataFrame:
    texto = _texto_pdf(conteudo)
    topo = texto.upper()
    if "ENGEKRAFT" not in topo or ("0020269-9" not in texto and "20269" not in texto):
        raise ValueError("O PDF enviado não parece ser o extrato Itaú da empresa 969 - Engekraft.")

    linhas = [x.strip() for x in texto.splitlines() if x.strip()]
    padrao_data = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(.*)$")
    blocos, atual = [], None
    for linha in linhas:
        m = padrao_data.match(linha)
        if m:
            if atual:
                blocos.append(atual)
            atual = {"data": m.group(1), "partes": [m.group(2)]}
        elif atual:
            atual["partes"].append(linha)
    if atual:
        blocos.append(atual)

    regex_moeda = re.compile(r"(?<!\d)([-+]?\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)")
    registros = []
    for bloco in blocos:
        corpo = _espacos(" ".join(bloco["partes"]))
        upper = corpo.upper()
        if upper.startswith("SALDO ") or "SALDO TOTAL DISPONÍVEL" in upper or "SALDO EM CONTA" in upper:
            continue
        moedas = list(regex_moeda.finditer(corpo))
        if not moedas:
            continue
        achado = moedas[-1]
        valor = _moeda(achado.group(1))
        if abs(valor) < 0.005:
            continue
        historico_original = corpo[:achado.start()]
        historico = _limpar_historico(historico_original)
        if not historico:
            continue
        data = pd.to_datetime(bloco["data"], dayfirst=True, errors="coerce")
        if pd.isna(data):
            continue
        if re.search(r"\bRENDIMENTOS?\b", historico_original, flags=re.I):
            historico_final = "Recebido: RENDIMENTOS"
        else:
            historico = re.sub(r"^(?:Pago|Recebido):\s*", "", historico, flags=re.I).strip()
            prefixo = "Recebido: " if valor > 0 else "Pago: "
            historico_final = prefixo + historico
        registros.append({
            "DESCRIÇÃO": "BANCO ITAÚ",
            "DATA": data,
            "VALOR": round(valor, 2),
            "DÉBITO": CONTA_ITAU_969 if valor > 0 else "",
            "CRÉDITO": CONTA_ITAU_969 if valor < 0 else "",
            "HISTÓRICO": historico_final,
        })
    if not registros:
        raise ValueError("Nenhum lançamento foi reconhecido no extrato Itaú da Engekraft.")
    return pd.DataFrame(registros, columns=COLUNAS_MODELO)


def gerar_modelo_dominio_engekraft_969(df: pd.DataFrame, modelo_bytes: bytes) -> bytes:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(modelo_bytes))
    trans = str.maketrans("ÁÀÃÂÉÊÍÓÔÕÚÇ", "AAAAEEIOOOUC")
    norm = lambda v: re.sub(r"[^A-Z0-9]", "", str(v or "").upper().translate(trans))
    esperadas = [norm(c) for c in COLUNAS_MODELO]
    ws_alvo = None
    linha_cab = None
    mapa = None
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 25) + 1):
            m = {norm(ws.cell(r, c).value): c for c in range(1, ws.max_column + 1)}
            if all(k in m for k in esperadas):
                ws_alvo, linha_cab, mapa = ws, r, m
                break
        if ws_alvo:
            break
    if not ws_alvo:
        raise ValueError("Cabeçalho do Modelo Domínio não encontrado.")

    primeira = linha_cab + 1
    for r in range(primeira, ws_alvo.max_row + 1):
        for col in COLUNAS_MODELO:
            ws_alvo.cell(r, mapa[norm(col)]).value = None

    estilo = {col: copy(ws_alvo.cell(primeira, mapa[norm(col)])._style) for col in COLUNAS_MODELO}
    formatos = {col: ws_alvo.cell(primeira, mapa[norm(col)]).number_format for col in COLUNAS_MODELO}
    for i, (_, row) in enumerate(df.iterrows(), start=primeira):
        for col in COLUNAS_MODELO:
            cel = ws_alvo.cell(i, mapa[norm(col)])
            valor = row[col]
            if col == "DATA" and pd.notna(valor):
                valor = pd.to_datetime(valor).to_pydatetime()
            cel.value = valor
            cel._style = copy(estilo[col])
            cel.number_format = formatos[col]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
