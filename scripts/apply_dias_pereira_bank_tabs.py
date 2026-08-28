from pathlib import Path

path = Path('app.py')
text = path.read_text(encoding='utf-8')

old1 = """                    quadros_nibo = []
                    configs_nibo_processados = []
                    with st.spinner('Lendo e organizando os relatórios Nibo...'):
                        for banco_nibo_rotulo in bancos_nibo_selecionados:
                            config_banco_nibo = bancos_dias_pereira[banco_nibo_rotulo]
                            arquivo_nibo = arquivos_nibo_por_banco[banco_nibo_rotulo]
                            df_nibo = processar_extrato_nibo_pdf(arquivo_nibo.getvalue())
                            df_banco_nibo = df_nibo[
                                ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
                            ].copy()
                            df_banco_nibo['DESCRIÇÃO'] = config_banco_nibo['descricao']
                            quadros_nibo.append(df_banco_nibo)
                            configs_nibo_processados.append(config_banco_nibo)
"""

new1 = """                    quadros_nibo = []
                    quadros_nibo_por_slug = {}
                    configs_nibo_processados = []
                    with st.spinner('Lendo e organizando os relatórios Nibo...'):
                        for banco_nibo_rotulo in bancos_nibo_selecionados:
                            config_banco_nibo = bancos_dias_pereira[banco_nibo_rotulo]
                            arquivo_nibo = arquivos_nibo_por_banco[banco_nibo_rotulo]
                            df_nibo = processar_extrato_nibo_pdf(arquivo_nibo.getvalue())
                            df_banco_nibo = df_nibo[
                                ['DESCRIÇÃO', 'DATA', 'VALOR', 'DÉBITO', 'CRÉDITO', 'HISTÓRICO']
                            ].copy()
                            df_banco_nibo['DESCRIÇÃO'] = config_banco_nibo['descricao']
                            quadros_nibo.append(df_banco_nibo)
                            quadros_nibo_por_slug[config_banco_nibo['slug']] = df_banco_nibo
                            configs_nibo_processados.append(config_banco_nibo)
"""

if old1 not in text:
    raise SystemExit('Bloco de leitura Nibo não encontrado')
text = text.replace(old1, new1, 1)

old2 = """                    excel_nibo = gerar_excel_modelo_dominio(df_export_nibo)
                    datas_validas = datas_nibo.dropna()
                    bancos_nome_arquivo = '_'.join(
                        config['arquivo'] for config in configs_nibo_processados
                    )
                    if not datas_validas.empty:
                        nome_nibo = (
                            f\"1529_Dias_Pereira_{bancos_nome_arquivo}_\"
                            f\"{datas_validas.min().strftime('%m_%Y')}.xlsx\"
                        )
                    else:
                        nome_nibo = (
                            f\"1529_Dias_Pereira_{bancos_nome_arquivo}_Modelo_Dominio.xlsx\"
                        )

                    arquivo_saida_nibo = excel_nibo
                    resumo_nibo = {}
                    try:
                        base_dias_pereira = carregar_classificacoes_online('dias_pereira')
                        bancos_slugs_nibo = {
                            config['slug'] for config in configs_nibo_processados
                        }
                        base_bancos_nibo = [
                            item for item in base_dias_pereira
                            if item.get('banco') in bancos_slugs_nibo
                        ]
                        arquivo_saida_nibo, resumo_nibo = classificar_planilha_final(
                            excel_nibo,
                            nome_nibo,
                            base_bancos_nibo,
                            contas_dias_pereira
                        )
                    except Exception as erro_base_nibo:
                        st.info(
                            'O Modelo Domínio foi gerado normalmente, mas a Base Inteligente '
                            f'não pôde ser aplicada agora: {erro_base_nibo}'
                        )
"""

new2 = """                    datas_validas = datas_nibo.dropna()
                    bancos_nome_arquivo = '_'.join(
                        config['arquivo'] for config in configs_nibo_processados
                    )
                    if not datas_validas.empty:
                        nome_nibo = (
                            f\"1529_Dias_Pereira_{bancos_nome_arquivo}_\"
                            f\"{datas_validas.min().strftime('%m_%Y')}.xlsx\"
                        )
                    else:
                        nome_nibo = (
                            f\"1529_Dias_Pereira_{bancos_nome_arquivo}_Modelo_Dominio.xlsx\"
                        )

                    def _bytes_excel_nibo(arquivo_excel):
                        if isinstance(arquivo_excel, (bytes, bytearray)):
                            return bytes(arquivo_excel)
                        if hasattr(arquivo_excel, 'getvalue'):
                            return arquivo_excel.getvalue()
                        return bytes(arquivo_excel)

                    def _combinar_modelos_nibo_em_abas(modelos_por_banco):
                        from copy import copy as copiar_estilo_nibo
                        from openpyxl import Workbook, load_workbook

                        wb_saida = Workbook()
                        wb_saida.remove(wb_saida.active)

                        for nome_aba, arquivo_banco in modelos_por_banco:
                            wb_origem = load_workbook(io.BytesIO(_bytes_excel_nibo(arquivo_banco)))
                            ws_origem = wb_origem.active
                            ws_destino = wb_saida.create_sheet(title=nome_aba[:31])

                            for row in ws_origem.iter_rows():
                                for celula in row:
                                    nova = ws_destino[celula.coordinate]
                                    nova.value = celula.value
                                    if celula.has_style:
                                        nova._style = copiar_estilo_nibo(celula._style)
                                    if celula.number_format:
                                        nova.number_format = celula.number_format
                                    if celula.font:
                                        nova.font = copiar_estilo_nibo(celula.font)
                                    if celula.fill:
                                        nova.fill = copiar_estilo_nibo(celula.fill)
                                    if celula.border:
                                        nova.border = copiar_estilo_nibo(celula.border)
                                    if celula.alignment:
                                        nova.alignment = copiar_estilo_nibo(celula.alignment)
                                    if celula.protection:
                                        nova.protection = copiar_estilo_nibo(celula.protection)

                            for chave, dimensao in ws_origem.column_dimensions.items():
                                ws_destino.column_dimensions[chave].width = dimensao.width
                                ws_destino.column_dimensions[chave].hidden = dimensao.hidden
                                ws_destino.column_dimensions[chave].bestFit = dimensao.bestFit

                            for indice, dimensao in ws_origem.row_dimensions.items():
                                ws_destino.row_dimensions[indice].height = dimensao.height
                                ws_destino.row_dimensions[indice].hidden = dimensao.hidden

                            for intervalo in ws_origem.merged_cells.ranges:
                                ws_destino.merge_cells(str(intervalo))

                            ws_destino.freeze_panes = ws_origem.freeze_panes
                            ws_destino.sheet_format = copiar_estilo_nibo(ws_origem.sheet_format)
                            ws_destino.sheet_properties = copiar_estilo_nibo(ws_origem.sheet_properties)
                            ws_destino.page_margins = copiar_estilo_nibo(ws_origem.page_margins)
                            ws_destino.page_setup = copiar_estilo_nibo(ws_origem.page_setup)
                            ws_destino.print_options = copiar_estilo_nibo(ws_origem.print_options)
                            ws_destino.sheet_view.showGridLines = ws_origem.sheet_view.showGridLines
                            if ws_origem.auto_filter.ref:
                                ws_destino.auto_filter.ref = ws_origem.auto_filter.ref

                        saida = io.BytesIO()
                        wb_saida.save(saida)
                        saida.seek(0)
                        return saida.getvalue()

                    modelos_nibo_por_banco = []
                    resumo_nibo = {'automaticos': 0, 'somente_banco': 0}
                    base_dias_pereira = []
                    erro_base_nibo = ''
                    try:
                        base_dias_pereira = carregar_classificacoes_online('dias_pereira')
                    except Exception as erro_base:
                        erro_base_nibo = str(erro_base)

                    for config_banco_nibo in configs_nibo_processados:
                        df_banco_nibo = quadros_nibo_por_slug[config_banco_nibo['slug']]
                        excel_banco_nibo = gerar_excel_modelo_dominio(df_banco_nibo)
                        arquivo_banco_nibo = excel_banco_nibo
                        resumo_banco_nibo = {}

                        if not erro_base_nibo:
                            try:
                                base_banco_nibo = [
                                    item for item in base_dias_pereira
                                    if item.get('banco') == config_banco_nibo['slug']
                                ]
                                arquivo_banco_nibo, resumo_banco_nibo = classificar_planilha_final(
                                    excel_banco_nibo,
                                    nome_nibo,
                                    base_banco_nibo,
                                    contas_dias_pereira
                                )
                            except Exception as erro_classificacao_banco:
                                st.info(
                                    f\"A aba {config_banco_nibo['nome']} foi gerada normalmente, \"
                                    'mas a Base Inteligente não pôde ser aplicada nela agora: '
                                    f'{erro_classificacao_banco}'
                                )

                        resumo_nibo['automaticos'] += int(
                            resumo_banco_nibo.get('automaticos', 0) or 0
                        )
                        resumo_nibo['somente_banco'] += int(
                            resumo_banco_nibo.get('somente_banco', 0) or 0
                        )
                        modelos_nibo_por_banco.append(
                            (config_banco_nibo['nome'], arquivo_banco_nibo)
                        )

                    if erro_base_nibo:
                        st.info(
                            'O Modelo Domínio foi gerado normalmente em abas separadas por banco, '
                            'mas a Base Inteligente não pôde ser carregada agora: '
                            f'{erro_base_nibo}'
                        )

                    arquivo_saida_nibo = _combinar_modelos_nibo_em_abas(
                        modelos_nibo_por_banco
                    )
"""

if old2 not in text:
    raise SystemExit('Bloco de geração/classificação Nibo não encontrado')
text = text.replace(old2, new2, 1)

old3 = """                    if resumo_nibo:
                        c_auto_1, c_auto_2 = st.columns(2)
"""
new3 = """                    if any(resumo_nibo.values()):
                        c_auto_1, c_auto_2 = st.columns(2)
"""
if old3 not in text:
    raise SystemExit('Bloco de resumo Nibo não encontrado')
text = text.replace(old3, new3, 1)

old4 = """                        'Baixar Modelo Domínio',
"""
new4 = """                        'Baixar Modelo Domínio por banco',
"""
if old4 not in text:
    raise SystemExit('Botão de download Nibo não encontrado')
text = text.replace(old4, new4, 1)

path.write_text(text, encoding='utf-8')
print('Patch aplicado com sucesso')
