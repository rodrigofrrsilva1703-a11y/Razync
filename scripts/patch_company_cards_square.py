from pathlib import Path

path = Path('app.py')
text = path.read_text(encoding='utf-8')

# ============================================================================
# 1) MELHORIAS VISUAIS LEVES (sem bibliotecas extras e sem aumentar o peso)
# ============================================================================
css_anchor = '''        .stTextInput { margin-top: -2px; }
'''
css_extra = '''        /* Hierarquia visual das ferramentas: mais clara sem deixar o app pesado. */
        [data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 8px !important;
            border-bottom: 1px solid var(--hc-border) !important;
            margin-bottom: 12px !important;
        }
        [data-testid="stTabs"] button[role="tab"] {
            min-height: 42px !important;
            padding: 8px 14px !important;
            border-radius: 8px 8px 0 0 !important;
            color: #aab9c6 !important;
            font-weight: 600 !important;
        }
        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #eef8ff !important;
            background: rgba(19, 185, 232, 0.08) !important;
            border-bottom: 2px solid var(--hc-accent) !important;
        }
        [data-testid="stFileUploader"] {
            margin: 10px 0 14px !important;
        }
        [data-testid="stFileUploaderDropzone"] {
            min-height: 92px !important;
            padding: 14px !important;
        }
        [data-testid="stMetric"] {
            background: rgba(17, 24, 32, 0.72) !important;
            border: 1px solid var(--hc-border) !important;
            border-radius: 8px !important;
            padding: 12px 14px !important;
        }
        .hc-review-box {
            margin: 14px 0 12px;
            padding: 14px 16px;
            border: 1px solid rgba(19, 185, 232, 0.28);
            border-left: 4px solid var(--hc-accent);
            border-radius: 8px;
            background: rgba(19, 185, 232, 0.055);
        }
        .hc-review-title {
            color: #eef8ff;
            font-size: 17px;
            font-weight: 700;
            margin-bottom: 4px;
        }
        .hc-review-text {
            color: #b9cad7;
            font-size: 13.5px;
            line-height: 1.5;
        }
        .hc-step-badge {
            display: inline-block;
            padding: 4px 9px;
            margin: 0 6px 6px 0;
            border-radius: 999px;
            background: #0a1722;
            border: 1px solid #1c4057;
            color: #cbe9f6;
            font-size: 11px;
            font-weight: 650;
        }

        .stTextInput { margin-top: -2px; }
'''
if '.hc-review-box {' not in text:
    if text.count(css_anchor) != 1:
        raise SystemExit(f'Âncora CSS encontrada {text.count(css_anchor)} vezes.')
    text = text.replace(css_anchor, css_extra, 1)

# ============================================================================
# 2) MOTOR DA TELA DE REVISÃO INTELIGENTE
# ============================================================================
helper_marker = '''@st.cache_data(show_spinner=False, max_entries=12)
def processar_nova_geracao_banco(file_bytes, nome_aba, conta_esperada, descricao_banco):
'''
helper_code = r'''def extrair_pendencias_revisao_inteligente(file_bytes, contas_bancarias):
    """Lista somente lançamentos que ainda precisam da conta de contrapartida."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    pendencias = []
    contas_bancarias = contas_bancarias or {}

    for ws in wb.worksheets:
        if 'retir' in normalizar_texto(ws.title):
            continue

        linha_cabecalho = None
        mapa_colunas = {}
        for numero_linha in range(1, min(ws.max_row, 30) + 1):
            mapa_teste = {
                normalizar_texto(texto_celula_seguro(ws.cell(numero_linha, coluna).value)).strip(): coluna
                for coluna in range(1, ws.max_column + 1)
            }
            if all(nome in mapa_teste for nome in ['historico', 'debito', 'credito']):
                linha_cabecalho = numero_linha
                mapa_colunas = mapa_teste
                break
        if linha_cabecalho is None:
            continue

        col_hist = mapa_colunas['historico']
        col_debito = mapa_colunas['debito']
        col_credito = mapa_colunas['credito']
        col_valor = mapa_colunas.get('valor')
        col_data = mapa_colunas.get('data')
        col_descricao = mapa_colunas.get('descricao')
        banco_aba = identificar_chave_banco_empresa(ws.title)

        for numero_linha in range(linha_cabecalho + 1, ws.max_row + 1):
            historico = texto_celula_seguro(ws.cell(numero_linha, col_hist).value)
            if not historico:
                continue

            banco = (
                identificar_chave_banco_empresa(ws.cell(numero_linha, col_descricao).value)
                if col_descricao is not None else ''
            ) or banco_aba
            if banco not in contas_bancarias:
                continue

            debito = texto_celula_seguro(ws.cell(numero_linha, col_debito).value)
            credito = texto_celula_seguro(ws.cell(numero_linha, col_credito).value)
            if debito and credito:
                continue

            assinatura = criar_assinatura_classificacao(historico)
            natureza = assinatura.split('|', 1)[0] if assinatura else ''
            valor = 0.0
            if col_valor is not None:
                valor = limpar_valor_monetario(ws.cell(numero_linha, col_valor).value)
            if natureza not in {'pago', 'recebido'}:
                if valor < 0:
                    natureza = 'pago'
                elif valor > 0:
                    natureza = 'recebido'

            if natureza == 'pago':
                coluna_destino = col_debito
                coluna_banco = col_credito
                lado = 'DÉBITO'
            elif natureza == 'recebido':
                coluna_destino = col_credito
                coluna_banco = col_debito
                lado = 'CRÉDITO'
            else:
                continue

            contrapartida_atual = texto_celula_seguro(
                ws.cell(numero_linha, coluna_destino).value
            )
            if contrapartida_atual:
                continue

            data_texto = ''
            if col_data is not None:
                data_raw = ws.cell(numero_linha, col_data).value
                data_parseada = pd.to_datetime(data_raw, dayfirst=True, errors='coerce')
                if not pd.isna(data_parseada):
                    data_texto = data_parseada.strftime('%d/%m/%Y')
                else:
                    data_texto = texto_celula_seguro(data_raw)

            pendencias.append({
                'Banco': nome_banco_por_chave(banco),
                'Data': data_texto,
                'Valor': valor,
                'Histórico': historico,
                'Classificar em': lado,
                'Conta bancária': texto_celula_seguro(contas_bancarias.get(banco, '')),
                'Conta da contrapartida': '',
                '_aba': ws.title,
                '_linha': numero_linha,
                '_col_destino': coluna_destino,
                '_col_banco': coluna_banco,
                '_banco': banco,
                '_col_data': col_data or 0,
            })

    return pd.DataFrame(pendencias)


def aplicar_revisoes_inteligentes(
    file_bytes, revisoes, filename, empresa, contas_bancarias
):
    """Aplica as contas revisadas e gera somente os novos padrões confirmados pelo usuário."""
    from openpyxl import load_workbook

    if revisoes is None or revisoes.empty:
        return file_bytes, 0, []

    wb = load_workbook(io.BytesIO(file_bytes))
    registros = []
    aplicadas = 0

    def valor_conta_excel(conta):
        texto = texto_celula_seguro(conta)
        if texto.isdigit() and (texto == '0' or not texto.startswith('0')):
            return int(texto)
        return texto

    for _, item in revisoes.iterrows():
        conta_contrapartida = texto_celula_seguro(item.get('Conta da contrapartida'))
        if not conta_contrapartida:
            continue

        nome_aba = texto_celula_seguro(item.get('_aba'))
        if nome_aba not in wb.sheetnames:
            continue
        ws = wb[nome_aba]
        numero_linha = int(item.get('_linha'))
        col_destino = int(item.get('_col_destino'))
        col_banco = int(item.get('_col_banco'))
        banco = texto_celula_seguro(item.get('_banco'))
        conta_banco = texto_celula_seguro(contas_bancarias.get(banco, ''))

        if conta_banco and not texto_celula_seguro(ws.cell(numero_linha, col_banco).value):
            ws.cell(numero_linha, col_banco).value = valor_conta_excel(conta_banco)
        ws.cell(numero_linha, col_destino).value = valor_conta_excel(conta_contrapartida)
        aplicadas += 1

        # Lê o par final exatamente como ficou na planilha e aprende somente esta revisão.
        mapa_colunas = {}
        linha_cabecalho = None
        for linha_teste in range(1, min(ws.max_row, 30) + 1):
            mapa_teste = {
                normalizar_texto(texto_celula_seguro(ws.cell(linha_teste, coluna).value)).strip(): coluna
                for coluna in range(1, ws.max_column + 1)
            }
            if all(nome in mapa_teste for nome in ['historico', 'debito', 'credito']):
                mapa_colunas = mapa_teste
                linha_cabecalho = linha_teste
                break
        if linha_cabecalho is None:
            continue

        historico = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['historico']).value)
        debito = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['debito']).value)
        credito = texto_celula_seguro(ws.cell(numero_linha, mapa_colunas['credito']).value)
        assinatura = criar_assinatura_classificacao(historico)
        if not assinatura or not debito or not credito:
            continue

        periodo = normalizar_texto(filename)
        col_data = mapa_colunas.get('data')
        if col_data is not None:
            data_lancamento = pd.to_datetime(
                ws.cell(numero_linha, col_data).value, dayfirst=True, errors='coerce'
            )
            if not pd.isna(data_lancamento):
                periodo = data_lancamento.strftime('%Y-%m')

        identificador = hashlib.sha256(
            f"{empresa}|{banco}|{assinatura}|{debito}|{credito}".encode('utf-8')
        ).hexdigest()
        registros.append({
            'id': identificador,
            'empresa': empresa,
            'banco': banco,
            'assinatura': assinatura,
            'debito': debito,
            'credito': credito,
            'ocorrencias': 1,
            'periodos': [periodo],
            'exemplo_historico': historico[:500]
        })

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue(), aplicadas, registros


def renderizar_revisao_inteligente(
    arquivo_classificado,
    arquivo_original_bytes,
    filename,
    empresa,
    contas_bancarias,
    senha_admin,
    prefixo_chave,
):
    """Tela de revisão dos lançamentos que a Base Inteligente ainda não resolveu."""
    fingerprint = hashlib.sha256(
        arquivo_original_bytes + empresa.encode('utf-8')
    ).hexdigest()
    chave_fp = f'{prefixo_chave}_review_fp'
    chave_bytes = f'{prefixo_chave}_review_bytes'

    if st.session_state.get(chave_fp) != fingerprint:
        st.session_state[chave_fp] = fingerprint
        st.session_state[chave_bytes] = arquivo_classificado

    arquivo_trabalho = st.session_state.get(chave_bytes, arquivo_classificado)
    pendencias = extrair_pendencias_revisao_inteligente(
        arquivo_trabalho, contas_bancarias
    )

    st.markdown(
        """
        <div class="hc-review-box">
            <div class="hc-review-title">Revisão Inteligente</div>
            <div class="hc-review-text">
                O Razync mostra somente os lançamentos cuja contrapartida ainda não foi identificada.
                Preencha a conta correta, aplique a revisão e o novo padrão será aprendido apenas
                pela base desta empresa.
            </div>
        </div>
        <span class="hc-step-badge">1 · Base classifica</span>
        <span class="hc-step-badge">2 · Você revisa pendências</span>
        <span class="hc-step-badge">3 · Razync aprende</span>
        <span class="hc-step-badge">4 · Baixar planilha</span>
        """,
        unsafe_allow_html=True
    )

    if pendencias.empty:
        st.success("Nenhuma pendência de classificação. A planilha está pronta para download.")
    else:
        c_rev1, c_rev2 = st.columns(2)
        c_rev1.metric("Pendências para revisar", len(pendencias))
        c_rev2.metric(
            "Bancos envolvidos",
            int(pendencias['Banco'].nunique()) if 'Banco' in pendencias.columns else 0
        )
        st.caption(
            "Edite apenas a coluna Conta da contrapartida. As demais colunas servem "
            "como referência para você conferir o lançamento."
        )

        revisoes = st.data_editor(
            pendencias,
            use_container_width=True,
            hide_index=True,
            key=f'{prefixo_chave}_editor',
            disabled=[
                'Banco', 'Data', 'Valor', 'Histórico', 'Classificar em', 'Conta bancária'
            ],
            column_config={
                '_aba': None,
                '_linha': None,
                '_col_destino': None,
                '_col_banco': None,
                '_banco': None,
                '_col_data': None,
                'Conta da contrapartida': st.column_config.TextColumn(
                    'Conta da contrapartida',
                    help='Informe somente o número/código da conta contábil correta.',
                    width='medium'
                ),
                'Histórico': st.column_config.TextColumn('Histórico', width='large'),
                'Valor': st.column_config.NumberColumn('Valor', format='R$ %.2f'),
            },
            height=min(430, 92 + (len(pendencias) * 36))
        )

        senha_revisao = st.text_input(
            "Senha administrativa para salvar o aprendizado da revisão",
            type='password',
            key=f'{prefixo_chave}_senha_revisao'
        ) if senha_admin else ''

        preenchidas = int(
            revisoes['Conta da contrapartida'].fillna('').astype(str).str.strip().ne('').sum()
        )
        st.caption(f"{preenchidas} de {len(revisoes)} pendências preenchidas nesta revisão.")

        if st.button(
            "Aplicar revisões e ensinar a Base Inteligente",
            key=f'{prefixo_chave}_aplicar_revisao',
            use_container_width=True,
            disabled=preenchidas == 0
        ):
            if senha_admin and not hmac.compare_digest(str(senha_revisao), str(senha_admin)):
                st.error("Senha administrativa inválida.")
            else:
                try:
                    novo_arquivo, aplicadas, novos_padroes = executar_com_loading(
                        "Aplicando revisões e preparando o aprendizado...",
                        aplicar_revisoes_inteligentes,
                        arquivo_trabalho,
                        revisoes,
                        filename,
                        empresa,
                        contas_bancarias
                    )
                    if novos_padroes:
                        salvar_classificacoes_online(novos_padroes, empresa)
                    st.session_state[chave_bytes] = novo_arquivo
                    st.success(
                        f"{aplicadas} revisões aplicadas. "
                        f"{len(novos_padroes)} novos padrões foram enviados para a base desta empresa."
                    )
                    st.rerun()
                except Exception as erro_revisao:
                    st.error(f"Não foi possível aplicar a revisão: {erro_revisao}")

    nome_saida = os.path.splitext(filename)[0]
    st.download_button(
        "Baixar planilha classificada atual",
        data=arquivo_trabalho,
        file_name=f"{nome_saida}_Classificada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f'{prefixo_chave}_download_atual',
        use_container_width=True
    )

'''
if 'def renderizar_revisao_inteligente(' not in text:
    if helper_marker not in text:
        raise SystemExit('Marcador para inserir motor de revisão não encontrado.')
    text = text.replace(helper_marker, helper_code + helper_marker, 1)

# ============================================================================
# 3) AUTOKRAFT / PROJETOS / ISA: substitui download simples pela revisão
# ============================================================================
old_autokraft_download = '''                    nome_saida = os.path.splitext(planilha_final.name)[0]
                    st.download_button(
                        "Baixar planilha final classificada",
                        data=arquivo_classificado,
                        file_name=f"{nome_saida}_Classificada.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"base_download_classificada_{empresa}",
                        use_container_width=True
                    )
'''
new_autokraft_download = '''                    renderizar_revisao_inteligente(
                        arquivo_classificado,
                        planilha_final.getvalue(),
                        planilha_final.name,
                        empresa,
                        contas_bancarias,
                        senha_admin,
                        f"base_revisao_{empresa}"
                    )
'''
if old_autokraft_download in text:
    text = text.replace(old_autokraft_download, new_autokraft_download, 1)
elif new_autokraft_download not in text:
    raise SystemExit('Bloco de classificação Autokraft não encontrado.')

# ============================================================================
# 4) NOVA GERAÇÃO MATRIZ/FILIAL: mesma Tela de Revisão Inteligente
# ============================================================================
old_nova_download = '''                        nome_base_saida = os.path.splitext(
                            planilha_final_classificacao.name
                        )[0]
                        st.download_button(
                            "Baixar planilha final classificada",
                            data=arquivo_classificado,
                            file_name=f"{nome_base_saida}_Classificada.xlsx",
                            mime=(
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet"
                            ),
                            key=(
                                'org_download_planilha_final_classificada_nova_'
                                + chave_estabelecimento
                            ),
                            use_container_width=True
                        )
'''
new_nova_download = '''                        renderizar_revisao_inteligente(
                            arquivo_classificado,
                            planilha_final_classificacao.getvalue(),
                            planilha_final_classificacao.name,
                            empresa_base_nova,
                            contas_dominio_estabelecimento,
                            senha_admin_classificacao,
                            'org_revisao_nova_' + chave_estabelecimento
                        )
'''
if old_nova_download in text:
    text = text.replace(old_nova_download, new_nova_download, 1)
elif new_nova_download not in text:
    raise SystemExit('Bloco de classificação Nova Geração não encontrado.')

# Validações estáticas.
checks = [
    'def extrair_pendencias_revisao_inteligente(',
    'def aplicar_revisoes_inteligentes(',
    'def renderizar_revisao_inteligente(',
    'Aplicar revisões e ensinar a Base Inteligente',
    'Conta da contrapartida',
    'hc-review-box',
    "f\"base_revisao_{empresa}\"",
    "'org_revisao_nova_' + chave_estabelecimento",
]
for check in checks:
    if check not in text:
        raise SystemExit(f'Validação falhou: {check!r}')

path.write_text(text, encoding='utf-8')
print('Tela de Revisão Inteligente adicionada às 4 empresas e interface refinada.')
