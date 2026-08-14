from pathlib import Path

p = Path('app.py')
s = p.read_text(encoding='utf-8')

marcador = "def gerar_txt_dominio(df):"
if 'def gerar_excel_modelo_dominio(df):' not in s:
    if marcador not in s:
        raise SystemExit('Ponto de inserção do exportador não encontrado.')

    helper = r'''def gerar_excel_modelo_dominio(df):
    """Preenche uma cópia fiel do Modelo Domínio, preservando sua estrutura e estilos."""
    from copy import copy
    from openpyxl import load_workbook

    caminho_modelo = next(
        (caminho for caminho in ['Modelo dominio.xlsx', 'Modelo dominio(6).xlsx']
         if os.path.exists(caminho)),
        None
    )
    if not caminho_modelo:
        raise FileNotFoundError('Modelo Domínio não encontrado no sistema.')

    wb = load_workbook(caminho_modelo)
    ws = wb[wb.sheetnames[0]]

    # Localiza a linha real do cabeçalho sem presumir que seja sempre a primeira.
    cabecalho_linha = None
    mapa_colunas = {}
    for linha in range(1, min(ws.max_row, 25) + 1):
        mapa_temp = {}
        for coluna in range(1, ws.max_column + 1):
            valor = ws.cell(linha, coluna).value
            nome = normalizar_texto(str(valor or '')).strip()
            if nome:
                mapa_temp[nome] = coluna
        if 'data' in mapa_temp and 'valor' in mapa_temp and 'historico' in mapa_temp:
            cabecalho_linha = linha
            mapa_colunas = mapa_temp
            break

    if cabecalho_linha is None:
        raise ValueError('Cabeçalho do Modelo Domínio não foi localizado.')

    nomes_df = {normalizar_texto(str(c)).strip(): c for c in df.columns}
    linha_modelo = cabecalho_linha + 1

    # Guarda o estilo da primeira linha de dados do próprio modelo para replicá-lo.
    estilos = {}
    for coluna in range(1, ws.max_column + 1):
        celula = ws.cell(linha_modelo, coluna)
        estilos[coluna] = {
            'font': copy(celula.font),
            'fill': copy(celula.fill),
            'border': copy(celula.border),
            'alignment': copy(celula.alignment),
            'number_format': celula.number_format,
            'protection': copy(celula.protection),
        }

    # Remove somente conteúdos antigos da área de dados. Cabeçalho, larguras,
    # filtros, congelamentos, impressão e demais propriedades ficam intactos.
    for linha in range(cabecalho_linha + 1, ws.max_row + 1):
        for coluna in range(1, ws.max_column + 1):
            ws.cell(linha, coluna).value = None

    for indice, registro in enumerate(df.to_dict('records'), start=cabecalho_linha + 1):
        for nome_normalizado, coluna_excel in mapa_colunas.items():
            coluna_df = nomes_df.get(nome_normalizado)
            if coluna_df is None:
                continue
            valor = registro.get(coluna_df, '')
            if pd.isna(valor):
                valor = ''
            if nome_normalizado == 'data' and valor not in ('', None):
                data = pd.to_datetime(valor, dayfirst=True, errors='coerce')
                valor = data.to_pydatetime() if not pd.isna(data) else valor

            celula = ws.cell(indice, coluna_excel)
            celula.value = valor
            estilo = estilos.get(coluna_excel)
            if estilo:
                celula.font = copy(estilo['font'])
                celula.fill = copy(estilo['fill'])
                celula.border = copy(estilo['border'])
                celula.alignment = copy(estilo['alignment'])
                celula.number_format = estilo['number_format']
                celula.protection = copy(estilo['protection'])

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()

'''
    s = s.replace(marcador, helper + marcador, 1)

antigo_consolidado = """                            buf_excel_g = io.BytesIO()\n                            with pd.ExcelWriter(buf_excel_g, engine='openpyxl') as writer: df_geral_final.to_excel(writer, index=False)\n                            cc_dl1.download_button(\"Baixar Excel (.XLSX)\", data=buf_excel_g.getvalue(), file_name=f\"consolidado_geral_{data_geral_ini.strftime('%d%m%Y')}.xlsx\", mime=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\", key=\"dl_excel_geral\", use_container_width=True)"""
novo_consolidado = """                            excel_modelo_g = gerar_excel_modelo_dominio(df_geral_final)\n                            cc_dl1.download_button(\"Baixar Excel (.XLSX)\", data=excel_modelo_g, file_name=f\"consolidado_geral_{data_geral_ini.strftime('%d%m%Y')}.xlsx\", mime=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\", key=\"dl_excel_geral\", use_container_width=True)"""
if antigo_consolidado not in s:
    raise SystemExit('Exportação consolidada antiga não encontrada.')
s = s.replace(antigo_consolidado, novo_consolidado, 1)

antigo_individual = """                        buffer_excel = io.BytesIO()\n                        with pd.ExcelWriter(buffer_excel, engine='openpyxl') as writer: df_final.to_excel(writer, index=False)\n                        c_dl1.download_button(\"Baixar Excel (.XLSX)\", data=buffer_excel.getvalue(), file_name=f\"lancamentos_{os.path.splitext(arquivo.name)[0]}_{data_sel_ini.strftime('%d%m%Y')}.xlsx\", mime=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\", key=f\"excel_{idx_arq}\", use_container_width=True)"""
novo_individual = """                        excel_modelo = gerar_excel_modelo_dominio(df_final)\n                        c_dl1.download_button(\"Baixar Excel (.XLSX)\", data=excel_modelo, file_name=f\"lancamentos_{os.path.splitext(arquivo.name)[0]}_{data_sel_ini.strftime('%d%m%Y')}.xlsx\", mime=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\", key=f\"excel_{idx_arq}\", use_container_width=True)"""
if antigo_individual not in s:
    raise SystemExit('Exportação individual antiga não encontrada.')
s = s.replace(antigo_individual, novo_individual, 1)

checks = [
    'def gerar_excel_modelo_dominio(df):',
    'excel_modelo_g = gerar_excel_modelo_dominio(df_geral_final)',
    'excel_modelo = gerar_excel_modelo_dominio(df_final)',
]
for check in checks:
    if check not in s:
        raise SystemExit(f'Validação falhou: {check}')

p.write_text(s, encoding='utf-8')
print('Conversor agora exporta usando cópia fiel do Modelo Domínio.')
